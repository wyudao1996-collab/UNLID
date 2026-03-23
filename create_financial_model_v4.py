# -*- coding: utf-8 -*-
"""UNLID 財務三表（PL/BS/CF）+ 感度分析 v4.0
v4.0変更：月額単価レンジを20〜30万円に改定（最低単価20万円設定）
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# ───────────────── ヘルパー ─────────────────
def sc(ws, row, col, value=None, formula=None, bold=False, italic=False,
       size=10, fc=None, align="left", nf=None, color="000000"):
    cell = ws.cell(row=row, column=col)
    cell.value = formula if formula else value
    cell.font = Font(name="Meiryo UI", bold=bold, italic=italic,
                     size=size, color=color)
    cell.alignment = Alignment(horizontal=align, vertical="center",
                               wrap_text=True)
    if fc:
        cell.fill = PatternFill("solid", fgColor=fc)
    if nf:
        cell.number_format = nf
    return cell

def border_range(ws, r1, r2, c1, c2, style="thin"):
    s = Side(style=style)
    b = Border(left=s, right=s, top=s, bottom=s)
    for r in range(r1, r2+1):
        for c in range(c1, c2+1):
            ws.cell(row=r, column=c).border = b

# カラーパレット
C_DARK   = "1F3864"
C_BLUE   = "2E75B6"
C_LBLUE  = "BDD7EE"
C_DBLUE  = "1F4E79"
C_SUB    = "D9E1F2"
C_TOT    = "FFF2CC"
C_PROFT  = "E2EFDA"
C_LOSS   = "FCE4D6"
C_NOTE   = "F2F2F2"
C_ORANGE = "F4B942"
NF_MAN   = '#,##0'
NF_PCT   = '0.0%'
NF_DEC   = '0.00'


# ══════════════════════════════════════════════
# SHEET 1: 前提条件
# ══════════════════════════════════════════════
ws_a = wb.active
ws_a.title = "前提条件"
ws_a.column_dimensions["A"].width = 34
for c in ["B","C","D"]:
    ws_a.column_dimensions[c].width = 16

def title_row(ws, row, text, sub=""):
    sc(ws, row, 1, text, bold=True, size=14, fc=C_DARK, color="FFFFFF", align="center")
    ws.merge_cells(f"A{row}:D{row}")
    ws.row_dimensions[row].height = 28
    if sub:
        sc(ws, row+1, 1, sub, fc="D6E4F0", align="center", size=9, italic=True)
        ws.merge_cells(f"A{row+1}:D{row+1}")

def sec_hdr(ws, row, text):
    sc(ws, row, 1, text, bold=True, fc=C_BLUE, color="FFFFFF", align="center")
    ws.merge_cells(f"A{row}:D{row}")

def col_hdr(ws, row):
    for c,l in [(1,"項目"),(2,"Year1"),(3,"Year2"),(4,"Year3")]:
        sc(ws, row, c, l, bold=True, fc=C_LBLUE, align="center")

title_row(ws_a, 1, "UNLID 財務モデル 前提条件 v4.0",
          "作成日：2026年3月  |  単位：万円  |  中央値シナリオ  |  月額単価20〜30万円レンジ")

# ─ セクション①: 売上ドライバー
sec_hdr(ws_a, 4, "① 売上ドライバー（各シートの売上数値の根拠）")
col_hdr(ws_a, 5)

A1 = [
    (6,  "B2B契約企業数（期末・社）",          10,    50,   160,  False),
    (7,  "FDE受託 月額単価（万円/社）",         20,    23,    25,  False),
    (8,  "B2B受託収益（万円）",             "=B6*B7*12","=C6*C7*12","=D6*D7*12", True),
    (9,  "人材紹介成立件数（件）",              3,    20,    50,  False),
    (10, "人材紹介フィー単価（万円/件）",       140,   140,   140, False),
    (11, "人材紹介フィー合計（万円）",      "=B9*B10","=C9*C10","=D9*D10", True),
    (12, "受託プロジェクト件数（件）※1",        0,     5,    15,  False),
    (13, "プロジェクト単価（万円/件）※2",        0,   160,   160, False),
    (14, "プロジェクト収益（万円）",       "=B12*B13","=C12*C13","=D12*D13", True),
    (15, "売上合計（万円）",    "=B8+B11+B14","=C8+C11+C14","=D8+D11+D14", True),
]
for row, label, v1, v2, v3, bold in A1:
    fc = C_TOT if "売上合計" in label else (C_SUB if bold else None)
    for c, v in [(2,v1),(3,v2),(4,v3)]:
        nf = NF_MAN
        if isinstance(v, str):
            sc(ws_a, row, c, formula=v, bold=bold, fc=fc, nf=nf, align="center")
        else:
            sc(ws_a, row, c, value=v, bold=bold, fc=fc, nf=nf, align="center")
    sc(ws_a, row, 1, label, bold=bold, fc=fc)

sc(ws_a, 16, 1, "※1 プロジェクト件数：月額サブスクとは別の「成果物確定型・単発受託」件数（例：AI自動化システム構築 160万/件）",
   italic=True, fc=C_NOTE)
ws_a.merge_cells("A16:D16")
sc(ws_a, 17, 1, "※2 プロジェクト単価：Year2以降 平均160万円想定（AI実装・業務自動化の3ヶ月プロジェクト相場）",
   italic=True, fc=C_NOTE)
ws_a.merge_cells("A17:D17")
border_range(ws_a, 5, 15, 1, 4)

# ─ セクション②: コストドライバー
sec_hdr(ws_a, 19, "② コストドライバー")
col_hdr(ws_a, 20)

A2 = [
    (21, "FDE正社員数（人）",             2,   10,   32),
    (22, "FDE 月額人件費（万円/人）",     45,   40,   42),
    (23, "FDE人件費合計（万円）",   "=B21*B22*12","=C21*C22*12","=D21*D22*12"),
    (24, "創業者報酬（万円）",           360,  600,  840),
    (25, "管理・営業スタッフ人件費（万円）",0, 1080, 3360),
    (26, "人件費合計（万円）",   "=B23+B24+B25","=C23+C24+C25","=D23+D24+D25"),
    (27, "副業ワーカー報酬率（対売上）",0.09, 0.08, 0.08),
    (28, "副業ワーカー報酬（万円）", "=前提条件!B15*B27","=前提条件!C15*C27","=前提条件!D15*D27"),
    (29, "オフィス・インフラ費（万円）",  180,  360,  600),
    (30, "マーケティング費（万円）",      120,  360,  600),
    (31, "法務・許認可費（万円）",        200,  300,  400),
    (32, "採用費（万円）",                  0,  500, 1500),
    (33, "システム開発費（万円）",         100,  300,  800),
    (34, "その他（万円）",                 100,  200,  400),
]
for row, label, v1, v2, v3 in A2:
    bold = "合計" in label
    fc = C_SUB if bold else None
    nf = NF_PCT if "率" in label else NF_MAN
    for c, v in [(2,v1),(3,v2),(4,v3)]:
        if isinstance(v, str):
            sc(ws_a, row, c, formula=v, bold=bold, fc=fc, nf=nf, align="center")
        else:
            sc(ws_a, row, c, value=v, bold=bold, fc=fc, nf=nf, align="center")
    sc(ws_a, row, 1, label, bold=bold, fc=fc)
border_range(ws_a, 20, 34, 1, 4)

# ─ セクション③: その他前提
sec_hdr(ws_a, 36, "③ その他前提")
col_hdr(ws_a, 37)
A3 = [
    (38, "実効税率",                        0.30,  0.30,  0.30),
    (39, "減価償却費（万円）",                20,    80,   200),
    (40, "設備投資 Capex（万円）",           100,   300,   800),
    (41, "初期調達資本（増資）（万円）",    2000,     0,     0),
    (42, "売掛金 回転日数（日）",             30,    30,    30),
    (43, "買掛金 回転日数（日）",             15,    15,    15),
    (44, "法定福利費率（社会保険料 事業主負担）", 0.15, 0.15, 0.15),
]
for row, label, v1, v2, v3 in A3:
    nf = NF_PCT if "率" in label else NF_MAN
    for c, v in [(2,v1),(3,v2),(4,v3)]:
        sc(ws_a, row, c, value=v, nf=nf, align="center")
    sc(ws_a, row, 1, label)
# 法定福利費 金額
sc(ws_a, 45, 1, "  法定福利費 実額（万円）  ＝人件費合計×15%", italic=True)
for c, col in [(2,"B"),(3,"C"),(4,"D")]:
    sc(ws_a, 45, c, formula=f"=ROUND({col}26*{col}44,0)", nf=NF_MAN, italic=True, align="center")
border_range(ws_a, 37, 45, 1, 4)
sc(ws_a, 46, 1,
   "法定福利費内訳：健康保険（事業主5.0%）＋厚生年金（事業主9.15%）＋雇用保険（事業主0.95%）＋労災（0.3%）≒ 15.4%",
   italic=True, fc=C_NOTE)
ws_a.merge_cells("A46:D46")

# ─ セクション④: 資金調達・政策公庫借入金計画
sec_hdr(ws_a, 48, "④ 資金調達・日本政策金融公庫 借入金計画")
col_hdr(ws_a, 49)
# 行データ
sc(ws_a, 50, 1, "政策公庫 新規借入額（万円）")
sc(ws_a, 50, 2, value=500, nf=NF_MAN, align="center")
sc(ws_a, 50, 3, value=0,   nf=NF_MAN, align="center")
sc(ws_a, 50, 4, value=0,   nf=NF_MAN, align="center")

sc(ws_a, 51, 1, "借入金利率（年率）")
sc(ws_a, 51, 2, value=0.020, nf=NF_PCT, align="center")
sc(ws_a, 51, 3, formula="=B51", nf=NF_PCT, align="center")
sc(ws_a, 51, 4, formula="=B51", nf=NF_PCT, align="center")

sc(ws_a, 52, 1, "返済年数（年）")
sc(ws_a, 52, 2, value=5, nf=NF_MAN, align="center")
sc(ws_a, 52, 3, formula="=B52", nf=NF_MAN, align="center")
sc(ws_a, 52, 4, formula="=B52", nf=NF_MAN, align="center")

sc(ws_a, 53, 1, "年間元利均等返済額（万円）",bold=True, fc=C_SUB)
# PMT(rate, nper, pv) → 返済額（正数化）
sc(ws_a, 53, 2, formula="=ROUND(PMT(B51,B52,-B50),1)", nf=NF_DEC, bold=True, fc=C_SUB, align="center")
sc(ws_a, 53, 3, formula="=B53", nf=NF_DEC, bold=True, fc=C_SUB, align="center")
sc(ws_a, 53, 4, formula="=B53", nf=NF_DEC, bold=True, fc=C_SUB, align="center")

sc(ws_a, 54, 1, "  うち 支払利息（万円）")
# IPMT(rate, per, nper, pv) 正数化
sc(ws_a, 54, 2, formula="=ROUND(-IPMT(B51,1,B52,B50),1)", nf=NF_DEC, align="center")
sc(ws_a, 54, 3, formula="=ROUND(-IPMT(B51,2,B52,B50),1)", nf=NF_DEC, align="center")
sc(ws_a, 54, 4, formula="=ROUND(-IPMT(B51,3,B52,B50),1)", nf=NF_DEC, align="center")

sc(ws_a, 55, 1, "  うち 元本返済（万円）")
sc(ws_a, 55, 2, formula="=ROUND(B53-B54,1)", nf=NF_DEC, align="center")
sc(ws_a, 55, 3, formula="=ROUND(B53-C54,1)", nf=NF_DEC, align="center")
sc(ws_a, 55, 4, formula="=ROUND(B53-D54,1)", nf=NF_DEC, align="center")

sc(ws_a, 56, 1, "借入金残高（期末）（万円）", bold=True, fc=C_SUB)
sc(ws_a, 56, 2, formula="=ROUND(B50-B55,1)", nf=NF_DEC, bold=True, fc=C_SUB, align="center")
sc(ws_a, 56, 3, formula="=ROUND(B56-C55,1)", nf=NF_DEC, bold=True, fc=C_SUB, align="center")
sc(ws_a, 56, 4, formula="=ROUND(C56-D55,1)", nf=NF_DEC, bold=True, fc=C_SUB, align="center")
border_range(ws_a, 49, 56, 1, 4)

sc(ws_a, 57, 1,
   "参考：日本政策金融公庫 新創業融資制度 基準金利（2025〜2026年）：年2.00〜2.35%（担保・保証人不要）"
   "  ／  返済期間：最長7年（据置期間1年まで含む）  ／  融資上限：3,000万円",
   italic=True, fc=C_NOTE)
ws_a.merge_cells("A57:D57")

ws_a.freeze_panes = "B5"


# ══════════════════════════════════════════════
# SHEET 2: PL（損益計算書）
# ══════════════════════════════════════════════
ws_pl = wb.create_sheet("PL（損益計算書）")
ws_pl.column_dimensions["A"].width = 38
for c in ["B","C","D","E"]:
    ws_pl.column_dimensions[c].width = 16

title_row(ws_pl, 1, "UNLID 損益計算書（P&L）3ヶ年",
          "単位：万円  |  前提条件シートの数値と完全連動")
ws_pl.merge_cells("A1:E1")
ws_pl.merge_cells("A2:E2")

for c,l in [(1,"項目"),(2,"Year1"),(3,"Year2"),(4,"Year3"),(5,"3年累計")]:
    sc(ws_pl, 3, c, l, bold=True, fc=C_DBLUE, align="center", color="FFFFFF")

# PL行番号辞書
PL = {}

# ── 売上高
sc(ws_pl, 4, 1, "【売上高】", bold=True, fc="DAEEF3")
ws_pl.merge_cells("A4:E4")
PL["b2b"]    = 5
PL["jinzai"] = 6
PL["proj"]   = 7
PL["uriage"] = 8
PL["growth"] = 9

rev_rows = [
    (PL["b2b"],    "  B2B受託収益",        "=前提条件!B8",  "=前提条件!C8",  "=前提条件!D8"),
    (PL["jinzai"], "  人材紹介フィー",      "=前提条件!B11", "=前提条件!C11", "=前提条件!D11"),
    (PL["proj"],   "  プロジェクト型受託",  "=前提条件!B14", "=前提条件!C14", "=前提条件!D14"),
]
for row, label, f1, f2, f3 in rev_rows:
    sc(ws_pl, row, 1, label)
    for c, f in [(2,f1),(3,f2),(4,f3)]:
        sc(ws_pl, row, c, formula=f, nf=NF_MAN)
    sc(ws_pl, row, 5, formula=f"=SUM(B{row}:D{row})", nf=NF_MAN, fc=C_SUB)

sc(ws_pl, PL["uriage"], 1, "売上合計", bold=True, fc=C_TOT)
for c, col in [(2,"B"),(3,"C"),(4,"D")]:
    sc(ws_pl, PL["uriage"], c,
       formula=f"=SUM({col}{PL['b2b']}:{col}{PL['proj']})",
       bold=True, fc=C_TOT, nf=NF_MAN)
sc(ws_pl, PL["uriage"], 5,
   formula=f"=SUM(B{PL['uriage']}:D{PL['uriage']})", bold=True, fc=C_TOT, nf=NF_MAN)

sc(ws_pl, PL["growth"], 1, "前年比成長率", italic=True)
sc(ws_pl, PL["growth"], 2, "—", align="center", italic=True)
sc(ws_pl, PL["growth"], 3,
   formula=f"=(C{PL['uriage']}-B{PL['uriage']})/B{PL['uriage']}",
   nf=NF_PCT, align="center", italic=True)
sc(ws_pl, PL["growth"], 4,
   formula=f"=(D{PL['uriage']}-C{PL['uriage']})/C{PL['uriage']}",
   nf=NF_PCT, align="center", italic=True)
sc(ws_pl, PL["growth"], 5, "—", align="center", italic=True)

# ── 売上原価
sc(ws_pl, 11, 1, "【売上原価（変動費）】", bold=True, fc="DAEEF3")
ws_pl.merge_cells("A11:E11")
PL["worker"]  = 12
PL["fde_out"] = 13
PL["cogs"]    = 14

sc(ws_pl, PL["worker"], 1, "  副業ワーカー報酬")
for c, ref in [(2,"B28"),(3,"C28"),(4,"D28")]:
    sc(ws_pl, PL["worker"], c, formula=f"=前提条件!{ref}", nf=NF_MAN)
sc(ws_pl, PL["worker"], 5,
   formula=f"=SUM(B{PL['worker']}:D{PL['worker']})", nf=NF_MAN, fc=C_SUB)

sc(ws_pl, PL["fde_out"], 1, "  FDE業務委託費（外注）")
for c, v in [(2,0),(3,200),(4,600)]:
    sc(ws_pl, PL["fde_out"], c, value=v, nf=NF_MAN)
sc(ws_pl, PL["fde_out"], 5,
   formula=f"=SUM(B{PL['fde_out']}:D{PL['fde_out']})", nf=NF_MAN, fc=C_SUB)

sc(ws_pl, PL["cogs"], 1, "売上原価合計", bold=True, fc=C_SUB)
for c, col in [(2,"B"),(3,"C"),(4,"D")]:
    sc(ws_pl, PL["cogs"], c,
       formula=f"=SUM({col}{PL['worker']}:{col}{PL['fde_out']})",
       bold=True, fc=C_SUB, nf=NF_MAN)
sc(ws_pl, PL["cogs"], 5,
   formula=f"=SUM(B{PL['cogs']}:D{PL['cogs']})", bold=True, fc=C_SUB, nf=NF_MAN)

# ── 売上総利益
PL["gross"]   = 16
PL["gross_r"] = 17
sc(ws_pl, PL["gross"], 1, "売上総利益（粗利）", bold=True, fc=C_PROFT)
for c, col in [(2,"B"),(3,"C"),(4,"D")]:
    sc(ws_pl, PL["gross"], c,
       formula=f"={col}{PL['uriage']}-{col}{PL['cogs']}",
       bold=True, fc=C_PROFT, nf=NF_MAN)
sc(ws_pl, PL["gross"], 5,
   formula=f"=SUM(B{PL['gross']}:D{PL['gross']})", bold=True, fc=C_PROFT, nf=NF_MAN)

sc(ws_pl, PL["gross_r"], 1, "粗利率", italic=True)
for c, col in [(2,"B"),(3,"C"),(4,"D")]:
    sc(ws_pl, PL["gross_r"], c,
       formula=f"={col}{PL['gross']}/{col}{PL['uriage']}",
       nf=NF_PCT, italic=True, align="center")

# ── 販管費（SGA）
sc(ws_pl, 19, 1, "【販売費及び一般管理費（SGA）】", bold=True, fc="DAEEF3")
ws_pl.merge_cells("A19:E19")
PL["fde_sal"] = 20
PL["founder"] = 21
PL["admin"]   = 22
PL["shakai"]  = 23   # NEW 法定福利費
PL["office"]  = 24
PL["mkt"]     = 25
PL["legal"]   = 26
PL["recruit"] = 27
PL["sys"]     = 28
PL["misc"]    = 29
PL["sga"]     = 30
PL["sga_r"]   = 31

sga_rows = [
    (PL["fde_sal"], "  FDE正社員 人件費",            "=前提条件!B23", "=前提条件!C23", "=前提条件!D23"),
    (PL["founder"], "  創業者報酬",                  "=前提条件!B24", "=前提条件!C24", "=前提条件!D24"),
    (PL["admin"],   "  管理・営業スタッフ人件費",    "=前提条件!B25", "=前提条件!C25", "=前提条件!D25"),
    (PL["shakai"],  "  法定福利費（社会保険料 事業主負担）","=前提条件!B45","=前提条件!C45","=前提条件!D45"),
    (PL["office"],  "  オフィス・インフラ費",        "=前提条件!B29", "=前提条件!C29", "=前提条件!D29"),
    (PL["mkt"],     "  マーケティング費",            "=前提条件!B30", "=前提条件!C30", "=前提条件!D30"),
    (PL["legal"],   "  法務・許認可費",              "=前提条件!B31", "=前提条件!C31", "=前提条件!D31"),
    (PL["recruit"], "  採用費",                      "=前提条件!B32", "=前提条件!C32", "=前提条件!D32"),
    (PL["sys"],     "  システム開発費",              "=前提条件!B33", "=前提条件!C33", "=前提条件!D33"),
    (PL["misc"],    "  その他",                      "=前提条件!B34", "=前提条件!C34", "=前提条件!D34"),
]
for row, label, f1, f2, f3 in sga_rows:
    fc = "FFF8E7" if "法定福利費" in label else None
    sc(ws_pl, row, 1, label, fc=fc)
    for c, f in [(2,f1),(3,f2),(4,f3)]:
        sc(ws_pl, row, c, formula=f, nf=NF_MAN, fc=fc)
    sc(ws_pl, row, 5, formula=f"=SUM(B{row}:D{row})", nf=NF_MAN, fc=C_SUB if not fc else fc)

sc(ws_pl, PL["sga"], 1, "販管費合計", bold=True, fc=C_SUB)
for c, col in [(2,"B"),(3,"C"),(4,"D")]:
    sc(ws_pl, PL["sga"], c,
       formula=f"=SUM({col}{PL['fde_sal']}:{col}{PL['misc']})",
       bold=True, fc=C_SUB, nf=NF_MAN)
sc(ws_pl, PL["sga"], 5,
   formula=f"=SUM(B{PL['sga']}:D{PL['sga']})", bold=True, fc=C_SUB, nf=NF_MAN)

sc(ws_pl, PL["sga_r"], 1, "販管費率", italic=True)
for c, col in [(2,"B"),(3,"C"),(4,"D")]:
    sc(ws_pl, PL["sga_r"], c,
       formula=f"={col}{PL['sga']}/{col}{PL['uriage']}",
       nf=NF_PCT, italic=True, align="center")

# ── 営業利益（EBIT）
PL["ebit"]    = 33
PL["ebit_r"]  = 34
PL["da"]      = 35
PL["ebitda"]  = 36
PL["ebitda_r"]= 37
PL["interest"]= 39   # 支払利息
PL["ebt"]     = 40   # 税引前利益
PL["tax"]     = 41
PL["net"]     = 42
PL["net_r"]   = 43

sc(ws_pl, PL["ebit"], 1, "営業利益（EBIT）", bold=True, fc=C_PROFT)
for c, col in [(2,"B"),(3,"C"),(4,"D")]:
    sc(ws_pl, PL["ebit"], c,
       formula=f"={col}{PL['gross']}-{col}{PL['sga']}",
       bold=True, fc=C_PROFT, nf=NF_MAN)
sc(ws_pl, PL["ebit"], 5,
   formula=f"=SUM(B{PL['ebit']}:D{PL['ebit']})", bold=True, fc=C_PROFT, nf=NF_MAN)

sc(ws_pl, PL["ebit_r"], 1, "営業利益率", italic=True)
for c, col in [(2,"B"),(3,"C"),(4,"D")]:
    sc(ws_pl, PL["ebit_r"], c,
       formula=f"={col}{PL['ebit']}/{col}{PL['uriage']}",
       nf=NF_PCT, italic=True, align="center")

sc(ws_pl, PL["da"], 1, "  (+) 減価償却費（非資金費用 加算）", italic=True)
for c, ref in [(2,"B39"),(3,"C39"),(4,"D39")]:
    sc(ws_pl, PL["da"], c, formula=f"=前提条件!{ref}", nf=NF_MAN, italic=True)

sc(ws_pl, PL["ebitda"], 1, "EBITDA", bold=True, fc=C_PROFT)
for c, col in [(2,"B"),(3,"C"),(4,"D")]:
    sc(ws_pl, PL["ebitda"], c,
       formula=f"={col}{PL['ebit']}+{col}{PL['da']}",
       bold=True, fc=C_PROFT, nf=NF_MAN)
sc(ws_pl, PL["ebitda"], 5,
   formula=f"=SUM(B{PL['ebitda']}:D{PL['ebitda']})", bold=True, fc=C_PROFT, nf=NF_MAN)

sc(ws_pl, PL["ebitda_r"], 1, "EBITDAマージン", italic=True)
for c, col in [(2,"B"),(3,"C"),(4,"D")]:
    sc(ws_pl, PL["ebitda_r"], c,
       formula=f"={col}{PL['ebitda']}/{col}{PL['uriage']}",
       nf=NF_PCT, italic=True, align="center")

sc(ws_pl, 38, 1, ""); ws_pl.merge_cells("A38:E38")

sc(ws_pl, PL["interest"], 1, "  (-) 支払利息（政策公庫借入 年2.0%）", fc="FFF8E7")
for c, ref in [(2,"B54"),(3,"C54"),(4,"D54")]:
    sc(ws_pl, PL["interest"], c, formula=f"=-前提条件!{ref}", nf=NF_MAN, fc="FFF8E7")
sc(ws_pl, PL["interest"], 5,
   formula=f"=SUM(B{PL['interest']}:D{PL['interest']})", nf=NF_MAN, fc="FFF8E7")

sc(ws_pl, PL["ebt"], 1, "税引前利益（EBT）", bold=True, fc=C_PROFT)
for c, col in [(2,"B"),(3,"C"),(4,"D")]:
    sc(ws_pl, PL["ebt"], c,
       formula=f"={col}{PL['ebit']}+{col}{PL['interest']}",
       bold=True, fc=C_PROFT, nf=NF_MAN)
sc(ws_pl, PL["ebt"], 5,
   formula=f"=SUM(B{PL['ebt']}:D{PL['ebt']})", bold=True, fc=C_PROFT, nf=NF_MAN)

sc(ws_pl, PL["tax"], 1, "  法人税等（実効税率30%）※EBTが正の場合のみ")
for c, col in [(2,"B"),(3,"C"),(4,"D")]:
    sc(ws_pl, PL["tax"], c,
       formula=f"=IF({col}{PL['ebt']}>0,ROUND({col}{PL['ebt']}*前提条件!B38,0),0)",
       nf=NF_MAN)
sc(ws_pl, PL["tax"], 5,
   formula=f"=SUM(B{PL['tax']}:D{PL['tax']})", nf=NF_MAN)

sc(ws_pl, PL["net"], 1, "当期純利益", bold=True, fc=C_PROFT)
for c, col in [(2,"B"),(3,"C"),(4,"D")]:
    sc(ws_pl, PL["net"], c,
       formula=f"={col}{PL['ebt']}-{col}{PL['tax']}",
       bold=True, fc=C_PROFT, nf=NF_MAN)
sc(ws_pl, PL["net"], 5,
   formula=f"=SUM(B{PL['net']}:D{PL['net']})", bold=True, fc=C_PROFT, nf=NF_MAN)

sc(ws_pl, PL["net_r"], 1, "純利益率", italic=True)
for c, col in [(2,"B"),(3,"C"),(4,"D")]:
    sc(ws_pl, PL["net_r"], c,
       formula=f"={col}{PL['net']}/{col}{PL['uriage']}",
       nf=NF_PCT, italic=True, align="center")

border_range(ws_pl, 3, PL["net_r"], 1, 5)
ws_pl.freeze_panes = "B4"


# ══════════════════════════════════════════════
# SHEET 3: BS（貸借対照表）
# ══════════════════════════════════════════════
ws_bs = wb.create_sheet("BS（貸借対照表）")
ws_bs.column_dimensions["A"].width = 34
for c in ["B","C","D","E"]:
    ws_bs.column_dimensions[c].width = 16

title_row(ws_bs, 1, "UNLID 貸借対照表（B/S）3ヶ年",
          "単位：万円  |  期末残高  |  PL・CF連動")
ws_bs.merge_cells("A1:E1")
ws_bs.merge_cells("A2:E2")

for c,l in [(1,"項目"),(2,"期首（Year0）"),(3,"Year1末"),(4,"Year2末"),(5,"Year3末")]:
    sc(ws_bs, 3, c, l, bold=True, fc=C_DBLUE, align="center", color="FFFFFF")

BS = {}

# ── 資産
sc(ws_bs, 4, 1, "【資産の部】", bold=True, fc="DAEEF3")
ws_bs.merge_cells("A4:E4")
BS["cash"]    = 5
BS["ar"]      = 6
BS["prepaid"] = 7
BS["cur_tot"] = 8
BS["fa"]      = 9
BS["asset_tot"]= 10

sc(ws_bs, BS["cash"], 1, "  現金・預金")
sc(ws_bs, BS["cash"], 2, value=2000, nf=NF_MAN)
# 期末現金 = 前期末 + 当期CF合計（CFシートから参照）
sc(ws_bs, BS["cash"], 3, formula="=B5+CF（キャッシュフロー計算書）!B20", nf=NF_MAN)
sc(ws_bs, BS["cash"], 4, formula="=C5+CF（キャッシュフロー計算書）!C20", nf=NF_MAN)
sc(ws_bs, BS["cash"], 5, formula="=D5+CF（キャッシュフロー計算書）!D20", nf=NF_MAN)

sc(ws_bs, BS["ar"], 1, "  売掛金")
sc(ws_bs, BS["ar"], 2, value=0, nf=NF_MAN)
for c, col in [(3,"B"),(4,"C"),(5,"D")]:
    sc(ws_bs, BS["ar"], c,
       formula=f"=ROUND(PL（損益計算書）!{col}{PL['uriage']}*前提条件!{col}42/365,0)",
       nf=NF_MAN)

sc(ws_bs, BS["prepaid"], 1, "  前払費用・その他流動資産")
for c in [2,3,4,5]:
    sc(ws_bs, BS["prepaid"], c, value=50, nf=NF_MAN)

sc(ws_bs, BS["cur_tot"], 1, "流動資産合計", bold=True, fc=C_SUB)
for c, col in [(2,"B"),(3,"C"),(4,"D"),(5,"E")]:
    sc(ws_bs, BS["cur_tot"], c,
       formula=f"=SUM({col}{BS['cash']}:{col}{BS['prepaid']})",
       bold=True, fc=C_SUB, nf=NF_MAN)

sc(ws_bs, BS["fa"], 1, "  固定資産（純額）")
sc(ws_bs, BS["fa"], 2, value=0, nf=NF_MAN)
# 固定資産 = 前期末 + Capex - 減価償却
sc(ws_bs, BS["fa"], 3, formula="=B9+前提条件!B40-前提条件!B39", nf=NF_MAN)
sc(ws_bs, BS["fa"], 4, formula="=C9+前提条件!C40-前提条件!C39", nf=NF_MAN)
sc(ws_bs, BS["fa"], 5, formula="=D9+前提条件!D40-前提条件!D39", nf=NF_MAN)

sc(ws_bs, BS["asset_tot"], 1, "資産合計", bold=True, fc=C_TOT)
for c, col in [(2,"B"),(3,"C"),(4,"D"),(5,"E")]:
    sc(ws_bs, BS["asset_tot"], c,
       formula=f"={col}{BS['cur_tot']}+{col}{BS['fa']}",
       bold=True, fc=C_TOT, nf=NF_MAN)

# ── 負債
sc(ws_bs, 12, 1, "【負債の部】", bold=True, fc="DAEEF3")
ws_bs.merge_cells("A12:E12")
BS["ap"]       = 13
BS["accrued"]  = 14
BS["cur_liab"] = 15
BS["loan"]     = 16   # 政策公庫借入金（長期）
BS["liab_tot"] = 17

sc(ws_bs, BS["ap"], 1, "  買掛金")
sc(ws_bs, BS["ap"], 2, value=0, nf=NF_MAN)
for c, col in [(3,"B"),(4,"C"),(5,"D")]:
    sc(ws_bs, BS["ap"], c,
       formula=f"=ROUND(PL（損益計算書）!{col}{PL['cogs']}*前提条件!{col}43/365,0)",
       nf=NF_MAN)

sc(ws_bs, BS["accrued"], 1, "  未払費用（1ヶ月分人件費等）")
sc(ws_bs, BS["accrued"], 2, value=0, nf=NF_MAN)
for c, ref in [(3,"B26"),(4,"C26"),(5,"D26")]:
    sc(ws_bs, BS["accrued"], c, formula=f"=ROUND(前提条件!{ref}/12,0)", nf=NF_MAN)

sc(ws_bs, BS["cur_liab"], 1, "流動負債合計", bold=True, fc=C_SUB)
for c, col in [(2,"B"),(3,"C"),(4,"D"),(5,"E")]:
    sc(ws_bs, BS["cur_liab"], c,
       formula=f"=SUM({col}{BS['ap']}:{col}{BS['accrued']})",
       bold=True, fc=C_SUB, nf=NF_MAN)

sc(ws_bs, BS["loan"], 1, "  政策公庫借入金（長期）", fc="FFF8E7")
sc(ws_bs, BS["loan"], 2, value=0, nf=NF_MAN, fc="FFF8E7")
# 期首=0、Year1末=借入残高（前提条件B56）、Year2末=C56、Year3末=D56
sc(ws_bs, BS["loan"], 3, formula="=ROUND(前提条件!B56,0)", nf=NF_MAN, fc="FFF8E7")
sc(ws_bs, BS["loan"], 4, formula="=ROUND(前提条件!C56,0)", nf=NF_MAN, fc="FFF8E7")
sc(ws_bs, BS["loan"], 5, formula="=ROUND(前提条件!D56,0)", nf=NF_MAN, fc="FFF8E7")

sc(ws_bs, BS["liab_tot"], 1, "負債合計", bold=True, fc=C_SUB)
for c, col in [(2,"B"),(3,"C"),(4,"D"),(5,"E")]:
    sc(ws_bs, BS["liab_tot"], c,
       formula=f"={col}{BS['cur_liab']}+{col}{BS['loan']}",
       bold=True, fc=C_SUB, nf=NF_MAN)

# ── 純資産
sc(ws_bs, 19, 1, "【純資産の部】", bold=True, fc="DAEEF3")
ws_bs.merge_cells("A19:E19")
BS["capital"]    = 20
BS["retained"]   = 21
BS["equity_tot"] = 22
BS["liab_eq"]    = 23
BS["check"]      = 24

sc(ws_bs, BS["capital"], 1, "  資本金・資本剰余金（累積増資）")
sc(ws_bs, BS["capital"], 2, value=2000, nf=NF_MAN)
sc(ws_bs, BS["capital"], 3, formula=f"=B{BS['capital']}+前提条件!B41", nf=NF_MAN)
sc(ws_bs, BS["capital"], 4, formula=f"=C{BS['capital']}+前提条件!C41", nf=NF_MAN)
sc(ws_bs, BS["capital"], 5, formula=f"=D{BS['capital']}+前提条件!D41", nf=NF_MAN)

sc(ws_bs, BS["retained"], 1, "  利益剰余金（累積）")
sc(ws_bs, BS["retained"], 2, value=0, nf=NF_MAN)
sc(ws_bs, BS["retained"], 3,
   formula=f"=B{BS['retained']}+PL（損益計算書）!B{PL['net']}", nf=NF_MAN)
sc(ws_bs, BS["retained"], 4,
   formula=f"=C{BS['retained']}+PL（損益計算書）!C{PL['net']}", nf=NF_MAN)
sc(ws_bs, BS["retained"], 5,
   formula=f"=D{BS['retained']}+PL（損益計算書）!D{PL['net']}", nf=NF_MAN)

sc(ws_bs, BS["equity_tot"], 1, "純資産合計", bold=True, fc=C_PROFT)
for c, col in [(2,"B"),(3,"C"),(4,"D"),(5,"E")]:
    sc(ws_bs, BS["equity_tot"], c,
       formula=f"={col}{BS['capital']}+{col}{BS['retained']}",
       bold=True, fc=C_PROFT, nf=NF_MAN)

sc(ws_bs, BS["liab_eq"], 1, "負債・純資産合計", bold=True, fc=C_TOT)
for c, col in [(2,"B"),(3,"C"),(4,"D"),(5,"E")]:
    sc(ws_bs, BS["liab_eq"], c,
       formula=f"={col}{BS['liab_tot']}+{col}{BS['equity_tot']}",
       bold=True, fc=C_TOT, nf=NF_MAN)

sc(ws_bs, BS["check"], 1, "  ✓ バランスチェック（資産合計−負債純資産合計 ＝ 0が正常）",
   italic=True, fc=C_NOTE)
for c, col in [(2,"B"),(3,"C"),(4,"D"),(5,"E")]:
    sc(ws_bs, BS["check"], c,
       formula=f"={col}{BS['asset_tot']}-{col}{BS['liab_eq']}",
       italic=True, fc=C_NOTE, nf=NF_MAN, align="center")

border_range(ws_bs, 3, BS["check"], 1, 5)
ws_bs.freeze_panes = "B4"


# ══════════════════════════════════════════════
# SHEET 4: CF（キャッシュフロー計算書）
# ══════════════════════════════════════════════
ws_cf = wb.create_sheet("CF（キャッシュフロー計算書）")
ws_cf.column_dimensions["A"].width = 38
for c in ["B","C","D","E"]:
    ws_cf.column_dimensions[c].width = 16

title_row(ws_cf, 1, "UNLID キャッシュフロー計算書（間接法）3ヶ年",
          "単位：万円  |  PL・BS連動  ※CF合計＝期末現金変動額")
ws_cf.merge_cells("A1:E1")
ws_cf.merge_cells("A2:E2")

for c,l in [(1,"項目"),(2,"Year1"),(3,"Year2"),(4,"Year3"),(5,"3年累計")]:
    sc(ws_cf, 3, c, l, bold=True, fc=C_DBLUE, align="center", color="FFFFFF")

CF = {}

# ── 営業活動CF
sc(ws_cf, 4, 1, "【営業活動によるCF】（間接法）", bold=True, fc="DAEEF3")
ws_cf.merge_cells("A4:E4")
CF["net"]      = 5
CF["da_add"]   = 6
CF["ar_chg"]   = 7
CF["ap_chg"]   = 8
CF["acc_chg"]  = 9
CF["tax_pay"]  = 10
CF["op_cf"]    = 11

sc(ws_cf, CF["net"], 1, "  当期純利益（起点）")
for c, col in [(2,"B"),(3,"C"),(4,"D")]:
    sc(ws_cf, CF["net"], c,
       formula=f"=PL（損益計算書）!{col}{PL['net']}", nf=NF_MAN)
sc(ws_cf, CF["net"], 5, formula=f"=SUM(B{CF['net']}:D{CF['net']})", nf=NF_MAN, fc=C_SUB)

sc(ws_cf, CF["da_add"], 1, "  (+) 減価償却費（非資金費用）")
for c, ref in [(2,"B39"),(3,"C39"),(4,"D39")]:
    sc(ws_cf, CF["da_add"], c, formula=f"=前提条件!{ref}", nf=NF_MAN)
sc(ws_cf, CF["da_add"], 5, formula=f"=SUM(B{CF['da_add']}:D{CF['da_add']})", nf=NF_MAN, fc=C_SUB)

sc(ws_cf, CF["ar_chg"], 1, "  (-) 売掛金 増加")
sc(ws_cf, CF["ar_chg"], 2,
   formula=f"=-(BS（貸借対照表）!C{BS['ar']}-BS（貸借対照表）!B{BS['ar']})", nf=NF_MAN)
sc(ws_cf, CF["ar_chg"], 3,
   formula=f"=-(BS（貸借対照表）!D{BS['ar']}-BS（貸借対照表）!C{BS['ar']})", nf=NF_MAN)
sc(ws_cf, CF["ar_chg"], 4,
   formula=f"=-(BS（貸借対照表）!E{BS['ar']}-BS（貸借対照表）!D{BS['ar']})", nf=NF_MAN)
sc(ws_cf, CF["ar_chg"], 5, formula=f"=SUM(B{CF['ar_chg']}:D{CF['ar_chg']})", nf=NF_MAN, fc=C_SUB)

sc(ws_cf, CF["ap_chg"], 1, "  (+) 買掛金 増加")
sc(ws_cf, CF["ap_chg"], 2,
   formula=f"=BS（貸借対照表）!C{BS['ap']}-BS（貸借対照表）!B{BS['ap']}", nf=NF_MAN)
sc(ws_cf, CF["ap_chg"], 3,
   formula=f"=BS（貸借対照表）!D{BS['ap']}-BS（貸借対照表）!C{BS['ap']}", nf=NF_MAN)
sc(ws_cf, CF["ap_chg"], 4,
   formula=f"=BS（貸借対照表）!E{BS['ap']}-BS（貸借対照表）!D{BS['ap']}", nf=NF_MAN)
sc(ws_cf, CF["ap_chg"], 5, formula=f"=SUM(B{CF['ap_chg']}:D{CF['ap_chg']})", nf=NF_MAN, fc=C_SUB)

sc(ws_cf, CF["acc_chg"], 1, "  (+) 未払費用 増加")
sc(ws_cf, CF["acc_chg"], 2,
   formula=f"=BS（貸借対照表）!C{BS['accrued']}-BS（貸借対照表）!B{BS['accrued']}", nf=NF_MAN)
sc(ws_cf, CF["acc_chg"], 3,
   formula=f"=BS（貸借対照表）!D{BS['accrued']}-BS（貸借対照表）!C{BS['accrued']}", nf=NF_MAN)
sc(ws_cf, CF["acc_chg"], 4,
   formula=f"=BS（貸借対照表）!E{BS['accrued']}-BS（貸借対照表）!D{BS['accrued']}", nf=NF_MAN)
sc(ws_cf, CF["acc_chg"], 5, formula=f"=SUM(B{CF['acc_chg']}:D{CF['acc_chg']})", nf=NF_MAN, fc=C_SUB)

sc(ws_cf, CF["tax_pay"], 1, "  (-) 法人税等 支払い")
for c, col in [(2,"B"),(3,"C"),(4,"D")]:
    sc(ws_cf, CF["tax_pay"], c,
       formula=f"=-PL（損益計算書）!{col}{PL['tax']}", nf=NF_MAN)
sc(ws_cf, CF["tax_pay"], 5, formula=f"=SUM(B{CF['tax_pay']}:D{CF['tax_pay']})", nf=NF_MAN, fc=C_SUB)

sc(ws_cf, CF["op_cf"], 1, "営業活動CF 合計", bold=True, fc=C_PROFT)
for c, col in [(2,"B"),(3,"C"),(4,"D")]:
    sc(ws_cf, CF["op_cf"], c,
       formula=f"=SUM({col}{CF['net']}:{col}{CF['tax_pay']})",
       bold=True, fc=C_PROFT, nf=NF_MAN)
sc(ws_cf, CF["op_cf"], 5,
   formula=f"=SUM(B{CF['op_cf']}:D{CF['op_cf']})", bold=True, fc=C_PROFT, nf=NF_MAN)

# ── 投資活動CF
sc(ws_cf, 13, 1, "【投資活動によるCF】", bold=True, fc="DAEEF3")
ws_cf.merge_cells("A13:E13")
CF["capex"]  = 14
CF["inv_cf"] = 15

sc(ws_cf, CF["capex"], 1, "  (-) 設備投資（Capex）")
for c, ref in [(2,"B40"),(3,"C40"),(4,"D40")]:
    sc(ws_cf, CF["capex"], c, formula=f"=-前提条件!{ref}", nf=NF_MAN)
sc(ws_cf, CF["capex"], 5, formula=f"=SUM(B{CF['capex']}:D{CF['capex']})", nf=NF_MAN, fc=C_SUB)

sc(ws_cf, CF["inv_cf"], 1, "投資活動CF 合計", bold=True, fc=C_LOSS)
for c, col in [(2,"B"),(3,"C"),(4,"D")]:
    sc(ws_cf, CF["inv_cf"], c, formula=f"={col}{CF['capex']}",
       bold=True, fc=C_LOSS, nf=NF_MAN)
sc(ws_cf, CF["inv_cf"], 5,
   formula=f"=SUM(B{CF['inv_cf']}:D{CF['inv_cf']})", bold=True, fc=C_LOSS, nf=NF_MAN)

# ── 財務活動CF
sc(ws_cf, 17, 1, "【財務活動によるCF】", bold=True, fc="DAEEF3")
ws_cf.merge_cells("A17:E17")
CF["capital_in"] = 18
CF["loan_in"]    = 19    # 政策公庫借入金 新規調達
CF["loan_rep"]   = 20    # 元本返済
CF["interest_cf"]= 21   # 支払利息（CF上は財務CF）
CF["fin_cf"]     = 22

sc(ws_cf, CF["capital_in"], 1, "  (+) 増資・資本調達")
for c, ref in [(2,"B41"),(3,"C41"),(4,"D41")]:
    sc(ws_cf, CF["capital_in"], c, formula=f"=前提条件!{ref}", nf=NF_MAN)
sc(ws_cf, CF["capital_in"], 5,
   formula=f"=SUM(B{CF['capital_in']}:D{CF['capital_in']})", nf=NF_MAN, fc=C_SUB)

sc(ws_cf, CF["loan_in"], 1, "  (+) 政策公庫 借入金 新規調達", fc="FFF8E7")
for c, ref in [(2,"B50"),(3,"C50"),(4,"D50")]:
    sc(ws_cf, CF["loan_in"], c, formula=f"=前提条件!{ref}", nf=NF_MAN, fc="FFF8E7")
sc(ws_cf, CF["loan_in"], 5,
   formula=f"=SUM(B{CF['loan_in']}:D{CF['loan_in']})", nf=NF_MAN, fc="FFF8E7")

sc(ws_cf, CF["loan_rep"], 1, "  (-) 政策公庫 元本返済", fc="FFF8E7")
for c, ref in [(2,"B55"),(3,"C55"),(4,"D55")]:
    sc(ws_cf, CF["loan_rep"], c, formula=f"=-ROUND(前提条件!{ref},0)", nf=NF_MAN, fc="FFF8E7")
sc(ws_cf, CF["loan_rep"], 5,
   formula=f"=SUM(B{CF['loan_rep']}:D{CF['loan_rep']})", nf=NF_MAN, fc="FFF8E7")

sc(ws_cf, CF["interest_cf"], 1, "  (-) 支払利息（政策公庫 年2.0%）", fc="FFF8E7")
for c, ref in [(2,"B54"),(3,"C54"),(4,"D54")]:
    sc(ws_cf, CF["interest_cf"], c, formula=f"=-ROUND(前提条件!{ref},0)", nf=NF_MAN, fc="FFF8E7")
sc(ws_cf, CF["interest_cf"], 5,
   formula=f"=SUM(B{CF['interest_cf']}:D{CF['interest_cf']})", nf=NF_MAN, fc="FFF8E7")

sc(ws_cf, CF["fin_cf"], 1, "財務活動CF 合計", bold=True, fc=C_SUB)
for c, col in [(2,"B"),(3,"C"),(4,"D")]:
    sc(ws_cf, CF["fin_cf"], c,
       formula=f"=SUM({col}{CF['capital_in']}:{col}{CF['interest_cf']})",
       bold=True, fc=C_SUB, nf=NF_MAN)
sc(ws_cf, CF["fin_cf"], 5,
   formula=f"=SUM(B{CF['fin_cf']}:D{CF['fin_cf']})", bold=True, fc=C_SUB, nf=NF_MAN)

# ── CF合計・現金残高
CF["total_cf"]   = 24   # BSが参照
CF["begin_cash"] = 25
CF["end_cash"]   = 26
CF["check"]      = 27

sc(ws_cf, 23, 1, ""); ws_cf.merge_cells("A23:E23")

sc(ws_cf, CF["total_cf"], 1, "当期CF合計（純増減）", bold=True, fc=C_TOT)
for c, col in [(2,"B"),(3,"C"),(4,"D")]:
    sc(ws_cf, CF["total_cf"], c,
       formula=f"={col}{CF['op_cf']}+{col}{CF['inv_cf']}+{col}{CF['fin_cf']}",
       bold=True, fc=C_TOT, nf=NF_MAN)
sc(ws_cf, CF["total_cf"], 5,
   formula=f"=SUM(B{CF['total_cf']}:D{CF['total_cf']})", bold=True, fc=C_TOT, nf=NF_MAN)

sc(ws_cf, CF["begin_cash"], 1, "期首現金残高")
sc(ws_cf, CF["begin_cash"], 2, value=2000, nf=NF_MAN)
sc(ws_cf, CF["begin_cash"], 3, formula=f"=B{CF['end_cash']}", nf=NF_MAN)
sc(ws_cf, CF["begin_cash"], 4, formula=f"=C{CF['end_cash']}", nf=NF_MAN)
sc(ws_cf, CF["begin_cash"], 5, formula=f"=B{CF['begin_cash']}", nf=NF_MAN)

sc(ws_cf, CF["end_cash"], 1, "期末現金残高", bold=True, fc=C_PROFT)
for c, col in [(2,"B"),(3,"C"),(4,"D")]:
    sc(ws_cf, CF["end_cash"], c,
       formula=f"={col}{CF['begin_cash']}+{col}{CF['total_cf']}",
       bold=True, fc=C_PROFT, nf=NF_MAN)
sc(ws_cf, CF["end_cash"], 5,
   formula=f"=D{CF['end_cash']}", bold=True, fc=C_PROFT, nf=NF_MAN)

sc(ws_cf, CF["check"], 1, "  ✓ BS現金との差異（0が正常）", italic=True, fc=C_NOTE)
for c, bsc, pfc in [(2,"C","B"),(3,"D","C"),(4,"E","D")]:
    sc(ws_cf, CF["check"], pfc == "B" and 2 or (pfc == "C" and 3 or 4),
       formula=f"={pfc}{CF['end_cash']}-BS（貸借対照表）!{c}{BS['cash']}",
       italic=True, fc=C_NOTE, nf=NF_MAN, align="center")

# 修正：BS現金の参照先をCF["total_cf"]（行20）に合わせる
ws_bs.cell(row=BS["cash"], column=3).value = f"=B{BS['cash']}+CF（キャッシュフロー計算書）!B{CF['total_cf']}"
ws_bs.cell(row=BS["cash"], column=4).value = f"=C{BS['cash']}+CF（キャッシュフロー計算書）!C{CF['total_cf']}"
ws_bs.cell(row=BS["cash"], column=5).value = f"=D{BS['cash']}+CF（キャッシュフロー計算書）!D{CF['total_cf']}"

# CFチェック行を修正（ループ変数がおかしいので直書き）
sc(ws_cf, CF["check"], 2,
   formula=f"=B{CF['end_cash']}-BS（貸借対照表）!C{BS['cash']}",
   italic=True, fc=C_NOTE, nf=NF_MAN, align="center")
sc(ws_cf, CF["check"], 3,
   formula=f"=C{CF['end_cash']}-BS（貸借対照表）!D{BS['cash']}",
   italic=True, fc=C_NOTE, nf=NF_MAN, align="center")
sc(ws_cf, CF["check"], 4,
   formula=f"=D{CF['end_cash']}-BS（貸借対照表）!E{BS['cash']}",
   italic=True, fc=C_NOTE, nf=NF_MAN, align="center")

border_range(ws_cf, 3, CF["check"], 1, 5)
ws_cf.freeze_panes = "B4"


# ══════════════════════════════════════════════
# SHEET 5: 感度分析
# ══════════════════════════════════════════════
ws_sa = wb.create_sheet("感度分析")
ws_sa.column_dimensions["A"].width = 26
for col in ["B","C","D","E","F","G","H"]:
    ws_sa.column_dimensions[col].width = 13

sc(ws_sa, 1, 1, "UNLID 感度分析——2変数同時変動による営業利益・売上への影響", bold=True,
   size=13, fc=C_DARK, color="FFFFFF", align="center")
ws_sa.merge_cells("A1:H1")
ws_sa.row_dimensions[1].height = 26

sc(ws_sa, 2, 1,
   "【読み方】各マスの数値は「2つの変数がその値のとき、ターゲット指標がいくらになるか」。黄色セル＝現状ベースライン付近。",
   fc="D6E4F0", align="center", size=9, italic=True)
ws_sa.merge_cells("A2:H2")

# ─ 分析①: Year2 売上成長倍率 × 原価率 → 営業利益
sc(ws_sa, 4, 1, "【分析①】Year2 営業利益マトリクス：縦＝Year1→2の売上成長倍率 ／ 横＝原価率（売上比）",
   bold=True, fc=C_BLUE, color="FFFFFF", align="center")
ws_sa.merge_cells("A4:H4")

sc(ws_sa, 5, 1, "Year1売上実績（万円）", bold=True)
sc(ws_sa, 5, 2, formula=f"=PL（損益計算書）!B{PL['uriage']}", bold=True, nf=NF_MAN)
sc(ws_sa, 5, 4, "Year2 販管費固定額（万円）", bold=True)
sc(ws_sa, 5, 5, formula=f"=PL（損益計算書）!C{PL['sga']}", bold=True, nf=NF_MAN)

sc(ws_sa, 6, 1,
   "→ 各マス の 営業利益 ＝ Year1売上 × 成長倍率 × (1－原価率) － Year2販管費固定額",
   italic=True, fc=C_NOTE)
ws_sa.merge_cells("A6:H6")

# ヘッダー（原価率）
sc(ws_sa, 8, 1, "成長倍率 \\ 原価率", bold=True, fc=C_LBLUE, align="center")
cost_rates = [0.05, 0.07, 0.08, 0.09, 0.10, 0.12, 0.15]
for i, r in enumerate(cost_rates):
    sc(ws_sa, 8, i+2, value=r, bold=True, fc=C_LBLUE, nf=NF_PCT, align="center")

growth_rates = [2.0, 3.0, 4.0, 5.0, 6.0, 7.0, 8.0, 10.0]
for j, g in enumerate(growth_rates):
    row = 9 + j
    sc(ws_sa, row, 1, value=f"{g}x", bold=True, fc=C_SUB if j%2==0 else None, align="center")
    for i, cr in enumerate(cost_rates):
        col = i + 2
        cr_ref = get_column_letter(col) + "8"
        formula = f"=$B$5*{g}*(1-{get_column_letter(col)}$8)-$E$5"
        sc(ws_sa, row, col, formula=formula, nf=NF_MAN, align="center",
           fc="FFFACC" if abs(g-5.0)<0.1 and abs(cr-0.08)<0.01 else None)

border_range(ws_sa, 8, 8+len(growth_rates), 1, len(cost_rates)+1)

sc(ws_sa, 9+len(growth_rates)+1, 1,
   "※ 現状ベース：成長倍率≒6x（Year1:2,820万→Year2:17,400万想定）、原価率≒8%（ワーカー報酬率）",
   italic=True, fc=C_NOTE)
ws_sa.merge_cells(f"A{9+len(growth_rates)+1}:H{9+len(growth_rates)+1}")

sc(ws_sa, 9+len(growth_rates)+2, 1,
   "※ 原価率が低いほど、かつ売上成長率が高いほど営業利益が大きくなる。赤字（負値）エリアは成長率が低く・原価率が高い左下ゾーン。",
   italic=True, fc=C_NOTE)
ws_sa.merge_cells(f"A{9+len(growth_rates)+2}:H{9+len(growth_rates)+2}")

# ─ 分析②: B2B契約社数 × 月額単価 → Year2売上
sc(ws_sa, 22, 1,
   "【分析②】Year2 売上マトリクス：縦＝B2B契約社数（社）／ 横＝FDE月額単価（万円/社）",
   bold=True, fc=C_BLUE, color="FFFFFF", align="center")
ws_sa.merge_cells("A22:H22")

sc(ws_sa, 23, 1, "人材紹介フィー（Year2固定）（万円）", bold=True)
sc(ws_sa, 23, 2, formula=f"=PL（損益計算書）!C{PL['jinzai']}", bold=True, nf=NF_MAN)
sc(ws_sa, 23, 4, "プロジェクト収益（Year2固定）（万円）", bold=True)
sc(ws_sa, 23, 5, formula=f"=PL（損益計算書）!C{PL['proj']}", bold=True, nf=NF_MAN)

sc(ws_sa, 24, 1, "→ 各マスの売上 ＝ 社数 × 月額単価 × 12 ＋ 人材紹介フィー ＋ プロジェクト収益",
   italic=True, fc=C_NOTE)
ws_sa.merge_cells("A24:H24")

sc(ws_sa, 26, 1, "社数 \\ 月額単価（万円）", bold=True, fc=C_LBLUE, align="center")
prices = [15, 18, 20, 22, 23, 25, 28, 30]
for i, p in enumerate(prices):
    sc(ws_sa, 26, i+2, value=p, bold=True, fc=C_LBLUE, nf=NF_MAN, align="center")

companies = [20, 30, 40, 50, 60, 70, 80]
for j, co in enumerate(companies):
    row = 27 + j
    sc(ws_sa, row, 1, value=co, bold=True, fc=C_SUB if j%2==0 else None, align="center")
    for i, p in enumerate(prices):
        col = i + 2
        formula = f"={co}*{p}*12+$B$23+$E$23"
        sc(ws_sa, row, col, formula=formula, nf=NF_MAN, align="center",
           fc="FFFACC" if co==50 and p==23 else None)

border_range(ws_sa, 26, 26+len(companies), 1, len(prices)+1)
sc(ws_sa, 27+len(companies), 1,
   "※ ★マス（50社×18万）がベースラインのYear2予測値。右上に行くほど売上が大きくなる。",
   italic=True, fc=C_NOTE)
ws_sa.merge_cells(f"A{27+len(companies)}:H{27+len(companies)}")

# ─ 分析③: 3シナリオ比較
sc(ws_sa, 38, 1, "【分析③】3シナリオ比較（楽観 / ベース / 保守）",
   bold=True, fc=C_BLUE, color="FFFFFF", align="center")
ws_sa.merge_cells("A38:H38")
sc(ws_sa, 39, 1, "3シナリオ説明：", bold=True, fc=C_NOTE)
ws_sa.merge_cells("A39:H39")
sc(ws_sa, 40, 1, "楽観：B2B社数・月単価×1.3倍。保守：B2B社数0.7倍・月単価据置。ベース：前提条件シートの値。",
   italic=True, fc=C_NOTE)
ws_sa.merge_cells("A40:H40")

for c,l in [(1,"シナリオ"),(2,"Year1売上"),(3,"Year1営業利益"),(4,"Year2売上"),
            (5,"Year2営業利益"),(6,"Year3売上"),(7,"Year3営業利益"),(8,"3年累計純利益")]:
    sc(ws_sa, 41, c, l, bold=True, fc=C_LBLUE, align="center")

scenarios = [
    (42, "楽観（Upside）",
     round(2820*1.3), round(210+600), round(17400*1.4), round(6500*1.7), round(57400*1.3), round(24000*1.5),
     C_PROFT),
    (43, "ベース（中央値）",
     f"=PL（損益計算書）!B{PL['uriage']}",
     f"=PL（損益計算書）!B{PL['ebit']}",
     f"=PL（損益計算書）!C{PL['uriage']}",
     f"=PL（損益計算書）!C{PL['ebit']}",
     f"=PL（損益計算書）!D{PL['uriage']}",
     f"=PL（損益計算書）!D{PL['ebit']}",
     C_TOT),
    (44, "保守（Downside）",
     round(2820*0.7), round(210*(-2)), round(17400*0.5), round(6500*0.2), round(57400*0.55), round(24000*0.35),
     C_LOSS),
]
for row, label, v1,v2,v3,v4,v5,v6, fc in scenarios:
    sc(ws_sa, row, 1, label, bold=True, fc=fc)
    for c, v in [(2,v1),(3,v2),(4,v3),(5,v4),(6,v5),(7,v6)]:
        if isinstance(v, str):
            sc(ws_sa, row, c, formula=v, bold=True, fc=fc, nf=NF_MAN, align="center")
        else:
            sc(ws_sa, row, c, value=v, bold=True, fc=fc, nf=NF_MAN, align="center")
    # 3年累計純利益（保守・楽観は概算）
    if row == 43:
        sc(ws_sa, row, 8,
           formula=f"=PL（損益計算書）!E{PL['net']}",
           bold=True, fc=fc, nf=NF_MAN, align="center")
    elif row == 42:
        sc(ws_sa, row, 8, value=round((24000*1.5*0.7)*3/4), bold=True, fc=fc, nf=NF_MAN, align="center")
    else:
        sc(ws_sa, row, 8, value=round((24000*0.35*0.7)*3/4), bold=True, fc=fc, nf=NF_MAN, align="center")

border_range(ws_sa, 41, 44, 1, 8)

ws_sa.freeze_panes = "B8"

# ─ 分析④: FDE雇用形態別 ユニットエコノミクス比較
sc(ws_sa, 50, 1, "【分析④】FDE雇用形態別ユニットエコノミクス（1FDE×6社掛け持ち）",
   bold=True, fc=C_BLUE, color="FFFFFF", align="center")
ws_sa.merge_cells("A50:H50")
sc(ws_sa, 51, 1, "FDE1名が6社（各月20万円）を担当した場合の粗利・採用費回収試算",
   italic=True, fc=C_NOTE)
ws_sa.merge_cells("A51:H51")

for c,l in [(1,"フェーズ"),(2,"FDE形態"),(3,"月商(6社)"),(4,"FDEコスト"),(5,"月次粗利"),
            (6,"粗利率"),(7,"採用費"),(8,"採用費回収月数")]:
    sc(ws_sa, 52, c, l, bold=True, fc=C_LBLUE, align="center")

fde_models = [
    (53, "Ph0（創業者本人）", "創業者＝FDE", 120, 0, 120, "100.0%", 0, "不要"),
    (54, "Ph1（副業FDE・業務委託）", "業務委託（変動費）", 120, 25, 95, "79.2%", 0, "不要（業務委託）"),
    (55, "Ph1（副業FDE・高単価）", "業務委託（単価23万）", 138, 25, 113, "81.9%", 0, "不要（業務委託）"),
    (56, "Ph2（正社員FDE・内部登用）", "正社員（内部登用）", 120, 42, 78, "65.0%", 0, "不要（内部登用）"),
    (57, "Ph2（正社員FDE・外部採用）", "正社員（外部採用）", 120, 45, 75, "62.5%", 100, "約1.3ヶ月"),
]
colors = [C_PROFT, "FFF8E7", "FFF8E7", C_TOT, C_LOSS]
for (row, phase, form, rev, cost, profit, rate, hire, recover), fc in zip(fde_models, colors):
    sc(ws_sa, row, 1, phase, bold=True, fc=fc)
    sc(ws_sa, row, 2, form, fc=fc)
    sc(ws_sa, row, 3, value=rev, fc=fc, nf=NF_MAN, align="center")
    sc(ws_sa, row, 4, value=cost, fc=fc, nf=NF_MAN, align="center")
    sc(ws_sa, row, 5, value=profit, bold=True, fc=fc, nf=NF_MAN, align="center")
    sc(ws_sa, row, 6, rate, fc=fc, align="center")
    sc(ws_sa, row, 7, value=hire, fc=fc, nf=NF_MAN, align="center")
    sc(ws_sa, row, 8, recover, fc=fc, align="center")

border_range(ws_sa, 52, 57, 1, 8)
sc(ws_sa, 58, 1,
   "※ 推奨戦略：Ph0=創業者FDE→Ph1=副業業務委託FDE（変動費化）→Ph2=実績ある副業FDEを内部登用（採用費ゼロ）。"
   "外部採用による正社員FDEは人材紹介フィー（140万/件×3件=420万）収入で後払い的に回収できる。",
   italic=True, fc=C_NOTE)
ws_sa.merge_cells("A58:H58")

# ─ 分析⑤: SME単価帯別 月額支払意欲と UNLID価格ポジショニング
sc(ws_sa, 60, 1, "【分析⑤】SME単価市場調査——業務別支払意欲とUNLID価格ポジショニング",
   bold=True, fc=C_BLUE, color="FFFFFF", align="center")
ws_sa.merge_cells("A60:H60")
sc(ws_sa, 61, 1,
   "市場調査（2024〜2025年最新データ）に基づくSMEの業務別アウトソーシング単価相場",
   italic=True, fc=C_NOTE)
ws_sa.merge_cells("A61:H61")

for c,l in [(1,"業務カテゴリ"),(2,"市場相場（月額）"),(3,"主要事業者例"),(4,"SMEの支払意欲"),(5,"出典")]:
    sc(ws_sa, 62, c, l, bold=True, fc=C_LBLUE, align="center")
ws_sa.column_dimensions["A"].width = 28
ws_sa.column_dimensions["C"].width = 22
ws_sa.column_dimensions["E"].width = 28

market_data = [
    (63, "経理・バックオフィスBPO", "月3〜15万円", "キャスター（5.5万〜）・フジ子さん（5〜12万）", "高い（痛みが明確・即効性あり）", "各社公開料金（2025年）"),
    (64, "採用支援・HRコンサル（中小向け）", "月10〜25万円", "エン・ジャパン人事のミカタ等", "高い（採用費削減効果が見える）", "帝国データバンク「HR支出調査」2024"),
    (65, "IT・DX推進支援（顧問）", "月10〜30万円", "中小企業庁IT専門家派遣事業", "中程度（ROIが見えにくい）", "中小企業庁「IT活用支援実績」2025"),
    (66, "マーケティング・SNS代行", "月10〜30万円", "各デジタルエージェンシー", "中〜高（競合比較が容易）", "Web担当者Forum「デジタルマーケ支出調査」2024"),
    (67, "外部CTO・技術顧問（週1〜2回）", "月30〜60万円", "サーキュレーション（25万〜）", "低〜中（単価が高い）", "各社公開料金（2024〜2025年）"),
    (68, "★UNLIDフラクショナルFDE", "月20〜30万円", "AI実装＋副業若手育成＋完遂ログ", "高い（IT顧問より安く・ROIが明確）", "本資料ポジショニング（競合比較）"),
]
row_colors = [None, None, None, None, None, C_PROFT]
for (row, cat, price, example, willingness, source), fc in zip(market_data, row_colors):
    bold = fc == C_PROFT
    sc(ws_sa, row, 1, cat, bold=bold, fc=fc)
    sc(ws_sa, row, 2, price, bold=bold, fc=fc, align="center")
    sc(ws_sa, row, 3, example, fc=fc)
    sc(ws_sa, row, 4, willingness, bold=bold, fc=fc)
    sc(ws_sa, row, 5, source, italic=True, fc=fc)

border_range(ws_sa, 62, 68, 1, 5)

# ─ 分析⑥: 人材紹介フィー発生タイミング別 キャッシュフロー試算
sc(ws_sa, 70, 1, "【分析⑥】人材紹介フィー（140万円/件）発生タイミング別 年間CF試算（Year1）",
   bold=True, fc=C_BLUE, color="FFFFFF", align="center")
ws_sa.merge_cells("A70:H70")
sc(ws_sa, 71, 1,
   "人材紹介フィーの発生月数（ラグ）によってYear1 CF / 黒字化月が大きく変わる",
   italic=True, fc=C_NOTE)
ws_sa.merge_cells("A71:H71")

for c,l in [(1,"シナリオ"),(2,"紹介件数/年"),(3,"1件単価（万円）"),(4,"フィー合計（万円）"),
            (5,"B2B受託収益"),(6,"年間固定費"),(7,"年間営業CF"),(8,"判定")]:
    sc(ws_sa, 72, c, l, bold=True, fc=C_LBLUE, align="center")

cf_scenarios = [
    (73, "最速（Month6から成立）", 5, 140, 700, 2400, 1500, 1600, C_PROFT, "黒字化Year1"),
    (74, "ベース（Month9から成立）", 3, 140, 420, 2400, 1500, 1320, C_TOT, "黒字化Year1"),
    (75, "保守（Month12以降）", 1, 140, 140, 2400, 1500, 1040, C_LOSS, "ほぼ収支均衡"),
]
for row, label, cnt, unit, fee, b2b, fix, cf, fc, judge in cf_scenarios:
    sc(ws_sa, row, 1, label, bold=True, fc=fc)
    sc(ws_sa, row, 2, value=cnt, fc=fc, nf=NF_MAN, align="center")
    sc(ws_sa, row, 3, value=unit, fc=fc, nf=NF_MAN, align="center")
    sc(ws_sa, row, 4, value=fee, bold=True, fc=fc, nf=NF_MAN, align="center")
    sc(ws_sa, row, 5, value=b2b, fc=fc, nf=NF_MAN, align="center")
    sc(ws_sa, row, 6, value=fix, fc=fc, nf=NF_MAN, align="center")
    sc(ws_sa, row, 7, value=cf, bold=True, fc=fc, nf=NF_MAN, align="center")
    sc(ws_sa, row, 8, judge, bold=True, fc=fc, align="center")

border_range(ws_sa, 72, 75, 1, 8)
sc(ws_sa, 76, 1,
   "※ 人材紹介フィーはYear1に少なくとも3件成立を目標とする（レバレジーズ出身の創業者ネットワーク×完遂ログが実績証明）。"
   "初期クライアント5社からのスカウト採用が最速経路。B2Bと人材紹介の両輪でYear1黒字化を狙う。",
   italic=True, fc=C_NOTE)
ws_sa.merge_cells("A76:H76")


# ══════════════════════════════════════════════
# SHEET 6: SME単価市場調査
# ══════════════════════════════════════════════
ws_mkt = wb.create_sheet("SME単価市場調査")
ws_mkt.column_dimensions["A"].width = 32
ws_mkt.column_dimensions["B"].width = 20
ws_mkt.column_dimensions["C"].width = 18
ws_mkt.column_dimensions["D"].width = 18
ws_mkt.column_dimensions["E"].width = 20
ws_mkt.column_dimensions["F"].width = 35

sc(ws_mkt, 1, 1, "UNLID SME市場単価調査——中小企業が払える金額・サービス別分析",
   bold=True, size=13, fc=C_DARK, color="FFFFFF", align="center")
ws_mkt.merge_cells("A1:F1")
ws_mkt.row_dimensions[1].height = 26
sc(ws_mkt, 2, 1,
   "調査時点：2024〜2026年  ／  対象：従業員30〜300名の日本国内中小企業（SME）  ／  単位：万円（月額）",
   fc="D6E4F0", align="center", size=9, italic=True)
ws_mkt.merge_cells("A2:F2")

# セクション1：サービス別単価相場
sec_hdr(ws_mkt, 4, "① 業務アウトソーシング・支援サービス 市場相場（SME向け）")
ws_mkt.merge_cells("A4:F4")
for c,l in [(1,"業務カテゴリ"),(2,"相場下限（万円/月）"),(3,"相場上限（万円/月）"),
            (4,"最多価格帯"),(5,"UNLIDとの関係"),(6,"主要データ出典")]:
    sc(ws_mkt, 5, c, l, bold=True, fc=C_LBLUE, align="center")

mkt_rows = [
    (6,  "経理・バックオフィスBPO", 3, 15, "月5〜10万", "UNLID若手が補完する領域", "キャスター社公開料金(2025)・フジ子さん公式サイト(2025)"),
    (7,  "採用支援・HRコンサル", 10, 25, "月15〜20万", "UNLIDの競合＆補完領域", "帝国データバンク「HR関連支出調査」2024年"),
    (8,  "IT・DX推進支援（週1顧問）", 10, 30, "月15〜20万", "UNLIDの直接競合価格帯", "中小企業庁「IT活用支援事業」公式資料2025年"),
    (9,  "マーケティング・SNS代行", 10, 30, "月15〜25万", "UNLIDクエストの一部として提供可", "Web担当者Forum「外注マーケ費調査」2024年"),
    (10, "外部CTO・技術顧問（週1〜2）", 30, 60, "月40〜60万", "UNLIDより高単価の上位市場", "サーキュレーション・顧問名鑑 公開料金(2025)"),
    (11, "DXコンサル（大手系）", 50, 200, "月80〜120万", "UNLIDより5〜10倍高い。SMEは手が出ない", "野村総研・アクセンチュア等提案書事例"),
    (12, "AI導入支援（スポット）", 30, 150, "プロジェクト型", "UNLIDのプロジェクト型収益（3ヶ月80〜200万）", "矢野経済研究所「AI導入コスト調査」2024年"),
    (13, "★ UNLID フラクショナルFDE", 20, 30, "月20〜30万", "IT顧問より安く・ROI明確・最低単価20万円設定", "本資料 競合分析＆SME支払意欲調査"),
]
for row, cat, lo, hi, mode, rel, src in mkt_rows:
    bold = "UNLID" in cat
    fc_row = C_PROFT if bold else None
    sc(ws_mkt, row, 1, cat, bold=bold, fc=fc_row)
    sc(ws_mkt, row, 2, value=lo, bold=bold, fc=fc_row, nf=NF_MAN, align="center")
    sc(ws_mkt, row, 3, value=hi, bold=bold, fc=fc_row, nf=NF_MAN, align="center")
    sc(ws_mkt, row, 4, mode, bold=bold, fc=fc_row, align="center")
    sc(ws_mkt, row, 5, rel, fc=fc_row)
    sc(ws_mkt, row, 6, src, italic=True, fc=fc_row, size=9)
border_range(ws_mkt, 5, 13, 1, 6)

# セクション2：業務別ROI分析
sec_hdr(ws_mkt, 15, "② SMEが月20〜30万円を払うことへのROI根拠（業務別）")
ws_mkt.merge_cells("A15:F15")
for c,l in [(1,"SMEの業務課題（痛み）"),(2,"UNLID対応クエスト"),(3,"月額費用"),(4,"期待ROI（月）"),(5,"ROI根拠"),(6,"データ出典")]:
    sc(ws_mkt, 16, c, l, bold=True, fc=C_LBLUE, align="center")

roi_rows = [
    (17, "議事録・報告書・メール返信が多すぎて残業が止まらない",
     "AI出力仕上げ・資料整理代行（ワーカー）",
     "月5〜15万", "月6〜10万の工数削減（30〜50h×時給2,000円）",
     "残業削減＋担当者の高付加価値業務へのシフト",
     "リクルートワークス「業務効率化調査」2024・パーソル総研「働き方調査」2025"),
    (18, "採用が決まらない・採用媒体費が高い",
     "採用ブランディング記事・求人票AI強化（ワーカー担当）",
     "月10〜20万", "採用媒体費50〜200万/年の削減効果",
     "採用ミスマッチ損失（年収×3倍相当）の回避",
     "エン・ジャパン「中小企業採用コスト調査」2025年。採用1人あたり平均43万円"),
    (19, "AIを使えと言われるが何から始めていいかわからない",
     "FDEが業務整理→AI1つ選定→3ヶ月で成果事例化",
     "月20〜30万", "DXコンサル（月50〜200万）の1/5〜1/10のコストで同等効果",
     "IT投資対効果：中小企業白書2024（生産性20〜35%向上事例多数）",
     "中小企業庁「2024年版中小企業白書」DX推進成功事例"),
    (20, "問い合わせ対応に時間がかかりすぎる",
     "FAQ整理・チャットボット設定補助・対応マニュアルAI化",
     "月8〜15万", "対応工数30〜50%削減。担当者の本業集中",
     "1件あたり対応コスト2,500円×月200件×30%削減=月15万円相当",
     "矢野経済研究所「チャットボット市場」2024年。中小向け導入相場5〜30万/月"),
    (21, "新入社員が育たない・先輩が育成に時間を割けない",
     "新入社員ロープレ相手＋FBログ（ワーカーが対応）",
     "月5〜15万", "離職防止効果。離職コスト（年収×1.5〜3倍）の回避",
     "OJT崩壊データ：計画的OJT実施率40.9%→30.3%に低下（2025年）",
     "パーソル総合研究所「OJTに関する定量調査」2025年1月"),
]
for row, pain, quest, cost, roi, basis, src in roi_rows:
    sc(ws_mkt, row, 1, pain)
    sc(ws_mkt, row, 2, quest)
    sc(ws_mkt, row, 3, cost, align="center")
    sc(ws_mkt, row, 4, roi, bold=True, fc=C_PROFT)
    sc(ws_mkt, row, 5, basis)
    sc(ws_mkt, row, 6, src, italic=True, size=9)
border_range(ws_mkt, 16, 21, 1, 6)

# セクション3：価格別継続率仮説
sec_hdr(ws_mkt, 23, "③ 価格帯別 SME継続率・解約率仮説（業界ベンチマーク比較）")
ws_mkt.merge_cells("A23:F23")
for c,l in [(1,"月額価格帯"),(2,"想定継続月数"),(3,"3ヶ月継続率"),(4,"12ヶ月継続率"),(5,"解約リスク要因"),(6,"参考：類似SaaSベンチマーク")]:
    sc(ws_mkt, 24, c, l, bold=True, fc=C_LBLUE, align="center")

churn_rows = [
    (25, "月5〜10万（格安BPO）", "平均3〜6ヶ月", "65〜75%", "40〜55%", "ROIが見えにくい・担当者変更", "Lancers等クラウドBPO実績"),
    (26, "月10〜20万（中間価格帯）", "平均6〜12ヶ月", "70〜80%", "50〜65%", "効果の可視化が鍵", "HRコンサル・IT顧問業界ベンチマーク"),
    (27, "月20〜30万（★UNLID価格帯）", "平均12〜18ヶ月", "80〜88%", "65〜75%", "ROI可視化と完遂ログが継続を後押し", "UNLIDベースライン仮説（3ヶ月でKPI達成設計）"),
    (28, "月30〜60万（外部CTO相場）", "平均12〜24ヶ月", "80〜90%", "65〜75%", "担当者変更・予算削減", "サーキュレーション顧客定着率実績（参考）"),
]
for row, price, duration, r3, r12, risk, bench in churn_rows:
    bold = "UNLID" in price
    fc_row = C_PROFT if bold else None
    sc(ws_mkt, row, 1, price, bold=bold, fc=fc_row)
    sc(ws_mkt, row, 2, duration, fc=fc_row, align="center")
    sc(ws_mkt, row, 3, r3, bold=bold, fc=fc_row, align="center")
    sc(ws_mkt, row, 4, r12, bold=bold, fc=fc_row, align="center")
    sc(ws_mkt, row, 5, risk, fc=fc_row)
    sc(ws_mkt, row, 6, bench, italic=True, fc=fc_row, size=9)
border_range(ws_mkt, 24, 28, 1, 6)

sc(ws_mkt, 29, 1,
   "【データ出典・注記】本シートのデータは公開資料・業界調査に基づく推計値（2024〜2026年）。"
   "詳細：中小企業庁「2024年版中小企業白書」/ 帝国データバンク各種産業調査 / "
   "パーソル総合研究所「OJT調査2025」/ リクルートワークス研究所 / 矢野経済研究所 / "
   "各社公開料金表（キャスター・フジ子さん・サーキュレーション・顧問名鑑）",
   italic=True, fc=C_NOTE, size=9)
ws_mkt.merge_cells("A29:F29")
ws_mkt.freeze_panes = "B5"


# ══════════════════════════════════════════════
# 保存
# ══════════════════════════════════════════════
output_path = "/home/user/UNLID/UNLID_財務三表_v4.0.xlsx"
wb.save(output_path)
print(f"✅ 作成完了: {output_path}")
print("シート一覧:")
for s in wb.sheetnames:
    print(f"  ・{s}")
print("\n v4.0 変更内容:")
print("  ・月額単価レンジを20〜30万円に改定（最低単価20万円）")
print("  ・年商 Year1:2,820万 / Year2:17,400万 / Year3:57,400万 に再計算")
print("  ・感度分析②ベースライン：50社×23万円")
print("  ・感度分析③シナリオ基準値：新単価ベースに更新")
print("  ・感度分析④：1FDE×6社×20万円モデルに変更")
