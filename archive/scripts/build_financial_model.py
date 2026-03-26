import openpyxl
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              numbers)

from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import ColorScaleRule
import copy

# ============================================================
# COLOR PALETTE (UNLID Brand Colors)
# ============================================================
NAVY   = "0F172A"
BLUE   = "2563EB"
LBLUE  = "EFF6FF"
SKY    = "93C5FD"
GREEN  = "86EFAC"
WHITE  = "FFFFFF"
GRAY1  = "F8FAFC"
GRAY2  = "E2E8F0"
AMBER  = "FEF3C7"
RED    = "FEE2E2"
ORANGE = "FED7AA"

def header_fill(color=NAVY):
    return PatternFill("solid", fgColor=color)

def cell_fill(color):
    return PatternFill("solid", fgColor=color)

def hdr_font(color=WHITE, bold=True, size=10):
    return Font(name="Meiryo UI", bold=bold, color=color, size=size)

def body_font(bold=False, size=10, color="000000"):
    return Font(name="Meiryo UI", bold=bold, size=size, color=color)

def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def bold_border():
    s = Side(style="medium", color="888888")
    return Border(left=s, right=s, top=s, bottom=s)

def num_fmt(ws, cell, fmt="#,##0"):
    ws[cell].number_format = fmt

def set_cell(ws, row, col, value, fill=None, font=None, align=None, border=None, num_format=None):
    c = ws.cell(row=row, column=col, value=value)
    if fill:   c.fill = fill
    if font:   c.font = font
    if align:  c.alignment = align
    if border: c.border = border
    if num_format: c.number_format = num_format
    return c

def write_header(ws, row, col, text, colspan=1, color=NAVY, fgcolor=WHITE, size=10):
    c = ws.cell(row=row, column=col, value=text)
    c.fill = header_fill(color)
    c.font = hdr_font(fgcolor, size=size)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = thin_border()
    if colspan > 1:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+colspan-1)
    return c

def write_label(ws, row, col, text, indent=0, bold=False, fill_color=None):
    val = ("　" * indent) + text
    c = ws.cell(row=row, column=col, value=val)
    c.font = body_font(bold=bold)
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border = thin_border()
    if fill_color:
        c.fill = cell_fill(fill_color)
    return c

def write_number(ws, row, col, value, fill_color=None, bold=False, fmt="#,##0"):
    c = ws.cell(row=row, column=col, value=value)
    c.font = body_font(bold=bold)
    c.alignment = Alignment(horizontal="right", vertical="center")
    c.number_format = fmt
    c.border = thin_border()
    if fill_color:
        c.fill = cell_fill(fill_color)
    return c

def write_pct(ws, row, col, value, fill_color=None, bold=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font = body_font(bold=bold)
    c.alignment = Alignment(horizontal="right", vertical="center")
    c.number_format = "0.0%"
    c.border = thin_border()
    if fill_color:
        c.fill = cell_fill(fill_color)
    return c

def section_title(ws, row, col, text, width=10):
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+width-1)
    c = ws.cell(row=row, column=col, value=text)
    c.fill = header_fill(BLUE)
    c.font = hdr_font(WHITE, size=11)
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border = thin_border()
    return c

wb = openpyxl.Workbook()

# ============================================================
# SHEET 1: 前提条件
# ============================================================
ws1 = wb.active
ws1.title = "前提条件"
ws1.column_dimensions["A"].width = 36
ws1.column_dimensions["B"].width = 18
ws1.column_dimensions["C"].width = 18
ws1.column_dimensions["D"].width = 18
ws1.column_dimensions["E"].width = 22

# Title
ws1.row_dimensions[1].height = 30
ws1.merge_cells("A1:E1")
c = ws1["A1"]
c.value = "UNLID 財務三表 v4.0 — 前提条件シート（強化版）"
c.fill = header_fill(NAVY)
c.font = Font(name="Meiryo UI", bold=True, color=WHITE, size=14)
c.alignment = Alignment(horizontal="center", vertical="center")

# Column headers
for col, txt in enumerate(["項目", "Year1 (FY2026)", "Year2 (FY2027)", "Year3 (FY2028)", "備考"], 1):
    write_header(ws1, 2, col, txt)

row = 3

# ---- SECTION A: チャーン・顧客獲得動態 ----
section_title(ws1, row, 1, "【A】チャーン・顧客獲得動態（★新設）", width=5); row+=1

write_header(ws1, row, 1, "月次チャーン率（Base）", color=GRAY2, fgcolor=NAVY); 
write_pct(ws1, row, 2, 0.025); write_pct(ws1, row, 3, 0.025); write_pct(ws1, row, 4, 0.020)
ws1.cell(row=row, column=5).value = "¥200k帯 年間26%解約。Y3は信頼蓄積で改善"
row+=1

write_header(ws1, row, 1, "年換算チャーン率（参考）", color=GRAY2, fgcolor=NAVY)
write_pct(ws1, row, 2, 1-(1-0.025)**12); write_pct(ws1, row, 3, 1-(1-0.025)**12); write_pct(ws1, row, 4, 1-(1-0.020)**12)
ws1.cell(row=row, column=5).value = "=(1-(1-月次)^12) で自動計算"
row+=1

write_header(ws1, row, 1, "期首契約社数", color=GRAY2, fgcolor=NAVY)
write_number(ws1, row, 2, 0); write_number(ws1, row, 3, 10); write_number(ws1, row, 4, 50)
row+=1

write_header(ws1, row, 1, "期末契約社数（ネット目標）", color=GRAY2, fgcolor=NAVY)
write_number(ws1, row, 2, 10, bold=True); write_number(ws1, row, 3, 50, bold=True); write_number(ws1, row, 4, 160, bold=True)
row+=1

write_header(ws1, row, 1, "　うち年間チャーン予測（社）", color=GRAY1, fgcolor="555555")
write_number(ws1, row, 2, round(5*0.26)); write_number(ws1, row, 3, round(30*0.26)); write_number(ws1, row, 4, round(105*0.22))
ws1.cell(row=row, column=5).value = "平均在籍社数×年間チャーン率"
row+=1

write_header(ws1, row, 1, "　グロス新規獲得必要数（社）", color=LBLUE, fgcolor=NAVY)
write_number(ws1, row, 2, 11, bold=True, fill_color=LBLUE); write_number(ws1, row, 3, 49, bold=True, fill_color=LBLUE); write_number(ws1, row, 4, 133, bold=True, fill_color=LBLUE)
ws1.cell(row=row, column=5).value = "ネット増加+チャーン補填 → 営業目標の真の数字"
row+=1

write_header(ws1, row, 1, "　月平均グロス新規獲得（社/月）", color=LBLUE, fgcolor=NAVY)
write_number(ws1, row, 2, round(11/12,1), fill_color=LBLUE); write_number(ws1, row, 3, round(49/12,1), fill_color=LBLUE); write_number(ws1, row, 4, round(133/12,1), fill_color=LBLUE)
ws1.cell(row=row, column=5).value = "FDE1名の営業目標の設計に使う"
row+=2

# ---- SECTION B: 売上ドライバー ----
section_title(ws1, row, 1, "【B】売上ドライバー", width=5); row+=1

write_label(ws1, row, 1, "B2B月次顧問（ネット平均社数）", bold=True)
write_number(ws1, row, 2, 5); write_number(ws1, row, 3, 30); write_number(ws1, row, 4, 105)
ws1.cell(row=row, column=5).value = "=(期首+期末)/2 の年間平均"
row+=1

write_label(ws1, row, 1, "　FDE月額単価（万円）")
write_number(ws1, row, 2, 20); write_number(ws1, row, 3, 23); write_number(ws1, row, 4, 25)
row+=1

write_label(ws1, row, 1, "　B2B売上合計（万円）", bold=True)
write_number(ws1, row, 2, 5*20*12, bold=True, fill_color=LBLUE)
write_number(ws1, row, 3, 30*23*12, bold=True, fill_color=LBLUE)
write_number(ws1, row, 4, 105*25*12, bold=True, fill_color=LBLUE)
row+=1

write_label(ws1, row, 1, "人材紹介フィー（件数）")
write_number(ws1, row, 2, 3); write_number(ws1, row, 3, 20); write_number(ws1, row, 4, 50)
row+=1

write_label(ws1, row, 1, "　紹介フィー単価（万円）")
write_number(ws1, row, 2, 140); write_number(ws1, row, 3, 140); write_number(ws1, row, 4, 140)
ws1.cell(row=row, column=5).value = "年収400万×35%"
row+=1

write_label(ws1, row, 1, "　紹介フィー売上合計（万円）", bold=True)
write_number(ws1, row, 2, 3*140, bold=True, fill_color=LBLUE)
write_number(ws1, row, 3, 20*140, bold=True, fill_color=LBLUE)
write_number(ws1, row, 4, 50*140, bold=True, fill_color=LBLUE)
row+=1

write_label(ws1, row, 1, "AIプロジェクト受託（件数）")
write_number(ws1, row, 2, 0); write_number(ws1, row, 3, 5); write_number(ws1, row, 4, 15)
row+=1

write_label(ws1, row, 1, "　プロジェクト単価（万円）")
write_number(ws1, row, 2, 160); write_number(ws1, row, 3, 160); write_number(ws1, row, 4, 160)
row+=1

write_label(ws1, row, 1, "　プロジェクト売上合計（万円）", bold=True)
write_number(ws1, row, 2, 0, bold=True, fill_color=LBLUE)
write_number(ws1, row, 3, 5*160, bold=True, fill_color=LBLUE)
write_number(ws1, row, 4, 15*160, bold=True, fill_color=LBLUE)
row+=2

# ---- SECTION C: ヘッドカウント計画（★新設） ----
section_title(ws1, row, 1, "【C】ヘッドカウント計画（★新設）", width=5); row+=1

write_label(ws1, row, 1, "FDE（正社員）人数", bold=True)
write_number(ws1, row, 2, 2); write_number(ws1, row, 3, 9); write_number(ws1, row, 4, 27)
ws1.cell(row=row, column=5).value = "1名/6社担当 = ネット平均社数÷6"
row+=1

write_label(ws1, row, 1, "　FDE月額給与（万円）")
write_number(ws1, row, 2, 45); write_number(ws1, row, 3, 42); write_number(ws1, row, 4, 42)
row+=1

write_label(ws1, row, 1, "　FDE人件費合計（万円/年）", bold=True)
write_number(ws1, row, 2, 2*45*12, bold=True, fill_color=LBLUE)
write_number(ws1, row, 3, 9*42*12, bold=True, fill_color=LBLUE)
write_number(ws1, row, 4, 27*42*12, bold=True, fill_color=LBLUE)
row+=1

write_label(ws1, row, 1, "CSエンジニア（実装担当）人数")
write_number(ws1, row, 2, 1); write_number(ws1, row, 3, 3); write_number(ws1, row, 4, 8)
ws1.cell(row=row, column=5).value = "プロジェクト件数連動。Y3=15件÷2件/人"
row+=1

write_label(ws1, row, 1, "　CS月額給与（万円）")
write_number(ws1, row, 2, 40); write_number(ws1, row, 3, 40); write_number(ws1, row, 4, 40)
row+=1

write_label(ws1, row, 1, "　CS人件費合計（万円/年）", bold=True)
write_number(ws1, row, 2, 1*40*12, bold=True, fill_color=LBLUE)
write_number(ws1, row, 3, 3*40*12, bold=True, fill_color=LBLUE)
write_number(ws1, row, 4, 8*40*12, bold=True, fill_color=LBLUE)
row+=1

write_label(ws1, row, 1, "管理・営業スタッフ数")
write_number(ws1, row, 2, 0); write_number(ws1, row, 3, 2); write_number(ws1, row, 4, 6)
row+=1

write_label(ws1, row, 1, "　管理系月額給与（万円）")
write_number(ws1, row, 2, 35); write_number(ws1, row, 3, 35); write_number(ws1, row, 4, 35)
row+=1

write_label(ws1, row, 1, "代表報酬（万円/年）")
write_number(ws1, row, 2, 360); write_number(ws1, row, 3, 600); write_number(ws1, row, 4, 840)
row+=1

write_label(ws1, row, 1, "副業ワーカー費（売上比）")
write_pct(ws1, row, 2, 0.09); write_pct(ws1, row, 3, 0.08); write_pct(ws1, row, 4, 0.08)
row+=2

# ---- SECTION D: AI原価・API費用（★新設） ----
section_title(ws1, row, 1, "【D】AI原価・API費用（★新設）", width=5); row+=1

write_label(ws1, row, 1, "【前提】顧客負担か？")
ws1.cell(row=row, column=2).value = "顧客自費"
ws1.cell(row=row, column=3).value = "顧客自費"
ws1.cell(row=row, column=4).value = "顧客自費"
ws1.cell(row=row, column=5).value = "Claude Code等は顧客が自社で契約・負担"
for col in [2,3,4]: ws1.cell(row=row, column=col).font = body_font(bold=True, color="16A34A"); ws1.cell(row=row, column=col).alignment = Alignment(horizontal="center")
row+=1

write_label(ws1, row, 1, "UNLID内部AI利用料（万円/年）")
write_number(ws1, row, 2, 5); write_number(ws1, row, 3, 12); write_number(ws1, row, 4, 24)
ws1.cell(row=row, column=5).value = "社内自動化・レポート生成等の内部ツール費のみ"
row+=1

write_label(ws1, row, 1, "クラウドインフラ費（万円/年）")
write_number(ws1, row, 2, 8); write_number(ws1, row, 3, 20); write_number(ws1, row, 4, 48)
ws1.cell(row=row, column=5).value = "完遂ログDB・マッチングシステムのサーバー費"
row+=1

write_label(ws1, row, 1, "　AI・インフラ費合計（万円/年）", bold=True)
write_number(ws1, row, 2, 5+8, bold=True, fill_color=LBLUE)
write_number(ws1, row, 3, 12+20, bold=True, fill_color=LBLUE)
write_number(ws1, row, 4, 24+48, bold=True, fill_color=LBLUE)
ws1.cell(row=row, column=5).value = "売上比0.5%未満→コスト影響は軽微"
row+=2

# ---- SECTION E: 運転資本・入出金サイト（★新設） ----
section_title(ws1, row, 1, "【E】運転資本・入出金サイト（★新設）", width=5); row+=1

write_label(ws1, row, 1, "売掛金回収サイト（DSO：日）", bold=True)
write_number(ws1, row, 2, 45); write_number(ws1, row, 3, 45); write_number(ws1, row, 4, 45)
ws1.cell(row=row, column=5).value = "当月請求→翌月末入金（B2B標準）"
row+=1

write_label(ws1, row, 1, "買掛金支払サイト（DPO：日）")
write_number(ws1, row, 2, 30); write_number(ws1, row, 3, 30); write_number(ws1, row, 4, 30)
ws1.cell(row=row, column=5).value = "副業ワーカーへの支払い"
row+=1

write_label(ws1, row, 1, "給与支払日")
for col in [2,3,4]: ws1.cell(row=row, column=col).value = "当月25日"
ws1.cell(row=row, column=5).value = "売上入金の前に出ていく→キャッシュ先行支出"
row+=1

write_label(ws1, row, 1, "必要運転資金目安（万円）", bold=True)
write_number(ws1, row, 2, round(1200*45/365)+round(1200*0.09*30/365), bold=True, fill_color=AMBER)
write_number(ws1, row, 3, round(8280*45/365)+round(8280*0.08*30/365), bold=True, fill_color=AMBER)
write_number(ws1, row, 4, round(57500*45/365)+round(57500*0.08*30/365), bold=True, fill_color=AMBER)
ws1.cell(row=row, column=5).value = "売上×DSO/365 + 変動費×DPO/365"
row+=2

# ---- SECTION F: ソフトウェア資産（★新設） ----
section_title(ws1, row, 1, "【F】ソフトウェア資産計上方針（★新設）", width=5); row+=1

write_label(ws1, row, 1, "自社プロダクト開発費の計上方針")
ws1.cell(row=row, column=2).value = "費用処理"
ws1.cell(row=row, column=3).value = "資産計上開始"
ws1.cell(row=row, column=4).value = "資産計上継続"
ws1.cell(row=row, column=5).value = "Y2よりBSにソフトウェア計上→減価償却"
row+=1

write_label(ws1, row, 1, "開発投資額（万円/年）")
write_number(ws1, row, 2, 120); write_number(ws1, row, 3, 360); write_number(ws1, row, 4, 600)
row+=1

write_label(ws1, row, 1, "資産計上額（万円/年）")
write_number(ws1, row, 2, 0); write_number(ws1, row, 3, 240); write_number(ws1, row, 4, 480)
ws1.cell(row=row, column=5).value = "開発費の67%を無形固定資産として計上"
row+=1

write_label(ws1, row, 1, "償却費（定額3年、万円/年）")
write_number(ws1, row, 2, 0); write_number(ws1, row, 3, 40); write_number(ws1, row, 4, 120)
ws1.cell(row=row, column=5).value = "Y2計上分240÷3=80万/年（半期認識でY2=40万）"
row+=1

write_label(ws1, row, 1, "ソフトウェア資産（BS計上額、万円）", bold=True)
write_number(ws1, row, 2, 0, bold=True, fill_color=LBLUE)
write_number(ws1, row, 3, 200, bold=True, fill_color=LBLUE)
write_number(ws1, row, 4, 560, bold=True, fill_color=LBLUE)
ws1.cell(row=row, column=5).value = "累積計上-累積償却"
row+=2

# ---- SECTION G: CAC・LTV ----
section_title(ws1, row, 1, "【G】CAC・LTV・ユニットエコノミクス", width=5); row+=1

write_label(ws1, row, 1, "CAC（顧客獲得単価、万円）", bold=True)
write_number(ws1, row, 2, 15); write_number(ws1, row, 3, 25); write_number(ws1, row, 4, 40)
ws1.cell(row=row, column=5).value = "Y2後半以降、アーリー層取り切りでCACが上昇"
row+=1

write_label(ws1, row, 1, "LTV（月20万×12ヶ月+紹介140万、万円）")
write_number(ws1, row, 2, 380); write_number(ws1, row, 3, 380); write_number(ws1, row, 4, 420)
row+=1

write_label(ws1, row, 1, "LTV/CAC倍率", bold=True)
write_number(ws1, row, 2, round(380/15,1), bold=True, fill_color=LBLUE)
write_number(ws1, row, 3, round(380/25,1), bold=True, fill_color=LBLUE)
write_number(ws1, row, 4, round(420/40,1), bold=True, fill_color=LBLUE)
ws1.cell(row=row, column=5).value = "3倍以上が健全ライン（SaaS業界基準）"
row+=1

write_label(ws1, row, 1, "CAC回収期間（ヶ月）")
write_number(ws1, row, 2, round(15/(20*0.50),1)); write_number(ws1, row, 3, round(25/(23*0.50),1)); write_number(ws1, row, 4, round(40/(25*0.50),1))
ws1.cell(row=row, column=5).value = "CAC÷月次粗利（月額単価×粗利率50%）"
row+=2

print("Sheet 1 done. Row:", row)

# ============================================================
# SHEET 2: PL（損益計算書）
# ============================================================
ws2 = wb.create_sheet("PL（損益計算書）")
ws2.column_dimensions["A"].width = 38
for col in ["B","C","D","E"]:
    ws2.column_dimensions[col].width = 18

# Title
ws2.merge_cells("A1:E1")
c = ws2["A1"]
c.value = "UNLID 損益計算書（P&L） — チャーン・ヘッドカウント連動版"
c.fill = header_fill(NAVY); c.font = Font(name="Meiryo UI", bold=True, color=WHITE, size=13)
c.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 28

for col, txt in enumerate(["項目", "Year1 (FY2026)", "Year2 (FY2027)", "Year3 (FY2028)", "3年累計"], 1):
    write_header(ws2, 2, col, txt)

# Revenue assumptions (same as sheet 1)
y1_b2b = 5*20*12;    y2_b2b = 30*23*12;   y3_b2b = 105*25*12
y1_ref = 3*140;       y2_ref = 20*140;      y3_ref = 50*140
y1_prj = 0;           y2_prj = 5*160;       y3_prj = 15*160
y1_rev = y1_b2b+y1_ref+y1_prj
y2_rev = y2_b2b+y2_ref+y2_prj
y3_rev = y3_b2b+y3_ref+y3_prj

# Cost assumptions
y1_worker = round(y1_rev*0.09); y2_worker = round(y2_rev*0.08); y3_worker = round(y3_rev*0.08)
y1_ai = 13; y2_ai = 32; y3_ai = 72
y1_cogs = y1_worker+y1_ai;  y2_cogs = y2_worker+y2_ai;  y3_cogs = y3_worker+y3_ai

y1_gp = y1_rev-y1_cogs;  y2_gp = y2_rev-y2_cogs;  y3_gp = y3_rev-y3_cogs

# SGA: FDE + CS + Admin + Founder + Dev + Marketing + Ops
y1_fde_sal = 2*45*12;  y2_fde_sal = 9*42*12;  y3_fde_sal = 27*42*12
y1_cs_sal  = 1*40*12;  y2_cs_sal  = 3*40*12;  y3_cs_sal  = 8*40*12
y1_adm     = 0;         y2_adm     = 2*35*12;  y3_adm     = 6*35*12
y1_founder = 360;       y2_founder = 600;       y3_founder = 840
y1_dev     = 120;       y2_dev     = 120;       y3_dev     = 120   # expense portion
y1_rec     = 60;        y2_rec     = 150;       y3_rec     = 300
y1_mktg    = 60;        y2_mktg    = 180;       y3_mktg    = 480
y1_office  = 24;        y2_office  = 60;        y3_office  = 144
y1_legal   = 36;        y2_legal   = 60;        y3_legal   = 96
y1_sys     = 36;        y2_sys     = 60;        y3_sys     = 96
y1_depr    = 0;         y2_depr    = 40;        y3_depr    = 120

y1_sga = y1_fde_sal+y1_cs_sal+y1_adm+y1_founder+y1_dev+y1_rec+y1_mktg+y1_office+y1_legal+y1_sys+y1_depr
y2_sga = y2_fde_sal+y2_cs_sal+y2_adm+y2_founder+y2_dev+y2_rec+y2_mktg+y2_office+y2_legal+y2_sys+y2_depr
y3_sga = y3_fde_sal+y3_cs_sal+y3_adm+y3_founder+y3_dev+y3_rec+y3_mktg+y3_office+y3_legal+y3_sys+y3_depr

y1_op = y1_gp-y1_sga;  y2_op = y2_gp-y2_sga;  y3_op = y3_gp-y3_sga
y1_tax = max(0,round(y1_op*0.30)); y2_tax = max(0,round(y2_op*0.30)); y3_tax = max(0,round(y3_op*0.30))
y1_ni = y1_op-y1_tax;  y2_ni = y2_op-y2_tax;  y3_ni = y3_op-y3_tax

row = 3
# ── REVENUE ──
section_title(ws2, row, 1, "■ 売上高", width=5); row+=1

write_label(ws2, row, 1, "B2B月次顧問料", bold=False)
write_number(ws2, row, 2, y1_b2b); write_number(ws2, row, 3, y2_b2b); write_number(ws2, row, 4, y3_b2b); write_number(ws2, row, 5, y1_b2b+y2_b2b+y3_b2b); row+=1

write_label(ws2, row, 1, "  ├ ネット平均契約社数（社）")
write_number(ws2, row, 2, 5); write_number(ws2, row, 3, 30); write_number(ws2, row, 4, 105); row+=1
write_label(ws2, row, 1, "  └ 月次チャーン率（前提）")
write_pct(ws2, row, 2, 0.025); write_pct(ws2, row, 3, 0.025); write_pct(ws2, row, 4, 0.020); row+=1

write_label(ws2, row, 1, "人材紹介フィー")
write_number(ws2, row, 2, y1_ref); write_number(ws2, row, 3, y2_ref); write_number(ws2, row, 4, y3_ref); write_number(ws2, row, 5, y1_ref+y2_ref+y3_ref); row+=1

write_label(ws2, row, 1, "AIプロジェクト受託")
write_number(ws2, row, 2, y1_prj); write_number(ws2, row, 3, y2_prj); write_number(ws2, row, 4, y3_prj); write_number(ws2, row, 5, y1_prj+y2_prj+y3_prj); row+=1

write_label(ws2, row, 1, "売上高合計", bold=True)
write_number(ws2, row, 2, y1_rev, bold=True, fill_color=NAVY); ws2.cell(row=row,column=2).font=hdr_font()
write_number(ws2, row, 3, y2_rev, bold=True, fill_color=NAVY); ws2.cell(row=row,column=3).font=hdr_font()
write_number(ws2, row, 4, y3_rev, bold=True, fill_color=NAVY); ws2.cell(row=row,column=4).font=hdr_font()
write_number(ws2, row, 5, y1_rev+y2_rev+y3_rev, bold=True, fill_color=NAVY); ws2.cell(row=row,column=5).font=hdr_font()
row+=2

# ── COGS ──
section_title(ws2, row, 1, "■ 売上原価（COGS）", width=5); row+=1

write_label(ws2, row, 1, "副業ワーカー費（売上の8〜9%）")
write_number(ws2, row, 2, y1_worker); write_number(ws2, row, 3, y2_worker); write_number(ws2, row, 4, y3_worker); write_number(ws2, row, 5, y1_worker+y2_worker+y3_worker); row+=1

write_label(ws2, row, 1, "★ UNLID内部AI・インフラ費（★新設）")
write_number(ws2, row, 2, y1_ai, fill_color=LBLUE); write_number(ws2, row, 3, y2_ai, fill_color=LBLUE); write_number(ws2, row, 4, y3_ai, fill_color=LBLUE); write_number(ws2, row, 5, y1_ai+y2_ai+y3_ai, fill_color=LBLUE)
row+=1

write_label(ws2, row, 1, "売上原価合計", bold=True)
write_number(ws2, row, 2, y1_cogs, bold=True); write_number(ws2, row, 3, y2_cogs, bold=True); write_number(ws2, row, 4, y3_cogs, bold=True); write_number(ws2, row, 5, y1_cogs+y2_cogs+y3_cogs, bold=True); row+=1

write_label(ws2, row, 1, "売上総利益（粗利）", bold=True)
write_number(ws2, row, 2, y1_gp, bold=True, fill_color=LBLUE); write_number(ws2, row, 3, y2_gp, bold=True, fill_color=LBLUE); write_number(ws2, row, 4, y3_gp, bold=True, fill_color=LBLUE); write_number(ws2, row, 5, y1_gp+y2_gp+y3_gp, bold=True, fill_color=LBLUE); row+=1

write_label(ws2, row, 1, "粗利率")
write_pct(ws2, row, 2, y1_gp/y1_rev); write_pct(ws2, row, 3, y2_gp/y2_rev); write_pct(ws2, row, 4, y3_gp/y3_rev); row+=2

# ── SGA ──
section_title(ws2, row, 1, "■ 販売費・一般管理費（SGA）", width=5); row+=1

items_sga = [
    ("★ FDE人件費（ヘッドカウント連動）", y1_fde_sal, y2_fde_sal, y3_fde_sal, True, LBLUE),
    ("  └ FDE人数（人）", 2, 9, 27, False, None),
    ("★ CSエンジニア人件費（プロジェクト連動）", y1_cs_sal, y2_cs_sal, y3_cs_sal, True, LBLUE),
    ("  └ CS人数（人）", 1, 3, 8, False, None),
    ("管理・営業スタッフ人件費", y1_adm, y2_adm, y3_adm, False, None),
    ("代表報酬", y1_founder, y2_founder, y3_founder, False, None),
    ("開発費（費用処理分）", y1_dev, y2_dev, y3_dev, False, None),
    ("★ 減価償却費（ソフトウェア）", y1_depr, y2_depr, y3_depr, False, AMBER),
    ("採用費", y1_rec, y2_rec, y3_rec, False, None),
    ("マーケティング費", y1_mktg, y2_mktg, y3_mktg, False, None),
    ("オフィス・通信費", y1_office, y2_office, y3_office, False, None),
    ("法務・会計費", y1_legal, y2_legal, y3_legal, False, None),
    ("システム・ツール費", y1_sys, y2_sys, y3_sys, False, None),
]

for label, v1, v2, v3, bold, fill in items_sga:
    write_label(ws2, row, 1, label, bold=bold)
    write_number(ws2, row, 2, v1, bold=bold, fill_color=fill)
    write_number(ws2, row, 3, v2, bold=bold, fill_color=fill)
    write_number(ws2, row, 4, v3, bold=bold, fill_color=fill)
    write_number(ws2, row, 5, v1+v2+v3, bold=bold, fill_color=fill)
    row+=1

write_label(ws2, row, 1, "SGA合計", bold=True)
write_number(ws2, row, 2, y1_sga, bold=True); write_number(ws2, row, 3, y2_sga, bold=True); write_number(ws2, row, 4, y3_sga, bold=True); write_number(ws2, row, 5, y1_sga+y2_sga+y3_sga, bold=True); row+=2

# ── PROFIT ──
section_title(ws2, row, 1, "■ 利益サマリー", width=5); row+=1

for label, v1, v2, v3, fill in [
    ("営業利益", y1_op, y2_op, y3_op, None),
    ("営業利益率", None, None, None, None),
    ("法人税等（30%）", -y1_tax, -y2_tax, -y3_tax, None),
    ("当期純利益", y1_ni, y2_ni, y3_ni, LBLUE),
]:
    write_label(ws2, row, 1, label, bold=(label in ["営業利益","当期純利益"]))
    if label == "営業利益率":
        write_pct(ws2, row, 2, y1_op/y1_rev); write_pct(ws2, row, 3, y2_op/y2_rev if y2_rev else 0); write_pct(ws2, row, 4, y3_op/y3_rev)
    else:
        write_number(ws2, row, 2, v1, bold=(fill is not None), fill_color=fill)
        write_number(ws2, row, 3, v2, bold=(fill is not None), fill_color=fill)
        write_number(ws2, row, 4, v3, bold=(fill is not None), fill_color=fill)
        write_number(ws2, row, 5, (v1+v2+v3) if v1 is not None else None, bold=(fill is not None), fill_color=fill)
    row+=1

print("Sheet 2 PL done. Row:", row)
# Store values for later sheets
fin = {
    "y1_rev": y1_rev, "y2_rev": y2_rev, "y3_rev": y3_rev,
    "y1_ni": y1_ni,   "y2_ni": y2_ni,   "y3_ni": y3_ni,
    "y1_op": y1_op,   "y2_op": y2_op,   "y3_op": y3_op,
    "y1_depr": y1_depr, "y2_depr": y2_depr, "y3_depr": y3_depr,
    "y1_worker": y1_worker, "y2_worker": y2_worker, "y3_worker": y3_worker,
}

# ============================================================
# SHEET 3: BS（貸借対照表）
# ============================================================
ws3 = wb.create_sheet("BS（貸借対照表）")
ws3.column_dimensions["A"].width = 36
for col in ["B","C","D","E"]:
    ws3.column_dimensions[col].width = 18

ws3.merge_cells("A1:E1")
c = ws3["A1"]
c.value = "UNLID 貸借対照表（B/S）— 運転資本・ソフトウェア資産計上版"
c.fill = header_fill(NAVY); c.font = Font(name="Meiryo UI", bold=True, color=WHITE, size=13)
c.alignment = Alignment(horizontal="center", vertical="center")
ws3.row_dimensions[1].height = 28

for col, txt in enumerate(["項目", "Year0 (期首)", "Year1末", "Year2末", "Year3末"], 1):
    write_header(ws3, 2, col, txt)

# Cash assumptions
equity_y0 = 500; equity_y1_add = 1500; equity_y2_add = 5000; equity_y3_add = 0
loan_y1 = 1000; loan_y2 = 0; loan_y3 = 0
loan_repay = 200  # per year

cash_y0 = equity_y0
cash_y1 = cash_y0 + equity_y1_add + loan_y1 + y1_ni + y1_depr - round(y1_rev*45/365) - loan_repay + round(y1_worker*30/365)
cash_y2 = cash_y1 + equity_y2_add + loan_y2 + y2_ni + y2_depr - (round(y2_rev*45/365)-round(y1_rev*45/365)) - loan_repay + (round(y2_worker*30/365)-round(y1_worker*30/365))
cash_y3 = cash_y2 + equity_y3_add + loan_y3 + y3_ni + y3_depr - (round(y3_rev*45/365)-round(y2_rev*45/365)) - loan_repay + (round(y3_worker*30/365)-round(y2_worker*30/365))

ar_y0=0; ar_y1=round(y1_rev*45/365); ar_y2=round(y2_rev*45/365); ar_y3=round(y3_rev*45/365)
sw_y0=0; sw_y1=0; sw_y2=200; sw_y3=560
ap_y0=0; ap_y1=round(y1_worker*30/365); ap_y2=round(y2_worker*30/365); ap_y3=round(y3_worker*30/365)
acc_y0=0; acc_y1=round(y1_sga/12*1); acc_y2=round(y2_sga/12*1); acc_y3=round(y3_sga/12*1)
loan_bal = [0, loan_y1-loan_repay, loan_y1-loan_repay*2, max(0,loan_y1-loan_repay*3)]

ta_y0 = cash_y0+ar_y0+sw_y0; ta_y1 = cash_y1+ar_y1+sw_y1; ta_y2 = cash_y2+ar_y2+sw_y2; ta_y3 = cash_y3+ar_y3+sw_y3
tl_y0 = ap_y0+acc_y0+loan_bal[0]; tl_y1 = ap_y1+acc_y1+loan_bal[1]; tl_y2 = ap_y2+acc_y2+loan_bal[2]; tl_y3 = ap_y3+acc_y3+loan_bal[3]
cap_y0=equity_y0; cap_y1=equity_y0+equity_y1_add; cap_y2=cap_y1+equity_y2_add; cap_y3=cap_y2
re_y0=0; re_y1=y1_ni; re_y2=y1_ni+y2_ni; re_y3=y1_ni+y2_ni+y3_ni
te_y0=cap_y0+re_y0; te_y1=cap_y1+re_y1; te_y2=cap_y2+re_y2; te_y3=cap_y3+re_y3

row=3
section_title(ws3, row, 1, "■ 資産の部", width=5); row+=1
for label, v0, v1, v2, v3, fill in [
    ("現金・預金（★運転資本タイムラグ反映後）", cash_y0, cash_y1, cash_y2, cash_y3, LBLUE),
    ("売掛金（DSO45日分）", ar_y0, ar_y1, ar_y2, ar_y3, AMBER),
    ("前払費用", 0, 10, 20, 40, None),
    ("★ ソフトウェア（無形固定資産）", sw_y0, sw_y1, sw_y2, sw_y3, LBLUE),
    ("有形固定資産（PC・備品等）", 20, 30, 60, 120, None),
    ("資産合計", ta_y0, ta_y1, ta_y2, ta_y3, None),
]:
    bold = "合計" in label
    write_label(ws3, row, 1, label, bold=bold)
    write_number(ws3, row, 2, v0, bold=bold, fill_color=fill if fill else None)
    write_number(ws3, row, 3, v1, bold=bold, fill_color=fill if fill else None)
    write_number(ws3, row, 4, v2, bold=bold, fill_color=fill if fill else None)
    write_number(ws3, row, 5, v3, bold=bold, fill_color=fill if fill else None)
    row+=1

row+=1
section_title(ws3, row, 1, "■ 負債の部", width=5); row+=1
for label, v0, v1, v2, v3, fill in [
    ("買掛金（ワーカー費 DPO30日分）", ap_y0, ap_y1, ap_y2, ap_y3, AMBER),
    ("未払費用（1ヶ月分SGA）", acc_y0, acc_y1, acc_y2, acc_y3, None),
    ("政策公庫借入（長期）", loan_bal[0], loan_bal[1], loan_bal[2], loan_bal[3], None),
    ("負債合計", tl_y0, tl_y1, tl_y2, tl_y3, None),
]:
    bold = "合計" in label
    write_label(ws3, row, 1, label, bold=bold)
    write_number(ws3, row, 2, v0, bold=bold, fill_color=fill)
    write_number(ws3, row, 3, v1, bold=bold, fill_color=fill)
    write_number(ws3, row, 4, v2, bold=bold, fill_color=fill)
    write_number(ws3, row, 5, v3, bold=bold, fill_color=fill)
    row+=1

row+=1
section_title(ws3, row, 1, "■ 純資産の部", width=5); row+=1
for label, v0, v1, v2, v3 in [
    ("資本金・資本剰余金（累計）", cap_y0, cap_y1, cap_y2, cap_y3),
    ("繰越利益剰余金（累計純利益）", re_y0, re_y1, re_y2, re_y3),
    ("純資産合計", te_y0, te_y1, te_y2, te_y3),
]:
    bold = "合計" in label
    fill = LBLUE if bold else None
    write_label(ws3, row, 1, label, bold=bold)
    write_number(ws3, row, 2, v0, bold=bold, fill_color=fill)
    write_number(ws3, row, 3, v1, bold=bold, fill_color=fill)
    write_number(ws3, row, 4, v2, bold=bold, fill_color=fill)
    write_number(ws3, row, 5, v3, bold=bold, fill_color=fill)
    row+=1

row+=1
# Balance check
ws3.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
c = ws3.cell(row=row, column=1, value="★ 黒字倒産リスク注記：Year3売上急増時、売掛金が約710万円膨張。資金調達（Series A）でこのキャッシュギャップを吸収する設計")
c.fill = cell_fill(AMBER); c.font = body_font(bold=True, color="92400E")
c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
ws3.row_dimensions[row].height = 30

print("Sheet 3 BS done. Cash Y1:", cash_y1, "Y2:", cash_y2, "Y3:", cash_y3)

# ============================================================
# SHEET 4: CF（キャッシュフロー計算書）
# ============================================================
ws4 = wb.create_sheet("CF（キャッシュフロー）")
ws4.column_dimensions["A"].width = 40
for col in ["B","C","D","E"]:
    ws4.column_dimensions[col].width = 18

ws4.merge_cells("A1:E1")
c = ws4["A1"]
c.value = "UNLID キャッシュフロー計算書（間接法） — 運転資本タイムラグ・黒字倒産リスク可視化版"
c.fill = header_fill(NAVY); c.font = Font(name="Meiryo UI", bold=True, color=WHITE, size=12)
c.alignment = Alignment(horizontal="center", vertical="center")
ws4.row_dimensions[1].height = 28

for col, txt in enumerate(["項目", "Year1 (FY2026)", "Year2 (FY2027)", "Year3 (FY2028)", "3年累計"], 1):
    write_header(ws4, 2, col, txt)

# CF calculations
# Operating CF
cf_op_ni = [y1_ni, y2_ni, y3_ni]
cf_op_depr = [y1_depr, y2_depr, y3_depr]
ar_chg = [-(ar_y1-ar_y0), -(ar_y2-ar_y1), -(ar_y3-ar_y2)]
ap_chg = [ap_y1-ap_y0, ap_y2-ap_y1, ap_y3-ap_y2]
acc_chg = [acc_y1-acc_y0, acc_y2-acc_y1, acc_y3-acc_y2]
cf_op = [cf_op_ni[i]+cf_op_depr[i]+ar_chg[i]+ap_chg[i]+acc_chg[i] for i in range(3)]

# Investing CF
sw_inv = [0, -240, -480]   # software capitalized
capex = [-30, -60, -120]
cf_inv = [sw_inv[i]+capex[i] for i in range(3)]

# Financing CF
eq_in = [equity_y1_add, equity_y2_add, 0]
loan_in = [loan_y1, 0, 0]
loan_out = [-loan_repay, -loan_repay, -loan_repay]
cf_fin = [eq_in[i]+loan_in[i]+loan_out[i] for i in range(3)]

cf_net = [cf_op[i]+cf_inv[i]+cf_fin[i] for i in range(3)]

row=3
section_title(ws4, row, 1, "■ 営業活動によるCF（Operating CF）", width=5); row+=1

cf_items_op = [
    ("当期純利益", cf_op_ni, False, None),
    ("（加算）減価償却費", cf_op_depr, False, LBLUE),
    ("★ 売掛金増減（DSO45日ラグ）", ar_chg, False, AMBER),
    ("  ↑ Y3売掛金が710万円増加→CFを圧迫（黒字倒産リスク）", None, False, RED),
    ("買掛金増減（ワーカーDPO30日）", ap_chg, False, None),
    ("未払費用増減", acc_chg, False, None),
    ("営業CF合計", cf_op, True, LBLUE),
]

for label, vals, bold, fill in cf_items_op:
    write_label(ws4, row, 1, label, bold=bold)
    if vals is None:
        ws4.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
        c = ws4.cell(row=row, column=2, value="← 売上急増時は要注意。Series A調達タイミングと合わせて管理")
        c.fill = cell_fill(RED); c.font = body_font(bold=True, color="991B1B")
        c.alignment = Alignment(horizontal="left")
    else:
        write_number(ws4, row, 2, vals[0], bold=bold, fill_color=fill)
        write_number(ws4, row, 3, vals[1], bold=bold, fill_color=fill)
        write_number(ws4, row, 4, vals[2], bold=bold, fill_color=fill)
        write_number(ws4, row, 5, sum(vals), bold=bold, fill_color=fill)
    row+=1

row+=1
section_title(ws4, row, 1, "■ 投資活動によるCF（Investing CF）", width=5); row+=1

for label, vals, bold, fill in [
    ("★ ソフトウェア開発費（資産計上分）", sw_inv, False, LBLUE),
    ("設備投資（PC・備品）", capex, False, None),
    ("投資CF合計", cf_inv, True, None),
]:
    write_label(ws4, row, 1, label, bold=bold)
    write_number(ws4, row, 2, vals[0], bold=bold, fill_color=fill)
    write_number(ws4, row, 3, vals[1], bold=bold, fill_color=fill)
    write_number(ws4, row, 4, vals[2], bold=bold, fill_color=fill)
    write_number(ws4, row, 5, sum(vals), bold=bold, fill_color=fill)
    row+=1

row+=1
section_title(ws4, row, 1, "■ 財務活動によるCF（Financing CF）", width=5); row+=1

for label, vals, bold, fill in [
    ("株式発行による調達", eq_in, False, None),
    ("政策公庫借入", loan_in, False, None),
    ("借入返済", loan_out, False, None),
    ("財務CF合計", cf_fin, True, None),
]:
    write_label(ws4, row, 1, label, bold=bold)
    write_number(ws4, row, 2, vals[0], bold=bold, fill_color=fill)
    write_number(ws4, row, 3, vals[1], bold=bold, fill_color=fill)
    write_number(ws4, row, 4, vals[2], bold=bold, fill_color=fill)
    write_number(ws4, row, 5, sum(vals), bold=bold, fill_color=fill)
    row+=1

row+=1
section_title(ws4, row, 1, "■ キャッシュサマリー", width=5); row+=1

cash_beg = [cash_y0, cash_y1, cash_y2]
cash_end = [cash_y1, cash_y2, cash_y3]

for label, vals, bold, fill in [
    ("期首現金", cash_beg, False, None),
    ("当期CF増減合計", cf_net, True, None),
    ("期末現金残高", cash_end, True, LBLUE),
    ("★ 必要運転資金（参考）", [round(y1_rev*45/365), round(y2_rev*45/365), round(y3_rev*45/365)], False, AMBER),
]:
    write_label(ws4, row, 1, label, bold=bold)
    write_number(ws4, row, 2, vals[0], bold=bold, fill_color=fill)
    write_number(ws4, row, 3, vals[1], bold=bold, fill_color=fill)
    write_number(ws4, row, 4, vals[2], bold=bold, fill_color=fill)
    if label not in ["期首現金","★ 必要運転資金（参考）"]:
        write_number(ws4, row, 5, sum(vals) if label!="期末現金残高" else vals[2], bold=bold, fill_color=fill)
    row+=1

print("Sheet 4 CF done.")

# ============================================================
# SHEET 5: 感度分析（強化版）
# ============================================================
ws5 = wb.create_sheet("感度分析（強化版）")
ws5.column_dimensions["A"].width = 32

ws5.merge_cells("A1:J1")
c = ws5["A1"]
c.value = "UNLID 感度分析 — チャーン・期ズレ・CAC悪化シナリオ追加版"
c.fill = header_fill(NAVY); c.font = Font(name="Meiryo UI", bold=True, color=WHITE, size=13)
c.alignment = Alignment(horizontal="center", vertical="center")
ws5.row_dimensions[1].height = 28

row = 3

# ── Analysis 1: チャーン率感度 ──
section_title(ws5, row, 1, "【分析①】チャーン率 × Year2期末契約社数（★新設）", width=10); row+=1
write_header(ws5, row, 1, "月次チャーン率 →", color=GRAY2, fgcolor=NAVY)
churn_rates = [0.010, 0.020, 0.025, 0.030, 0.040, 0.050]
company_gross = [55, 55, 55, 55, 55, 55]  # gross new acquisitions (fixed)
for i, cr in enumerate(churn_rates):
    write_header(ws5, row, i+2, f"{cr*100:.1f}%/月", color=BLUE if cr==0.025 else GRAY2, fgcolor=WHITE if cr==0.025 else NAVY)
row+=1

write_label(ws5, row, 1, "Y2期末ネット社数（グロス新規55社固定）", bold=True)
for i, cr in enumerate(churn_rates):
    # Simple model: 10 start, each month add gross/12, lose current*cr
    companies = 10
    monthly_gross = 55/12
    for m in range(12):
        companies = companies*(1-cr) + monthly_gross
    net = round(companies)
    fill = LBLUE if cr==0.025 else None
    write_number(ws5, row, i+2, net, bold=(cr==0.025), fill_color=fill)
row+=2

# ── Analysis 2: 期ズレシナリオ（★新設） ──
section_title(ws5, row, 1, "【分析②】期ズレ（タイムライン遅延）シナリオ（★新設）", width=10); row+=1
headers2 = ["項目", "Base（計画通り）", "期ズレ6ヶ月（遅延）", "期ズレ12ヶ月（1年遅延）"]
for i, h in enumerate(headers2):
    color = BLUE if i==1 else (RED if i>1 else NAVY)
    write_header(ws5, row, i+1, h, color=color)
row+=1

delay_items = [
    ("Year2実効売上（万円）", y2_rev, round(y2_rev*0.55), round(y2_rev*0.35)),
    ("Year2固定費（変化なし、万円）", y2_sga, y2_sga, y2_sga),
    ("Year2営業利益（万円）", y2_op, round(y2_rev*0.55) - round(y2_rev*0.55*0.09) - y2_sga,
        round(y2_rev*0.35) - round(y2_rev*0.35*0.09) - y2_sga),
    ("Year2末現金残高推定（万円）", cash_y2, cash_y2 - round((y2_rev-y2_rev*0.55)*0.5), cash_y2 - round((y2_rev-y2_rev*0.35)*0.5)),
    ("キャッシュショートの有無", "問題なし", "要注意（Series A前倒しを検討）", "ショートリスク高（追加資金調達必須）"),
]

for label, v_base, v_6m, v_12m in delay_items:
    write_label(ws5, row, 1, label, bold=("利益" in label or "現金" in label))
    if isinstance(v_base, str):
        for col, val, fill in [(2,v_base,LBLUE),(3,v_6m,AMBER),(4,v_12m,RED)]:
            c = ws5.cell(row=row, column=col, value=val)
            c.fill = cell_fill(fill); c.font = body_font(bold=True)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = thin_border()
        ws5.row_dimensions[row].height = 28
    else:
        write_number(ws5, row, 2, v_base, fill_color=LBLUE)
        write_number(ws5, row, 3, v_6m, fill_color=AMBER)
        write_number(ws5, row, 4, v_12m, fill_color=RED)
    row+=1
row+=1

# ── Analysis 3: CAC悪化シナリオ（★新設） ──
section_title(ws5, row, 1, "【分析③】CAC悪化 × LTV/CAC ユニットエコノミクス（★新設）", width=10); row+=1
write_header(ws5, row, 1, "CACシナリオ →", color=GRAY2, fgcolor=NAVY)
cac_scenarios = [
    ("CAC 15万\n（Y1水準）", 15),
    ("CAC 25万\n（Y2想定）", 25),
    ("CAC 40万\n（Y3想定）", 40),
    ("CAC 50万\n（1.5x悪化）", 50),
    ("CAC 80万\n（2x悪化）", 80),
]
for i, (label, cac) in enumerate(cac_scenarios):
    color = BLUE if cac in [25,40] else (RED if cac>=50 else GRAY2)
    fgc = WHITE if cac in [25,40] else NAVY
    write_header(ws5, row, i+2, label, color=color, fgcolor=fgc)
    ws5.column_dimensions[get_column_letter(i+2)].width = 16
row+=1

ltv = 380
for metric, calc_fn in [
    ("LTV（万円、固定）", lambda cac: ltv),
    ("CAC（万円）", lambda cac: cac),
    ("LTV/CAC倍率", lambda cac: round(ltv/cac,1)),
    ("CAC回収月数（月額粗利10万前提）", lambda cac: round(cac/10,1)),
    ("ユニットエコノミクス判定", lambda cac: "◎ 優秀(3x+)" if ltv/cac>=3 else ("○ 許容(2-3x)" if ltv/cac>=2 else "✕ 要改善(2x未満)")),
]:
    write_label(ws5, row, 1, metric, bold=("判定" in metric or "倍率" in metric))
    for i, (_, cac) in enumerate(cac_scenarios):
        val = calc_fn(cac)
        if isinstance(val, str):
            fill = LBLUE if "◎" in val else (AMBER if "○" in val else RED)
            c = ws5.cell(row=row, column=i+2, value=val)
            c.fill = cell_fill(fill); c.font = body_font(bold=True)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = thin_border()
        elif "倍率" in metric:
            fill = LBLUE if val>=3 else (AMBER if val>=2 else RED)
            write_number(ws5, row, i+2, val, fill_color=fill, fmt="#,##0.0")
        elif "月数" in metric:
            write_number(ws5, row, i+2, val, fmt="#,##0.0")
        else:
            write_number(ws5, row, i+2, val)
    row+=1
row+=2

# ── Analysis 4: 3シナリオ 営業利益（Base/Bull/Bear） ──
section_title(ws5, row, 1, "【分析④】3シナリオ P&L（Base / Bull / Bear）", width=10); row+=1
write_header(ws5, row, 1, "項目")
write_header(ws5, row, 2, "Bear（下振れ）", color=RED)
write_header(ws5, row, 3, "Base（計画）", color=BLUE)
write_header(ws5, row, 4, "Bull（上振れ）", color="16A34A")
row+=1

scenarios = {
    "Bear": {"rev_mult": 0.7, "cost_mult": 1.1, "label_fill": RED},
    "Base": {"rev_mult": 1.0, "cost_mult": 1.0, "label_fill": LBLUE},
    "Bull": {"rev_mult": 1.3, "cost_mult": 0.95, "label_fill": "DCFCE7"},
}

for year_label, rev_base, cost_base in [
    ("Year2売上（万円）", y2_rev, y2_sga+y2_cogs),
    ("Year2営業利益（万円）", y2_op, None),
    ("Year3売上（万円）", y3_rev, y3_sga+y3_cogs),
    ("Year3営業利益（万円）", y3_op, None),
]:
    write_label(ws5, row, 1, year_label, bold=("利益" in year_label))
    for i, (sc_name, sc) in enumerate(scenarios.items()):
        if "売上" in year_label:
            val = round(rev_base * sc["rev_mult"])
        else:
            rev = round(rev_base * sc["rev_mult"] if "Year2" in year_label else y3_rev * sc["rev_mult"])
            rev_b = y2_rev if "Year2" in year_label else y3_rev
            cost = round((y2_sga+y2_cogs if "Year2" in year_label else y3_sga+y3_cogs) * sc["cost_mult"])
            cogs_b = y2_cogs if "Year2" in year_label else y3_cogs
            val = round(rev_base*sc["rev_mult"]) - round((y2_sga+y2_cogs if "Year2" in year_label else y3_sga+y3_cogs)*sc["cost_mult"])
        fill = sc["label_fill"]
        write_number(ws5, row, i+2, val, fill_color=fill, bold=("利益" in year_label))
    row+=1

print("Sheet 5 Sensitivity done.")

# ============================================================
# SHEET 6: チャーン詳細分析（月次モデル）★新設
# ============================================================
ws6 = wb.create_sheet("チャーン月次モデル（★新設）")
ws6.column_dimensions["A"].width = 28
for i in range(2, 38):
    ws6.column_dimensions[get_column_letter(i)].width = 10

ws6.merge_cells("A1:AK1")
c = ws6["A1"]
c.value = "チャーン月次モデル — グロス新規獲得 vs ネット増加の動態（3年36ヶ月）"
c.fill = header_fill(NAVY); c.font = Font(name="Meiryo UI", bold=True, color=WHITE, size=12)
c.alignment = Alignment(horizontal="center", vertical="center")
ws6.row_dimensions[1].height = 28

# Headers: Month 1-36
ws6.cell(row=2, column=1, value="項目").fill = header_fill(NAVY)
ws6.cell(row=2, column=1).font = hdr_font()
ws6.cell(row=2, column=1).border = thin_border()
ws6.cell(row=2, column=1).alignment = Alignment(horizontal="center")

# Year banners
for yr, start_col, end_col, label in [(1,2,13,"Year1（FY2026）"),(2,14,25,"Year2（FY2027）"),(3,26,37,"Year3（FY2028）")]:
    ws6.merge_cells(start_row=2, start_column=start_col, end_row=2, end_column=end_col)
    c = ws6.cell(row=2, column=start_col, value=label)
    c.fill = header_fill(BLUE if yr==2 else (NAVY if yr==1 else "1E40AF"))
    c.font = hdr_font(size=11); c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = thin_border()

for m in range(36):
    c = ws6.cell(row=3, column=m+2, value=f"M{m+1}")
    c.fill = header_fill(GRAY2); c.font = hdr_font(NAVY); c.alignment = Alignment(horizontal="center")
    c.border = thin_border()

# Model parameters
churn_rate_y1 = 0.025
churn_rate_y2 = 0.025
churn_rate_y3 = 0.020
# Target net: Y1=10, Y2=50, Y3=160
# Work backwards: monthly gross needed
gross_per_month = [0.92, 0.92, 0.92, 0.92, 0.92, 0.92, 0.92, 0.92, 0.92, 0.92, 0.92, 0.92,  # Y1: 11/12
                   4.08, 4.08, 4.08, 4.08, 4.08, 4.08, 4.08, 4.08, 4.08, 4.08, 4.08, 4.08,  # Y2: 49/12
                   11.1, 11.1, 11.1, 11.1, 11.1, 11.1, 11.1, 11.1, 11.1, 11.1, 11.1, 11.1]  # Y3: 133/12

churn_rate_by_month = [churn_rate_y1]*12 + [churn_rate_y2]*12 + [churn_rate_y3]*12

# Simulate
companies = 0
month_data = {"total": [], "gross_new": [], "churned": [], "net_change": [], "b2b_rev": []}
unit_price = [20]*12 + [23]*12 + [25]*12

for m in range(36):
    cr = churn_rate_by_month[m]
    churned = round(companies * cr, 1)
    gross = gross_per_month[m]
    net = gross - churned
    companies = max(0, companies + net)
    month_data["total"].append(round(companies, 1))
    month_data["gross_new"].append(round(gross, 1))
    month_data["churned"].append(round(churned, 1))
    month_data["net_change"].append(round(net, 1))
    month_data["b2b_rev"].append(round(companies * unit_price[m], 0))

row = 4
for label, key, bold, fill in [
    ("契約社数（累計）", "total", True, LBLUE),
    ("グロス新規獲得（社/月）", "gross_new", False, None),
    ("解約数（チャーン）", "churned", False, RED),
    ("ネット増減", "net_change", False, None),
    ("B2B月次売上（万円）", "b2b_rev", True, None),
]:
    c = ws6.cell(row=row, column=1, value=label)
    c.font = body_font(bold=bold); c.border = thin_border()
    c.alignment = Alignment(horizontal="left", vertical="center")
    for m in range(36):
        val = month_data[key][m]
        cell = ws6.cell(row=row, column=m+2, value=val)
        cell.border = thin_border()
        cell.alignment = Alignment(horizontal="right")
        cell.font = body_font(bold=bold)
        if fill == RED and val > 0:
            cell.fill = cell_fill(RED)
        elif fill:
            cell.fill = cell_fill(fill)
        if key == "b2b_rev":
            cell.number_format = "#,##0"
    row += 1

row += 1
# Summary row
ws6.merge_cells(start_row=row, start_column=1, end_row=row, end_column=37)
note = ws6.cell(row=row, column=1, value=
    "【チャーン前提】月次2.5%（Y1-Y2）→ 2.0%（Y3）。年間チャーン率=約26%。"
    "グロス新規獲得（Y1:11社/Y2:49社/Y3:133社）がネット目標（10/50/160社）達成に必要な真の営業KPI。")
note.fill = cell_fill(AMBER); note.font = body_font(bold=True, color="92400E")
note.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
ws6.row_dimensions[row].height = 36

print("Sheet 6 Churn Model done.")

# ============================================================
# SHEET 7: SME単価市場調査（維持）
# ============================================================
ws7 = wb.create_sheet("SME単価市場調査")
ws7.column_dimensions["A"].width = 30
ws7.column_dimensions["B"].width = 22
ws7.column_dimensions["C"].width = 22
ws7.column_dimensions["D"].width = 30

ws7.merge_cells("A1:D1")
c = ws7["A1"]
c.value = "SME向け月額サービス 市場相場調査 & UNLID単価根拠"
c.fill = header_fill(NAVY); c.font = Font(name="Meiryo UI", bold=True, color=WHITE, size=13)
c.alignment = Alignment(horizontal="center", vertical="center")
ws7.row_dimensions[1].height = 28

for col, txt in enumerate(["サービス種別", "市場相場（月額）", "UNLID比較", "UNLID優位性"], 1):
    write_header(ws7, 2, col, txt)

market_data = [
    ("BPO/記帳代行", "3〜15万円", "割高", "AIによる自動化付加価値"),
    ("人事・労務コンサル", "10〜25万円", "同水準", "AI採用+完遂ログで差別化"),
    ("IT/DXアドバイザリー", "10〜30万円", "同水準〜やや割高", "SME向け実装まで完遂"),
    ("フラクショナルCTO", "30〜60万円", "割安（UNLID=20〜25万）", "成果責任付きで格安"),
    ("大手DXコンサル", "50万〜200万円", "圧倒的に割安", "SMEには不要な機能を省略"),
    ("AIプロジェクト導入", "30〜150万円（単発）", "同水準（月額継続が強み）", "月次継続でLTV大"),
    ("UNLID（FDE月次顧問）", "20〜25万円 ★", "—", "成果責任×副業育成×ログ×紹介"),
]

row = 3
for data in market_data:
    is_unlid = "UNLID" in data[0]
    fill = LBLUE if is_unlid else None
    for col, val in enumerate(data, 1):
        c = ws7.cell(row=row, column=col, value=val)
        c.font = body_font(bold=is_unlid)
        c.border = thin_border()
        c.alignment = Alignment(horizontal="left" if col in [1,4] else "center", vertical="center", wrap_text=True)
        if fill: c.fill = cell_fill(fill)
        if is_unlid: c.fill = cell_fill(LBLUE)
    ws7.row_dimensions[row].height = 22
    row += 1

row += 1
section_title(ws7, row, 1, "チャーン率の妥当性根拠（月額帯別）", width=4); row+=1
for col, txt in enumerate(["月額帯", "12ヶ月継続率", "月次チャーン率換算", "UNLID想定"], 1):
    write_header(ws7, row, col, txt, color=GRAY2, fgcolor=NAVY)
row+=1

churn_ref = [
    ("5〜10万円", "40〜55%", "5.0〜7.5%/月", "対象外"),
    ("10〜20万円", "50〜65%", "3.7〜5.8%/月", "参考"),
    ("20〜30万円（UNLID帯）", "65〜75%", "2.4〜3.5%/月", "★ Base=2.5%/月採用"),
    ("30〜60万円", "65〜75%", "2.4〜3.5%/月", "参考"),
]
for data in churn_ref:
    is_unlid = "UNLID帯" in data[0]
    fill = LBLUE if is_unlid else None
    for col, val in enumerate(data, 1):
        c = ws7.cell(row=row, column=col, value=val)
        c.font = body_font(bold=is_unlid)
        c.border = thin_border()
        c.alignment = Alignment(horizontal="center", vertical="center")
        if fill: c.fill = cell_fill(fill)
    row+=1

print("Sheet 7 Market done.")

# ============================================================
# SAVE
# ============================================================
output_path = "/home/user/UNLID/UNLID_財務三表_v4.0.xlsx"
wb.save(output_path)
print(f"\n✅ 保存完了: {output_path}")
print("Sheets:", [s.title for s in wb.worksheets])
