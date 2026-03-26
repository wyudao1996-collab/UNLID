# -*- coding: utf-8 -*-
"""UNLID 財務三表 v4.2 ─ Part 1
helpers + 前提条件シート + 人員計画シート
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# ───────────────── ヘルパー ─────────────────
def sc(ws, r, c, v=None, f=None, bold=False, italic=False, size=10,
       fc=None, align="left", nf=None, color="000000"):
    cell = ws.cell(row=r, column=c)
    cell.value = f if f else v
    cell.font = Font(name="Meiryo UI", bold=bold, italic=italic, size=size, color=color)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    if fc:
        cell.fill = PatternFill("solid", fgColor=fc)
    if nf:
        cell.number_format = nf
    return cell

def bdr(ws, r1, r2, c1, c2, style="thin"):
    s = Side(style=style)
    b = Border(left=s, right=s, top=s, bottom=s)
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(row=r, column=c).border = b

def gl(c): return get_column_letter(c)
def mrg(ws, r, c1, c2): ws.merge_cells(f"{gl(c1)}{r}:{gl(c2)}{r}")

def title_row(ws, r, text, ncols=6):
    sc(ws, r, 1, text, bold=True, size=13, fc=C_NAVY, color="FFFFFF", align="center")
    mrg(ws, r, 1, ncols)
    ws.row_dimensions[r].height = 26

def sec_hdr(ws, r, text, ncols=6):
    sc(ws, r, 1, text, bold=True, fc=C_BLUE, color="FFFFFF", align="left")
    mrg(ws, r, 1, ncols)
    ws.row_dimensions[r].height = 20

def note_row(ws, r, text, ncols=6):
    sc(ws, r, 1, text, italic=True, fc=C_NOTE, size=9)
    mrg(ws, r, 1, ncols)

def col_hdrs(ws, r, labels, fcs=None, fc_def="BDD7EE"):
    for i, lbl in enumerate(labels, 1):
        fc = (fcs[i-1] if fcs else None) or fc_def
        sc(ws, r, i, lbl, bold=True, fc=fc, align="center")

# ─── カラーパレット ───
C_NAVY  = "0F172A"
C_BLUE  = "2563EB"
C_LBLUE = "EFF6FF"
C_BHDR  = "DBEAFE"
C_TOT   = "FFF9C4"
C_GPFT  = "DCFCE7"
C_LOSS  = "FEE2E2"
C_NOTE  = "F8FAFC"
C_WARN  = "FEF3C7"
C_SUB   = "E0E7FF"
C_ORANGE= "FED7AA"
NF_MAN  = '#,##0'
NF_PCT  = '0.0%'
NF_2D   = '0.00'

# ══════════════════════════════════════════════
# SHEET 1: 前提条件
# ══════════════════════════════════════════════
ws_a = wb.active
ws_a.title = "前提条件"
ws_a.column_dimensions["A"].width = 36
for col in ["B","C","D"]:
    ws_a.column_dimensions[col].width = 16

title_row(ws_a, 1, "UNLID 財務モデル 前提条件 v4.2  ─  単位：万円  |  中央値シナリオ", 4)
sc(ws_a, 2, 1,
   "v4.2変更：COGS/SG&A完全展開 ／ 人員計画連動 ／ 運転資本サイト ／ 設備投資・償却 ／ 資金調達前提 追加",
   italic=True, fc=C_NOTE, size=9)
mrg(ws_a, 2, 1, 4)

# ─ セクション①: 売上ドライバー（v4.1踏襲）
sec_hdr(ws_a, 4, "① 売上ドライバー", 4)
col_hdrs(ws_a, 5, ["項目", "Year1", "Year2", "Year3"], fc_def=C_BHDR)

A1 = [
    (6,  "B2B契約企業数（期末・社）",            10,   50,  160, False),
    (7,  "FDE受託 月額単価（万円/社）",           20,   23,   25, False),
    (8,  "B2B受託収益（万円）",          "=B6*B7*12","=C6*C7*12","=D6*D7*12", True),
    (9,  "人材紹介成立件数（件）",                 3,   20,   50, False),
    (10, "人材紹介フィー単価（万円/件）",          140,  140,  140, False),
    (11, "人材紹介フィー合計（万円）",         "=B9*B10","=C9*C10","=D9*D10", True),
    (12, "受託プロジェクト件数（件）",              0,    5,   15, False),
    (13, "プロジェクト単価（万円/件・ベース）",     0,  160,  160, False),
    (14, "プロジェクト収益（万円）",       "=B12*B13","=C12*C13","=D12*D13", True),
    (15, "売上合計（万円）",      "=B8+B11+B14","=C8+C11+C14","=D8+D11+D14", True),
]
for row, lbl, v1, v2, v3, bold in A1:
    fc = C_TOT if "売上合計" in lbl else (C_GPFT if bold else None)
    sc(ws_a, row, 1, lbl, bold=bold, fc=fc)
    for c, v in [(2,v1),(3,v2),(4,v3)]:
        nf = NF_MAN
        if isinstance(v, str):
            sc(ws_a, row, c, f=v, bold=bold, fc=fc, nf=nf, align="center")
        else:
            sc(ws_a, row, c, v=v, bold=bold, fc=fc, nf=nf, align="center")
bdr(ws_a, 5, 15, 1, 4)
note_row(ws_a, 16, "※ B2Bは期末社数×単価×12ヶ月（期中平均で計算する場合は別途月次CFシート参照）", 4)

# ─ セクション②: COGS（売上原価）前提
sec_hdr(ws_a, 18, "② COGS（売上原価）前提  ─  サーバー/API費・外注費", 4)
col_hdrs(ws_a, 19, ["項目", "Year1", "Year2", "Year3"], fc_def=C_BHDR)

A2 = [
    (20, "AIサーバー・インフラ費（万円）",           24,   72,  180),
    (21, "外部API利用料（OpenAI等・万円）",          12,   60,  180),
    (22, "副業ワーカー報酬率（対B2B受託売上）",     0.09, 0.08, 0.08),
    (23, "副業ワーカー報酬（万円）",  "=B8*B22","=C8*C22","=D8*D22"),
    (24, "外注エンジニア費（プロジェクト分・万円）",  0,   200,  600),
    (25, "COGS合計（万円）", "=SUM(B20:B24)","=SUM(C20:C24)","=SUM(D20:D24)"),
    (26, "売上総利益（万円）",       "=B15-B25","=C15-C25","=D15-D25"),
    (27, "粗利率",                "=B26/B15","=C26/C15","=D26/D15"),
]
for row, lbl, v1, v2, v3 in A2:
    bold = ("合計" in lbl or "粗利" in lbl)
    fc = C_SUB if "COGS合計" in lbl else (C_GPFT if "売上総利益" in lbl else None)
    nf = NF_PCT if "率" in lbl else NF_MAN
    sc(ws_a, row, 1, lbl, bold=bold, fc=fc)
    for c, v in [(2,v1),(3,v2),(4,v3)]:
        if isinstance(v, str):
            sc(ws_a, row, c, f=v, bold=bold, fc=fc, nf=nf, align="center")
        else:
            sc(ws_a, row, c, v=v, bold=bold, fc=fc, nf=nf, align="center")
bdr(ws_a, 19, 27, 1, 4)

# ─ セクション③: SG&A前提（人員計画シートと連動）
sec_hdr(ws_a, 29, "③ SG&A（販売費及び一般管理費）前提  ─  人員計画シートと連動", 4)
col_hdrs(ws_a, 30, ["項目", "Year1", "Year2", "Year3"], fc_def=C_BHDR)

A3 = [
    (31, "人件費合計（万円）※人員計画シート参照",
         "=人員計画!D32", "=人員計画!E32", "=人員計画!F32"),
    (32, "採用費合計（万円）※人員計画シート参照",
         "=人員計画!D42", "=人員計画!E42", "=人員計画!F42"),
    (33, "B2Bマーケティング・CAC費（万円）",        120,  360,  600),
    (34, "人材紹介 候補者集客広告費（万円）",         30,  200,  500),
    (35, "地代家賃・オフィス費（万円）",              60,  180,  360),
    (36, "SaaSツール・システム費（万円）",            60,  120,  240),
    (37, "法務・許認可・顧問費（万円）",             200,  300,  400),
    (38, "その他固定費（万円）",                      60,  120,  200),
    (39, "SG&A合計（万円）",
         "=SUM(B31:B38)","=SUM(C31:C38)","=SUM(D31:D38)"),
    (40, "営業利益（万円）",     "=B26-B39","=C26-C39","=D26-D39"),
    (41, "営業利益率",           "=B40/B15","=C40/C15","=D40/D15"),
]
for row, lbl, v1, v2, v3 in A3:
    bold = ("合計" in lbl or "営業利益" in lbl)
    fc = C_SUB if "SG&A合計" in lbl else (C_TOT if "営業利益（万円）" in lbl else None)
    nf = NF_PCT if "率" in lbl else NF_MAN
    sc(ws_a, row, 1, lbl, bold=bold, fc=fc)
    for c, v in [(2,v1),(3,v2),(4,v3)]:
        if isinstance(v, str):
            sc(ws_a, row, c, f=v, bold=bold, fc=fc, nf=nf, align="center")
        else:
            sc(ws_a, row, c, v=v, bold=bold, fc=fc, nf=nf, align="center")
bdr(ws_a, 30, 41, 1, 4)

# ─ セクション④: 運転資本サイト（BS・CF連動）
sec_hdr(ws_a, 43, "④ 運転資本サイト（売掛金・買掛金）", 4)
col_hdrs(ws_a, 44, ["項目", "Year1", "Year2", "Year3"], fc_def=C_BHDR)

A4 = [
    (45, "売掛金 回収サイト（日）─ 月末締め翌月末払い", 30, 30, 30),
    (46, "買掛金 支払サイト（日）─ 月末締め翌月末払い", 30, 30, 30),
    (47, "期末売掛金残高（万円）=売上/12",
         "=ROUND(B15/12,0)","=ROUND(C15/12,0)","=ROUND(D15/12,0)"),
    (48, "期末買掛金残高（万円）=COGS/12",
         "=ROUND(B25/12,0)","=ROUND(C25/12,0)","=ROUND(D25/12,0)"),
]
for row, lbl, v1, v2, v3 in A4:
    nf = NF_MAN if "万円" in lbl else "#,##0"
    sc(ws_a, row, 1, lbl)
    for c, v in [(2,v1),(3,v2),(4,v3)]:
        if isinstance(v, str):
            sc(ws_a, row, c, f=v, nf=NF_MAN, align="center")
        else:
            sc(ws_a, row, c, v=v, nf=NF_MAN, align="center")
bdr(ws_a, 44, 48, 1, 4)

# ─ セクション⑤: 設備投資・ソフトウェア資産・償却
sec_hdr(ws_a, 50, "⑤ 設備投資・ソフトウェア資産・減価償却", 4)
col_hdrs(ws_a, 51, ["項目", "Year1", "Year2", "Year3"], fc_def=C_BHDR)

A5 = [
    (52, "自社システム開発費（資産計上・万円）",    100,  300,  800),
    (53, "償却期間（年）",                            3,    3,    3),
    (54, "当期償却費（万円）",
         "=ROUND(B52/B53,0)","=ROUND(B52/B53+C52/C53,0)","=ROUND(B52/B53+C52/C53+D52/D53,0)"),
    (55, "ソフトウェア資産残高（万円・期末）",
         "=B52-B54","=B52+C52-C54-B54","=B52+C52+D52-D54-C54-B54"),
]
for row, lbl, v1, v2, v3 in A5:
    nf = NF_MAN if "万円" in lbl else NF_MAN
    sc(ws_a, row, 1, lbl)
    for c, v in [(2,v1),(3,v2),(4,v3)]:
        if isinstance(v, str):
            sc(ws_a, row, c, f=v, nf=NF_MAN, align="center")
        else:
            sc(ws_a, row, c, v=v, nf=NF_MAN, align="center")
bdr(ws_a, 51, 55, 1, 4)

# ─ セクション⑥: 資金調達前提
sec_hdr(ws_a, 57, "⑥ 資金調達前提（エクイティ・デット）", 4)
col_hdrs(ws_a, 58, ["項目", "Year1", "Year2", "Year3"], fc_def=C_BHDR)

A6 = [
    (59, "期首資本金（払込済・万円）",              500,    0,    0),
    (60, "シード調達（エクイティ・万円）",            0, 3000,    0),
    (61, "シリーズA調達（万円）",                     0,    0, 10000),
    (62, "銀行借入（万円）",                        500,    0,    0),
    (63, "借入金利（年率）",                       0.02, 0.02, 0.02),
    (64, "支払利息（万円）",
         "=ROUND(B62*B63,0)","=ROUND(B62*C63,0)","=ROUND(B62*D63,0)"),
    (65, "税引前利益（万円）",  "=B40-B64","=C40-C64","=D40-D64"),
    (66, "法人税等（万円）30%", "=IF(B65>0,ROUND(B65*0.3,0),0)",
                                "=IF(C65>0,ROUND(C65*0.3,0),0)",
                                "=IF(D65>0,ROUND(D65*0.3,0),0)"),
    (67, "当期純利益（万円）",  "=B65-B66","=C65-C66","=D65-D66"),
]
for row, lbl, v1, v2, v3 in A6:
    bold = ("税引前利益" in lbl or "当期純利益" in lbl)
    fc = C_TOT if "当期純利益" in lbl else None
    nf = NF_PCT if "金利" in lbl else NF_MAN
    sc(ws_a, row, 1, lbl, bold=bold, fc=fc)
    for c, v in [(2,v1),(3,v2),(4,v3)]:
        if isinstance(v, str):
            sc(ws_a, row, c, f=v, bold=bold, fc=fc, nf=nf, align="center")
        else:
            sc(ws_a, row, c, v=v, bold=bold, fc=fc, nf=nf, align="center")
bdr(ws_a, 58, 67, 1, 4)

# ─ セクション⑦: 解約率（Churn）・CVR前提
sec_hdr(ws_a, 69, "⑦ 解約率（Churn）・成約率（CVR）前提  ─  感度分析シート連動", 4)
col_hdrs(ws_a, 70, ["項目", "Year1", "Year2", "Year3"], fc_def=C_BHDR)

A7 = [
    (71, "月次Churn率（B2B解約率）ベース",         0.02, 0.02, 0.015),
    (72, "年間解約予測社数（期初社数×年次換算）",
         "=ROUND(B6*B71*12,0)","=ROUND(C6*C71*12,0)","=ROUND(D6*D71*12,0)"),
    (73, "グロス新規獲得必要社数（純増＋解約補填）",
         "=B6+B72","=C6-B6+C72","=D6-C6+D72"),
    (74, "人材紹介 候補者充足率（CVR・ベース）",  0.15, 0.18, 0.20),
    (75, "プロジェクト案件 成約率（ベース）",     0.30, 0.35, 0.35),
]
for row, lbl, v1, v2, v3 in A7:
    nf = NF_PCT if "率" in lbl else NF_MAN
    sc(ws_a, row, 1, lbl)
    for c, v in [(2,v1),(3,v2),(4,v3)]:
        if isinstance(v, str):
            sc(ws_a, row, c, f=v, nf=nf, align="center")
        else:
            sc(ws_a, row, c, v=v, nf=nf, align="center")
bdr(ws_a, 70, 75, 1, 4)
note_row(ws_a, 76, "※ Churn率・CVRの感度分析（悲観/中央/楽観シナリオ別）は「感度分析」シート参照", 4)

# ══════════════════════════════════════════════
# SHEET 2: 人員計画
# ══════════════════════════════════════════════
ws_h = wb.create_sheet("人員計画")
ws_h.column_dimensions["A"].width = 28
for col in ["B","C","D","E","F","G","H","I"]:
    ws_h.column_dimensions[col].width = 14

title_row(ws_h, 1, "UNLID 人員計画（採用計画）v4.2  ─  採用数・人件費がPL・感度分析に自動連動", 8)
sc(ws_h, 2, 1,
   "★ この表の人件費合計・採用費合計が「前提条件」シートのSG&A（行31,32）に自動参照される",
   bold=True, fc=C_WARN, size=9)
mrg(ws_h, 2, 1, 8)

# ─ 役割定義テーブル
sec_hdr(ws_h, 4, "① 役職定義・単価テーブル（入力エリア）", 8)
col_hdrs(ws_h, 5,
    ["役職", "月額給与(万円)", "法定福利費率", "採用費単価(万円)", "備考"],
    fc_def=C_BHDR)

ROLES = [
    ("CEO / 創業者",           30,  0.15,   0, "役員報酬（Y1）、段階的引上げ"),
    ("FDE（フラクショナルFDE）",45,  0.15, 140, "正社員化。副業→内部登用でコスト圧縮"),
    ("CSM（カスタマーサクセス）",40, 0.15, 100, "1名が25〜30社を担当"),
    ("PM（プロジェクトMgr）",   50,  0.15, 150, "プロジェクト5件/人目安"),
    ("AIエンジニア",            60,  0.15, 200, "自社ツール・大型PJ担当"),
    ("Admin / BizDev",          35,  0.15,  80, "バックオフィス・採用支援"),
]
for i, (role, salary, welf, recruit, note) in enumerate(ROLES, 6):
    sc(ws_h, i, 1, role, bold=True)
    sc(ws_h, i, 2, salary, nf=NF_MAN, align="center")
    sc(ws_h, i, 3, welf, nf=NF_PCT, align="center")
    sc(ws_h, i, 4, recruit, nf=NF_MAN, align="center")
    sc(ws_h, i, 5, note, italic=True, fc=C_NOTE, size=9)
bdr(ws_h, 5, 11, 1, 5)

# ─ 採用スケジュールテーブル
sec_hdr(ws_h, 13, "② 採用スケジュール（期末時点ヘッドカウント）", 8)
col_hdrs(ws_h, 14,
    ["役職", "Y1期首", "Y1採用", "Y1期末", "Y2採用", "Y2期末", "Y3採用", "Y3期末"],
    fc_def=C_BHDR)

# (role, Y1_start, Y1_hire, Y2_hire, Y3_hire)
HIRE = [
    ("CEO / 創業者",           1, 0, 0, 0),
    ("FDE（フラクショナルFDE）",0, 2, 8,22),
    ("CSM（カスタマーサクセス）",0,0, 3, 5),
    ("PM（プロジェクトMgr）",   0, 0, 2, 4),
    ("AIエンジニア",            0, 1, 3, 5),
    ("Admin / BizDev",          0, 0, 1, 2),
]

# ヘッドカウントをsheet上の数式で計算
ROW_ROLE_START = 15
for i, (role, y1s, y1h, y2h, y3h) in enumerate(HIRE):
    r = ROW_ROLE_START + i
    y1e = y1s + y1h
    y2e = y1e + y2h
    y3e = y2e + y3h
    sc(ws_h, r, 1, role, bold=True)
    sc(ws_h, r, 2, y1s, nf=NF_MAN, align="center")      # Y1期首
    sc(ws_h, r, 3, y1h, nf=NF_MAN, align="center")      # Y1採用
    sc(ws_h, r, 4, y1e, bold=True, nf=NF_MAN, align="center", fc=C_GPFT)  # Y1期末
    sc(ws_h, r, 5, y2h, nf=NF_MAN, align="center")      # Y2採用
    sc(ws_h, r, 6, y2e, bold=True, nf=NF_MAN, align="center", fc=C_GPFT)  # Y2期末
    sc(ws_h, r, 7, y3h, nf=NF_MAN, align="center")      # Y3採用
    sc(ws_h, r, 8, y3e, bold=True, nf=NF_MAN, align="center", fc=C_GPFT)  # Y3期末

# 合計行
r_tot = ROW_ROLE_START + len(HIRE)
sc(ws_h, r_tot, 1, "合計（人）", bold=True, fc=C_NAVY, color="FFFFFF")
for c in range(2, 9):
    sc(ws_h, r_tot, c, f=f"=SUM({gl(c)}{ROW_ROLE_START}:{gl(c)}{r_tot-1})",
       bold=True, fc=C_TOT, nf=NF_MAN, align="center")
bdr(ws_h, 14, r_tot, 1, 8)

# ─ 年次人件費計算テーブル（→前提条件シートが参照するキーセル）
sec_hdr(ws_h, 24, "③ 年次人件費・採用費サマリー（→前提条件シートに自動連動）", 8)
col_hdrs(ws_h, 25, ["役職", "月額給与(万)", "法定福利率", "Y1人件費", "Y2人件費", "Y3人件費"], fc_def=C_BHDR)

# 各役職の年次人件費 = 給与 × (1+法定福利費率) × 12 × 期末人数
# 給与・福利費率はROLES、期末人数はHIREから
salary_map = {r[0]: (r[1], r[2]) for r in ROLES}
headcount_map = {
    role: (y1s+y1h, y1s+y1h+y2h, y1s+y1h+y2h+y3h)
    for role, y1s, y1h, y2h, y3h in HIRE
}

for i, (role, salary, welf, recruit, _) in enumerate(ROLES):
    r = 26 + i
    y1e, y2e, y3e = headcount_map[role]
    y1_cost = round(salary * (1 + welf) * 12 * y1e)
    y2_cost = round(salary * (1 + welf) * 12 * y2e)
    y3_cost = round(salary * (1 + welf) * 12 * y3e)
    sc(ws_h, r, 1, role)
    sc(ws_h, r, 2, salary, nf=NF_MAN, align="center")
    sc(ws_h, r, 3, welf, nf=NF_PCT, align="center")
    sc(ws_h, r, 4, y1_cost, nf=NF_MAN, align="center")
    sc(ws_h, r, 5, y2_cost, nf=NF_MAN, align="center")
    sc(ws_h, r, 6, y3_cost, nf=NF_MAN, align="center")

# 人件費合計行（Row 32 = ROW 26+6）
r_sal_tot = 26 + len(ROLES)
sc(ws_h, r_sal_tot, 1, "人件費合計（万円）", bold=True, fc=C_NAVY, color="FFFFFF")
for c in [4, 5, 6]:
    sc(ws_h, r_sal_tot, c, f=f"=SUM({gl(c)}26:{gl(c)}{r_sal_tot-1})",
       bold=True, fc=C_TOT, nf=NF_MAN, align="center")
bdr(ws_h, 25, r_sal_tot, 1, 6)

# 採用費計算
sec_hdr(ws_h, 34, "④ 採用費サマリー（採用人数 × 採用費単価）", 8)
col_hdrs(ws_h, 35, ["役職", "採用費単価(万)", "", "Y1採用費", "Y2採用費", "Y3採用費"], fc_def=C_BHDR)

recruit_map = {r[0]: r[3] for r in ROLES}
for i, (role, y1s, y1h, y2h, y3h) in enumerate(HIRE):
    r = 36 + i
    recruit_fee = recruit_map[role]
    sc(ws_h, r, 1, role)
    sc(ws_h, r, 2, recruit_fee, nf=NF_MAN, align="center")
    sc(ws_h, r, 4, y1h * recruit_fee, nf=NF_MAN, align="center")
    sc(ws_h, r, 5, y2h * recruit_fee, nf=NF_MAN, align="center")
    sc(ws_h, r, 6, y3h * recruit_fee, nf=NF_MAN, align="center")

r_rec_tot = 36 + len(HIRE)
sc(ws_h, r_rec_tot, 1, "採用費合計（万円）", bold=True, fc=C_NAVY, color="FFFFFF")
for c in [4, 5, 6]:
    sc(ws_h, r_rec_tot, c, f=f"=SUM({gl(c)}36:{gl(c)}{r_rec_tot-1})",
       bold=True, fc=C_ORANGE, nf=NF_MAN, align="center")
bdr(ws_h, 35, r_rec_tot, 1, 6)

# ★ キーセル定義（前提条件シートからの参照先）
# D36=人件費Y1, E36=人件費Y2, F36=人件費Y3
# D37=採用費Y1, E37=採用費Y2, F37=採用費Y3
# → 前提条件シート行31,32がこれを参照する
# 実際のセルアドレス確認用にコメント
note_row(ws_h, r_rec_tot+1,
    f"★ 前提条件シート参照セル: 人件費={gl(4)}{r_sal_tot}〜{gl(6)}{r_sal_tot} / 採用費={gl(4)}{r_rec_tot}〜{gl(6)}{r_rec_tot}", 8)

# 前提条件の参照式を確定セルアドレスに合わせて修正
# （人員計画シートのD36・E36・F36 が人件費合計行）
# Part1終了 ─ Part2に続く
# ══════════════════════════════════════════════
# _v42_part2a.py: PL（損益計算書）+ BS（貸借対照表）
# Part1(_v42_part1.py)の続き。wb変数を引き継ぐ前提。
# ══════════════════════════════════════════════

# ──── SHEET 3: PL（損益計算書）────
ws_pl = wb.create_sheet("PL（損益計算書）")
ws_pl.column_dimensions["A"].width = 32
for c in ["B","C","D","E"]:
    ws_pl.column_dimensions[c].width = 15

title_row(ws_pl, 1, "UNLID 損益計算書（P/L）v4.2  ─  単位：万円  |  前提条件シート連動", 5)
sc(ws_pl, 2, 1,
   "★ 人件費・採用費は人員計画シートから自動参照。全数値は前提条件シートのドライバーで算出。",
   italic=True, fc=C_NOTE, size=9)
mrg(ws_pl, 2, 1, 5)
col_hdrs(ws_pl, 4, ["項目", "Year1", "Year2", "Year3", "3年累計"], fc_def=C_BHDR)

def p(r, y):
    """前提条件シートのセル参照（r=行番号, y=1/2/3→B/C/D列）"""
    return f"前提条件!{['B','C','D'][y-1]}{r}"

# セクションヘッダー（merge行）
for row, txt in [
    (5,  "▌売上高"),
    (10, "▌売上原価（COGS）"),
    (17, "▌SG&A（販売費及び一般管理費）"),
    (33, "▌参考指標"),
]:
    sc(ws_pl, row, 1, txt, bold=True, fc=C_NAVY, color="FFFFFF")
    mrg(ws_pl, row, 1, 5)

# (行, ラベル, Year1式, Year2式, Year3式, bold, fc)
PL_DATA = [
    # 売上高
    (6,  "  B2B受託収益",         f"={p(8,1)}",  f"={p(8,2)}",  f"={p(8,3)}",  False, None),
    (7,  "  人材紹介フィー",       f"={p(11,1)}", f"={p(11,2)}", f"={p(11,3)}", False, None),
    (8,  "  プロジェクト収益",     f"={p(14,1)}", f"={p(14,2)}", f"={p(14,3)}", False, None),
    (9,  "売上合計",               f"={p(15,1)}", f"={p(15,2)}", f"={p(15,3)}", True,  C_TOT),
    # COGS
    (11, "  AIサーバー・API費",    f"={p(20,1)}+{p(21,1)}", f"={p(20,2)}+{p(21,2)}", f"={p(20,3)}+{p(21,3)}", False, None),
    (12, "  副業ワーカー報酬",     f"={p(23,1)}", f"={p(23,2)}", f"={p(23,3)}", False, None),
    (13, "  外注エンジニア費",     f"={p(24,1)}", f"={p(24,2)}", f"={p(24,3)}", False, None),
    (14, "売上原価合計",           f"={p(25,1)}", f"={p(25,2)}", f"={p(25,3)}", True,  C_SUB),
    (15, "売上総利益",             f"={p(26,1)}", f"={p(26,2)}", f"={p(26,3)}", True,  C_GPFT),
    (16, "  粗利率",               f"={p(27,1)}", f"={p(27,2)}", f"={p(27,3)}", False, None),
    # SG&A
    (18, "  人件費合計",           "=人員計画!D32", "=人員計画!E32", "=人員計画!F32", False, None),
    (19, "  採用費合計",           "=人員計画!D42", "=人員計画!E42", "=人員計画!F42", False, None),
    (20, "  B2Bマーケ・CAC費",    f"={p(33,1)}", f"={p(33,2)}", f"={p(33,3)}", False, None),
    (21, "  人材紹介広告費",       f"={p(34,1)}", f"={p(34,2)}", f"={p(34,3)}", False, None),
    (22, "  地代家賃・オフィス",   f"={p(35,1)}", f"={p(35,2)}", f"={p(35,3)}", False, None),
    (23, "  SaaS・システム費",     f"={p(36,1)}", f"={p(36,2)}", f"={p(36,3)}", False, None),
    (24, "  法務・許認可費",       f"={p(37,1)}", f"={p(37,2)}", f"={p(37,3)}", False, None),
    (25, "  その他固定費",         f"={p(38,1)}", f"={p(38,2)}", f"={p(38,3)}", False, None),
    (26, "SG&A合計",              f"={p(39,1)}", f"={p(39,2)}", f"={p(39,3)}", True,  C_SUB),
    # 利益
    (27, "営業利益",               f"={p(40,1)}", f"={p(40,2)}", f"={p(40,3)}", True,  C_TOT),
    (28, "  営業利益率",           f"={p(41,1)}", f"={p(41,2)}", f"={p(41,3)}", False, None),
    (29, "  支払利息",             f"={p(64,1)}", f"={p(64,2)}", f"={p(64,3)}", False, C_LOSS),
    (30, "税引前利益",             f"={p(65,1)}", f"={p(65,2)}", f"={p(65,3)}", True,  C_TOT),
    (31, "  法人税等（30%）",      f"={p(66,1)}", f"={p(66,2)}", f"={p(66,3)}", False, None),
    (32, "当期純利益",             f"={p(67,1)}", f"={p(67,2)}", f"={p(67,3)}", True,  C_GPFT),
    # 参考
    (34, "  減価償却費（参考）",   f"={p(54,1)}", f"={p(54,2)}", f"={p(54,3)}", False, C_NOTE),
    (35, "  EBITDA（参考）",       f"={p(40,1)}+{p(54,1)}", f"={p(40,2)}+{p(54,2)}", f"={p(40,3)}+{p(54,3)}", True, C_GPFT),
]

for row, lbl, f1, f2, f3, bold, fc in PL_DATA:
    sc(ws_pl, row, 1, lbl, bold=bold, fc=fc)
    nf = NF_PCT if "率" in lbl else NF_MAN
    for c, fml in [(2, f1), (3, f2), (4, f3)]:
        sc(ws_pl, row, c, f=fml, bold=bold, fc=fc, nf=nf, align="center")
    if "率" not in lbl:
        sc(ws_pl, row, 5, f=f"=B{row}+C{row}+D{row}", bold=bold, fc=fc, nf=NF_MAN, align="center")

bdr(ws_pl, 4, 35, 1, 5)

# ──── SHEET 4: BS（貸借対照表）────
ws_bs = wb.create_sheet("BS（貸借対照表）")
ws_bs.column_dimensions["A"].width = 30
for c in ["B", "C", "D", "E"]:
    ws_bs.column_dimensions[c].width = 16

title_row(ws_bs, 1, "UNLID 貸借対照表（B/S）v4.2  ─  単位：万円  |  期末残高ベース", 5)
sc(ws_bs, 2, 1,
   "注：現金残高は「営業CF + 投資CF + 財務CF」の累積。貸借差額セル（Row28）で一致確認。",
   italic=True, fc=C_NOTE, size=9)
mrg(ws_bs, 2, 1, 5)
col_hdrs(ws_bs, 4, ["項目", "Y0（期首）", "Y1末", "Y2末", "Y3末"], fc_def=C_BHDR)

Q = "前提条件!"  # 前提条件シート参照プレフィックス

# セクションヘッダー
for row, txt in [
    (5,  "▌流動資産"),
    (9,  "▌固定資産"),
    (14, "▌流動負債"),
    (18, "▌固定負債"),
    (23, "▌純資産"),
]:
    sc(ws_bs, row, 1, txt, bold=True, fc=C_NAVY, color="FFFFFF")
    mrg(ws_bs, row, 1, 5)

# 現金・預金（累積CF計算）
# Y0 = 資本金払込 + 借入金
# Y1 = Y0 + 純利益 + 償却費 - ソフト投資 - 売掛金 + 買掛金
# Y2 = Y1 + 上記変動分 + シード調達
# Y3 = Y2 + 上記変動分 + シリーズA
sc(ws_bs, 6, 1, "現金・預金", bold=True)
sc(ws_bs, 6, 2, f"={Q}B59+{Q}B62",
   bold=True, nf=NF_MAN, align="center")
sc(ws_bs, 6, 3,
   f"=B6+{Q}B67+{Q}B54-{Q}B52-{Q}B47+{Q}B48",
   bold=True, nf=NF_MAN, align="center")
sc(ws_bs, 6, 4,
   f"=C6+{Q}C67+{Q}C54-{Q}C52-({Q}C47-{Q}B47)+({Q}C48-{Q}B48)+{Q}C60",
   bold=True, nf=NF_MAN, align="center")
sc(ws_bs, 6, 5,
   f"=D6+{Q}D67+{Q}D54-{Q}D52-({Q}D47-{Q}C47)+({Q}D48-{Q}C48)+{Q}D61",
   bold=True, nf=NF_MAN, align="center")

# 売掛金（前提条件シート参照）
sc(ws_bs, 7, 1, "売掛金")
sc(ws_bs, 7, 2, v=0, nf=NF_MAN, align="center")
for c, col in [(3, "B"), (4, "C"), (5, "D")]:
    sc(ws_bs, 7, c, f=f"={Q}{col}47", nf=NF_MAN, align="center")

# 流動資産合計
sc(ws_bs, 8, 1, "流動資産合計", bold=True, fc=C_SUB)
for c in [2, 3, 4, 5]:
    sc(ws_bs, 8, c, f=f"={gl(c)}6+{gl(c)}7", bold=True, fc=C_SUB, nf=NF_MAN, align="center")

# ソフトウェア（純額）
sc(ws_bs, 10, 1, "ソフトウェア（純額）")
sc(ws_bs, 10, 2, v=0, nf=NF_MAN, align="center")
for c, col in [(3, "B"), (4, "C"), (5, "D")]:
    sc(ws_bs, 10, c, f=f"={Q}{col}55", nf=NF_MAN, align="center")

# 固定資産合計
sc(ws_bs, 11, 1, "固定資産合計", bold=True, fc=C_SUB)
for c in [2, 3, 4, 5]:
    sc(ws_bs, 11, c, f=f"={gl(c)}10", bold=True, fc=C_SUB, nf=NF_MAN, align="center")

# 資産合計
sc(ws_bs, 12, 1, "資産合計", bold=True, fc=C_TOT)
for c in [2, 3, 4, 5]:
    sc(ws_bs, 12, c, f=f"={gl(c)}8+{gl(c)}11", bold=True, fc=C_TOT, nf=NF_MAN, align="center")
bdr(ws_bs, 4, 12, 1, 5)

# 買掛金
sc(ws_bs, 15, 1, "買掛金")
sc(ws_bs, 15, 2, v=0, nf=NF_MAN, align="center")
for c, col in [(3, "B"), (4, "C"), (5, "D")]:
    sc(ws_bs, 15, c, f=f"={Q}{col}48", nf=NF_MAN, align="center")

# 未払費用（人件費1ヶ月相当 → 人員計画シート参照）
sc(ws_bs, 16, 1, "未払費用（人件費1ヶ月相当）")
sc(ws_bs, 16, 2, v=0, nf=NF_MAN, align="center")
for c, col in [(3, "D"), (4, "E"), (5, "F")]:
    sc(ws_bs, 16, c, f=f"=ROUND(人員計画!{col}32/12,0)", nf=NF_MAN, align="center")

# 流動負債合計
sc(ws_bs, 17, 1, "流動負債合計", bold=True, fc=C_SUB)
for c in [2, 3, 4, 5]:
    sc(ws_bs, 17, c, f=f"={gl(c)}15+{gl(c)}16", bold=True, fc=C_SUB, nf=NF_MAN, align="center")

# 借入金（返済なし想定）
sc(ws_bs, 19, 1, "借入金（期中返済なし想定）")
for c in [2, 3, 4, 5]:
    sc(ws_bs, 19, c, f=f"={Q}B62", nf=NF_MAN, align="center")

# 固定負債合計
sc(ws_bs, 20, 1, "固定負債合計", bold=True, fc=C_SUB)
for c in [2, 3, 4, 5]:
    sc(ws_bs, 20, c, f=f"={gl(c)}19", bold=True, fc=C_SUB, nf=NF_MAN, align="center")

# 負債合計
sc(ws_bs, 21, 1, "負債合計", bold=True, fc=C_TOT)
for c in [2, 3, 4, 5]:
    sc(ws_bs, 21, c, f=f"={gl(c)}17+{gl(c)}20", bold=True, fc=C_TOT, nf=NF_MAN, align="center")
bdr(ws_bs, 14, 21, 1, 5)

# 払込資本（資本金＋調達額累計）
sc(ws_bs, 24, 1, "払込資本（資本金＋調達額累計）")
sc(ws_bs, 24, 2, f"={Q}B59", nf=NF_MAN, align="center")          # Y0: 資本金のみ
sc(ws_bs, 24, 3, f"={Q}B59", nf=NF_MAN, align="center")          # Y1: 新規調達なし
sc(ws_bs, 24, 4, f"={Q}B59+{Q}C60", nf=NF_MAN, align="center")  # Y2: +シード3,000万
sc(ws_bs, 24, 5, f"={Q}B59+{Q}C60+{Q}D61", nf=NF_MAN, align="center")  # Y3: +シリーズA10,000万

# 利益剰余金（累積純利益）
sc(ws_bs, 25, 1, "利益剰余金（累積純利益）")
sc(ws_bs, 25, 2, v=0, nf=NF_MAN, align="center")
sc(ws_bs, 25, 3, f"={Q}B67", nf=NF_MAN, align="center")
sc(ws_bs, 25, 4, f"={Q}B67+{Q}C67", nf=NF_MAN, align="center")
sc(ws_bs, 25, 5, f"={Q}B67+{Q}C67+{Q}D67", nf=NF_MAN, align="center")

# 純資産合計
sc(ws_bs, 26, 1, "純資産合計", bold=True, fc=C_GPFT)
for c in [2, 3, 4, 5]:
    sc(ws_bs, 26, c, f=f"={gl(c)}24+{gl(c)}25", bold=True, fc=C_GPFT, nf=NF_MAN, align="center")

# 検証：負債・純資産合計（資産合計と一致すれば貸借OK）
sc(ws_bs, 27, 1, "負債・純資産合計（検証）", bold=True, fc=C_TOT)
for c in [2, 3, 4, 5]:
    sc(ws_bs, 27, c, f=f"={gl(c)}21+{gl(c)}26", bold=True, fc=C_TOT, nf=NF_MAN, align="center")

sc(ws_bs, 28, 1, "貸借差額（0 = 正常・要確認）", bold=True)
for c in [2, 3, 4, 5]:
    sc(ws_bs, 28, c, f=f"={gl(c)}12-{gl(c)}27", bold=True,
       fc=C_GPFT, nf=NF_MAN, align="center")

bdr(ws_bs, 23, 28, 1, 5)
# ══════════════════════════════════════════════
# _v42_part2b.py: 月次CF（キャッシュフロー）シート — 36ヶ月
# Part1 + Part2a の続き。wb 変数引き継ぎ前提。
# ══════════════════════════════════════════════

LAST_COL = gl(37)  # AK列 = Month36

# ──── SHEET 5: 月次CF（36ヶ月）────
ws_cf = wb.create_sheet("月次CF（36ヶ月）")
ws_cf.column_dimensions["A"].width = 28
for m in range(1, 37):
    ws_cf.column_dimensions[gl(m + 1)].width = 9

title_row(ws_cf, 1, "UNLID 月次キャッシュフロー v4.2  ─  単位：万円  |  36ヶ月展開", 37)
sc(ws_cf, 2, 1,
   "★ 月次値 = 年次値÷12（平均配分）。設備投資は各年度初月に一括計上。財務調達は固定月（M13/M25）に計上。",
   italic=True, fc=C_NOTE, size=9)
ws_cf.merge_cells(f"A2:{LAST_COL}2")
ws_cf.freeze_panes = "B5"  # A列ラベル固定 + 上4行固定

# Row 3: Year帯ラベル（12ヶ月ずつmerge）
sc(ws_cf, 3, 1, "", fc=C_BHDR)
for start, end, lbl, fc_yr in [
    (1,  12, "Year 1  （M1〜M12）",  C_NAVY),
    (13, 24, "Year 2  （M13〜M24）", C_BLUE),
    (25, 36, "Year 3  （M25〜M36）", C_NAVY),
]:
    sc(ws_cf, 3, start + 1, lbl, bold=True, fc=fc_yr, color="FFFFFF", align="center")
    ws_cf.merge_cells(f"{gl(start+1)}3:{gl(end+1)}3")

# Row 4: 月番号ラベル
sc(ws_cf, 4, 1, "項目 / 月", bold=True, fc=C_BHDR, color="FFFFFF")
for m in range(1, 37):
    bg = C_NAVY if (m - 1) // 12 % 2 == 0 else C_BLUE
    sc(ws_cf, 4, m + 1, f"M{m}", bold=True, fc=bg, color="FFFFFF", align="center")

def yr_col(m): return ["B","C","D"][(m - 1) // 12]  # 年次列 → 前提条件シート列

# ── P&Lサマリーセクションヘッダー ──
for row, txt in [
    (5,  "▌売上高"),
    (10, "▌売上原価・粗利"),
    (13, "▌SG&A・営業利益"),
]:
    sc(ws_cf, row, 1, txt, bold=True, fc=C_NAVY, color="FFFFFF")
    ws_cf.merge_cells(f"A{row}:{LAST_COL}{row}")

# P&Lサマリー月次行: (row, label, 前提条件行番号, bold, fc)
PL_CF = [
    (6,  "  B2B受託収益",         8,  False, None),
    (7,  "  人材紹介フィー",       11, False, None),
    (8,  "  プロジェクト収益",     14, False, None),
    (9,  "売上合計",               15, True,  C_TOT),
    (11, "  売上原価合計",         25, False, C_SUB),
    (12, "売上総利益",             26, True,  C_GPFT),
    (14, "  SG&A合計",            39, False, C_SUB),
    (15, "営業利益",               40, True,  C_TOT),
    (16, "当月純利益（税引後）",    67, True,  C_GPFT),
    (18, "当月償却費（参考・非現金）", 54, False, C_NOTE),
]

for row, lbl, mae_row, bold, fc in PL_CF:
    sc(ws_cf, row, 1, lbl, bold=bold, fc=fc)
    for m in range(1, 37):
        yc = yr_col(m)
        sc(ws_cf, row, m + 1,
           f=f"=ROUND(前提条件!{yc}{mae_row}/12,0)",
           bold=bold, fc=fc, nf=NF_MAN, align="center")

# ── 月次CFセクション ──
sc(ws_cf, 19, 1, "▌月次キャッシュフロー", bold=True, fc=C_NAVY, color="FFFFFF")
ws_cf.merge_cells(f"A19:{LAST_COL}19")

# ── Row 20: 期首現金残高（BSシートY0起点 → 前月期末を連鎖）──
sc(ws_cf, 20, 1, "期首現金残高")
for m in range(1, 37):
    col = m + 1
    if m == 1:
        f_str = "=BS（貸借対照表）!B6"     # Y0期首現金 = 資本金+借入金
    else:
        f_str = f"={gl(m)}26"               # 前月Row26（期末）を参照
    sc(ws_cf, 20, col, f=f_str, nf=NF_MAN, align="center")

# ── Row 21: 当月純利益 ──
sc(ws_cf, 21, 1, "当月純利益")
for m in range(1, 37):
    yc = yr_col(m)
    sc(ws_cf, 21, m + 1, f=f"=ROUND(前提条件!{yc}67/12,0)", nf=NF_MAN, align="center")

# ── Row 22: 当月償却費（非現金加算） ──
sc(ws_cf, 22, 1, "当月償却費（非現金加算）")
for m in range(1, 37):
    yc = yr_col(m)
    sc(ws_cf, 22, m + 1, f=f"=ROUND(前提条件!{yc}54/12,0)", nf=NF_MAN, align="center")

# ── Row 23: 設備投資（各年度初月 M1/M13/M25 に一括計上）──
sc(ws_cf, 23, 1, "設備投資（－）")
for m in range(1, 37):
    col = m + 1
    if m == 1:
        sc(ws_cf, 23, col, f="=前提条件!B52", fc=C_LOSS, nf=NF_MAN, align="center")
    elif m == 13:
        sc(ws_cf, 23, col, f="=前提条件!C52", fc=C_LOSS, nf=NF_MAN, align="center")
    elif m == 25:
        sc(ws_cf, 23, col, f="=前提条件!D52", fc=C_LOSS, nf=NF_MAN, align="center")
    else:
        sc(ws_cf, 23, col, v=0, nf=NF_MAN, align="center")

# ── Row 24: 運転資本変動（ΔAR − ΔAP、正 = 現金流出）──
# Y1: AR・APともゼロ起点から年次値まで均等増
# Y2/Y3: 年次差分を均等配分
sc(ws_cf, 24, 1, "運転資本変動（ΔAR−ΔAP）")
for m in range(1, 37):
    col = m + 1
    yr = (m - 1) // 12
    if yr == 0:
        f_str = (
            "=ROUND(前提条件!B47/12,0)"
            "-ROUND(前提条件!B48/12,0)"
        )
    elif yr == 1:
        f_str = (
            "=ROUND((前提条件!C47-前提条件!B47)/12,0)"
            "-ROUND((前提条件!C48-前提条件!B48)/12,0)"
        )
    else:
        f_str = (
            "=ROUND((前提条件!D47-前提条件!C47)/12,0)"
            "-ROUND((前提条件!D48-前提条件!C48)/12,0)"
        )
    sc(ws_cf, 24, col, f=f_str, nf=NF_MAN, align="center")

# ── Row 25: 財務調達（シード M13 / シリーズA M25）──
sc(ws_cf, 25, 1, "財務調達（調達・借入金）")
for m in range(1, 37):
    col = m + 1
    if m == 13:
        sc(ws_cf, 25, col, f="=前提条件!C60", fc=C_GPFT, nf=NF_MAN, align="center")
    elif m == 25:
        sc(ws_cf, 25, col, f="=前提条件!D61", fc=C_GPFT, nf=NF_MAN, align="center")
    else:
        sc(ws_cf, 25, col, v=0, nf=NF_MAN, align="center")

# ── Row 26: 期末現金残高 ──
# = 期首 + 純利益 + 償却費 − 設備投資 − 運転資本変動 + 財務調達
sc(ws_cf, 26, 1, "期末現金残高", bold=True, fc=C_TOT)
for m in range(1, 37):
    cl = gl(m + 1)
    sc(ws_cf, 26, m + 1,
       f=f"={cl}20+{cl}21+{cl}22-{cl}23-{cl}24+{cl}25",
       bold=True, fc=C_TOT, nf=NF_MAN, align="center")

bdr(ws_cf, 4, 26, 1, 37)
# ══════════════════════════════════════════════
# _v42_part2c.py: 感度分析シート + wb.save()
# Part1 + 2a + 2b の続き。wb 変数引き継ぎ前提。
# ══════════════════════════════════════════════

# ──── SHEET 6: 感度分析 ────
ws_sa = wb.create_sheet("感度分析")
ws_sa.column_dimensions["A"].width = 32
for c in ["B","C","D","E","F","G","H","I"]:
    ws_sa.column_dimensions[c].width = 15

title_row(ws_sa, 1, "UNLID 感度分析 v4.2  ─  ユニットエコノミクス × シナリオ比較", 9)
sc(ws_sa, 2, 1,
   "★ マトリックスは静的計算。シナリオサマリーは前提条件シート基準値に倍率を乗算。",
   italic=True, fc=C_NOTE, size=9)
mrg(ws_sa, 2, 1, 9)

MAE = "前提条件!"  # 前提条件シート参照プレフィックス

# ────────────────────────────────────────────
# TABLE 1: 月商マトリックス（FDE人数 × 担当社数）
# ────────────────────────────────────────────
sc(ws_sa, 4, 1, "▌月商マトリックス（万円/月）  ─  単価：20万円/社 固定", bold=True, fc=C_NAVY, color="FFFFFF")
mrg(ws_sa, 4, 1, 9)

sc(ws_sa, 5, 1, "FDE人数 ╲ 担当社数", bold=True, fc=C_BHDR, color="FFFFFF")
CLIENTS_LIST = [4, 5, 6, 7, 8]
for ci, clients in enumerate(CLIENTS_LIST):
    lbl = f"{clients}社 ★BEP" if clients == 6 else f"{clients}社"
    bg  = C_BLUE if clients == 6 else C_BHDR
    sc(ws_sa, 5, ci + 2, lbl, bold=True, fc=bg, color="FFFFFF", align="center")

for fi, fde in enumerate([1, 2, 3, 4, 5]):
    row = 6 + fi
    sc(ws_sa, row, 1, f"FDE {fde}名", bold=True)
    for ci, clients in enumerate(CLIENTS_LIST):
        col  = ci + 2
        rev  = fde * clients * 20          # 月商（万円）
        is_bep = (fde == 1 and clients == 6)
        bg = C_GPFT if is_bep else (C_LBLUE if clients == 6 else None)
        sc(ws_sa, row, col, v=rev, bold=(clients == 6), fc=bg, nf=NF_MAN, align="center")
bdr(ws_sa, 5, 10, 1, 6)

# ────────────────────────────────────────────
# TABLE 2: 粗利マトリックス（ワーカー費15% + FDE給与40万/名を控除）
# ────────────────────────────────────────────
sc(ws_sa, 12, 1, "▌粗利マトリックス（万円/月）  ─  ワーカー費15%控除 + FDE給与40万/名控除", bold=True, fc=C_NAVY, color="FFFFFF")
mrg(ws_sa, 12, 1, 9)

sc(ws_sa, 13, 1, "FDE人数 ╲ 担当社数", bold=True, fc=C_BHDR, color="FFFFFF")
for ci, clients in enumerate(CLIENTS_LIST):
    lbl = f"{clients}社 ★BEP" if clients == 6 else f"{clients}社"
    bg  = C_BLUE if clients == 6 else C_BHDR
    sc(ws_sa, 13, ci + 2, lbl, bold=True, fc=bg, color="FFFFFF", align="center")

for fi, fde in enumerate([1, 2, 3, 4, 5]):
    row = 14 + fi
    sc(ws_sa, row, 1, f"FDE {fde}名", bold=True)
    for ci, clients in enumerate(CLIENTS_LIST):
        col   = ci + 2
        gross = round(fde * clients * 20 * 0.85 - fde * 40)  # 粗利（万円/月）
        is_bep = (fde == 1 and clients == 6)
        bg = C_GPFT if is_bep else (C_LBLUE if clients == 6 else (C_LOSS if gross < 0 else None))
        sc(ws_sa, row, col, v=gross, bold=(clients == 6), fc=bg, nf=NF_MAN, align="center")
bdr(ws_sa, 13, 18, 1, 6)

# ────────────────────────────────────────────
# TABLE 3: B2B単価感度（Year1）
# ────────────────────────────────────────────
sc(ws_sa, 20, 1, "▌B2B単価感度（Year1）  ─  FDE×社数構成不変、月額単価のみ変動", bold=True, fc=C_NAVY, color="FFFFFF")
mrg(ws_sa, 20, 1, 9)

col_hdrs(ws_sa, 21,
    ["月額単価", "Y1 売上（万円）", "Y1 粗利（万円）", "Y1 営業利益（万円）", "Base比変化率"],
    fc_def=C_BHDR)

for pi, price in enumerate([15, 17, 20, 23, 25]):
    row   = 22 + pi
    ratio = price / 20                     # Base比（20万基準）
    is_base = (price == 20)
    bg = C_LBLUE if is_base else (C_GPFT if price > 20 else C_LOSS)
    lbl = f"{price}万円/社{'  ★Base' if is_base else ''}"
    sc(ws_sa, row, 1, lbl, bold=is_base, fc=bg)
    sc(ws_sa, row, 2, f"=ROUND({MAE}B15*{ratio},0)", bold=is_base, fc=bg, nf=NF_MAN, align="center")
    sc(ws_sa, row, 3, f"=ROUND({MAE}B26*{ratio},0)", bold=is_base, fc=bg, nf=NF_MAN, align="center")
    sc(ws_sa, row, 4, f"=ROUND({MAE}B40*{ratio},0)", bold=is_base, fc=bg, nf=NF_MAN, align="center")
    sc(ws_sa, row, 5, v=round(ratio - 1, 3),        bold=is_base, fc=bg, nf=NF_PCT, align="center")
bdr(ws_sa, 21, 26, 1, 5)

# ────────────────────────────────────────────
# TABLE 4: 3シナリオサマリー（Bear / Base / Bull）
# ────────────────────────────────────────────
sc(ws_sa, 28, 1, "▌3シナリオサマリー（前提条件基準値 × 倍率）", bold=True, fc=C_NAVY, color="FFFFFF")
mrg(ws_sa, 28, 1, 9)

col_hdrs(ws_sa, 29,
    ["シナリオ", "Y1 売上", "Y1 粗利", "Y1 営業利益", "Y2 売上", "Y2 粗利", "Y2 営業利益", "Y3 売上", "Y3 営業利益"],
    fc_def=C_BHDR)

for si, (lbl, mult, fc) in enumerate([
    ("Bear（悲観）  ×0.70", 0.70, C_LOSS),
    ("Base（中央）  ×1.00", 1.00, C_LBLUE),
    ("Bull（楽観）  ×1.30", 1.30, C_GPFT),
]):
    row = 30 + si
    sc(ws_sa, row, 1, lbl, bold=(mult == 1.0), fc=fc)
    for c, fml in enumerate([
        f"=ROUND({MAE}B15*{mult},0)", f"=ROUND({MAE}B26*{mult},0)", f"=ROUND({MAE}B40*{mult},0)",
        f"=ROUND({MAE}C15*{mult},0)", f"=ROUND({MAE}C26*{mult},0)", f"=ROUND({MAE}C40*{mult},0)",
        f"=ROUND({MAE}D15*{mult},0)",                                f"=ROUND({MAE}D40*{mult},0)",
    ], start=2):
        sc(ws_sa, row, c, f=fml, bold=(mult == 1.0), fc=fc, nf=NF_MAN, align="center")
bdr(ws_sa, 29, 32, 1, 9)

# ════════════════════════════════════════════
# 保存
# ════════════════════════════════════════════
OUTPUT_FILE = "UNLID_財務モデル_v4.2.xlsx"
wb.save(OUTPUT_FILE)
print(f"✅  生成完了 → {OUTPUT_FILE}")
print(f"    シート一覧: {[s.title for s in wb.worksheets]}")
