# -*- coding: utf-8 -*-
"""UNLID 財務三表 v4.2 ─ Part 1
helpers + 前提条件シート + 人員計画シート
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# ───────────────── ヘルパー ─────────────────
def sc(ws, r, c, v=None, f=None, bold=False, italic=False, size=10,
       fc=None, align="left", nf=None, color="000000"):
    cell = ws.cell(row=r, column=c)
    cell.value = f if f else v
    cell.font = Font(name="Meiryo UI", bold=bold, italic=italic, size=size, color=color)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    if fc:
        cell.fill = PatternFill("solid", fgColor=fc)
    if nf:
        cell.number_format = nf
    return cell

def bdr(ws, r1, r2, c1, c2, style="thin"):
    s = Side(style=style)
    b = Border(left=s, right=s, top=s, bottom=s)
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(row=r, column=c).border = b

def gl(c): return get_column_letter(c)
def mrg(ws, r, c1, c2): ws.merge_cells(f"{gl(c1)}{r}:{gl(c2)}{r}")

def title_row(ws, r, text, ncols=6):
    sc(ws, r, 1, text, bold=True, size=13, fc=C_NAVY, color="FFFFFF", align="center")
    mrg(ws, r, 1, ncols)
    ws.row_dimensions[r].height = 26

def sec_hdr(ws, r, text, ncols=6):
    sc(ws, r, 1, text, bold=True, fc=C_BLUE, color="FFFFFF", align="left")
    mrg(ws, r, 1, ncols)
    ws.row_dimensions[r].height = 20

def note_row(ws, r, text, ncols=6):
    sc(ws, r, 1, text, italic=True, fc=C_NOTE, size=9)
    mrg(ws, r, 1, ncols)

def col_hdrs(ws, r, labels, fcs=None, fc_def="BDD7EE"):
    for i, lbl in enumerate(labels, 1):
        fc = (fcs[i-1] if fcs else None) or fc_def
        sc(ws, r, i, lbl, bold=True, fc=fc, align="center")

# ─── カラーパレット ───
C_NAVY  = "0F172A"
C_BLUE  = "2563EB"
C_LBLUE = "EFF6FF"
C_BHDR  = "DBEAFE"
C_TOT   = "FFF9C4"
C_GPFT  = "DCFCE7"
C_LOSS  = "FEE2E2"
C_NOTE  = "F8FAFC"
C_WARN  = "FEF3C7"
C_SUB   = "E0E7FF"
C_ORANGE= "FED7AA"
NF_MAN  = '#,##0'
NF_PCT  = '0.0%'
NF_2D   = '0.00'

# ══════════════════════════════════════════════
# SHEET 1: 前提条件
# ══════════════════════════════════════════════
ws_a = wb.active
ws_a.title = "前提条件"
ws_a.column_dimensions["A"].width = 36
for col in ["B","C","D"]:
    ws_a.column_dimensions[col].width = 16

title_row(ws_a, 1, "UNLID 財務モデル 前提条件 v4.2  ─  単位：万円  |  中央値シナリオ", 4)
sc(ws_a, 2, 1,
   "v4.2変更：COGS/SG&A完全展開 ／ 人員計画連動 ／ 運転資本サイト ／ 設備投資・償却 ／ 資金調達前提 追加",
   italic=True, fc=C_NOTE, size=9)
mrg(ws_a, 2, 1, 4)

# ─ セクション①: 売上ドライバー（v4.1踏襲）
sec_hdr(ws_a, 4, "① 売上ドライバー", 4)
col_hdrs(ws_a, 5, ["項目", "Year1", "Year2", "Year3"], fc_def=C_BHDR)

A1 = [
    (6,  "B2B契約企業数（期末・社）",            10,   50,  160, False),
    (7,  "FDE受託 月額単価（万円/社）",           20,   23,   25, False),
    (8,  "B2B受託収益（万円）",          "=B6*B7*12","=C6*C7*12","=D6*D7*12", True),
    (9,  "人材紹介成立件数（件）",                 3,   20,   50, False),
    (10, "人材紹介フィー単価（万円/件）",          140,  140,  140, False),
    (11, "人材紹介フィー合計（万円）",         "=B9*B10","=C9*C10","=D9*D10", True),
    (12, "受託プロジェクト件数（件）",              0,    5,   15, False),
    (13, "プロジェクト単価（万円/件・ベース）",     0,  160,  160, False),
    (14, "プロジェクト収益（万円）",       "=B12*B13","=C12*C13","=D12*D13", True),
    (15, "売上合計（万円）",      "=B8+B11+B14","=C8+C11+C14","=D8+D11+D14", True),
]
for row, lbl, v1, v2, v3, bold in A1:
    fc = C_TOT if "売上合計" in lbl else (C_GPFT if bold else None)
    sc(ws_a, row, 1, lbl, bold=bold, fc=fc)
    for c, v in [(2,v1),(3,v2),(4,v3)]:
        nf = NF_MAN
        if isinstance(v, str):
            sc(ws_a, row, c, f=v, bold=bold, fc=fc, nf=nf, align="center")
        else:
            sc(ws_a, row, c, v=v, bold=bold, fc=fc, nf=nf, align="center")
bdr(ws_a, 5, 15, 1, 4)
note_row(ws_a, 16, "※ B2Bは期末社数×単価×12ヶ月（期中平均で計算する場合は別途月次CFシート参照）", 4)

# ─ セクション②: COGS（売上原価）前提
sec_hdr(ws_a, 18, "② COGS（売上原価）前提  ─  サーバー/API費・外注費", 4)
col_hdrs(ws_a, 19, ["項目", "Year1", "Year2", "Year3"], fc_def=C_BHDR)

A2 = [
    (20, "AIサーバー・インフラ費（万円）",           24,   72,  180),
    (21, "外部API利用料（OpenAI等・万円）",          12,   60,  180),
    (22, "副業ワーカー報酬率（対B2B受託売上）",     0.09, 0.08, 0.08),
    (23, "副業ワーカー報酬（万円）",  "=B8*B22","=C8*C22","=D8*D22"),
    (24, "外注エンジニア費（プロジェクト分・万円）",  0,   200,  600),
    (25, "COGS合計（万円）", "=SUM(B20:B24)","=SUM(C20:C24)","=SUM(D20:D24)"),
    (26, "売上総利益（万円）",       "=B15-B25","=C15-C25","=D15-D25"),
    (27, "粗利率",                "=B26/B15","=C26/C15","=D26/D15"),
]
for row, lbl, v1, v2, v3 in A2:
    bold = ("合計" in lbl or "粗利" in lbl)
    fc = C_SUB if "COGS合計" in lbl else (C_GPFT if "売上総利益" in lbl else None)
    nf = NF_PCT if "率" in lbl else NF_MAN
    sc(ws_a, row, 1, lbl, bold=bold, fc=fc)
    for c, v in [(2,v1),(3,v2),(4,v3)]:
        if isinstance(v, str):
            sc(ws_a, row, c, f=v, bold=bold, fc=fc, nf=nf, align="center")
        else:
            sc(ws_a, row, c, v=v, bold=bold, fc=fc, nf=nf, align="center")
bdr(ws_a, 19, 27, 1, 4)

# ─ セクション③: SG&A前提（人員計画シートと連動）
sec_hdr(ws_a, 29, "③ SG&A（販売費及び一般管理費）前提  ─  人員計画シートと連動", 4)
col_hdrs(ws_a, 30, ["項目", "Year1", "Year2", "Year3"], fc_def=C_BHDR)

A3 = [
    (31, "人件費合計（万円）※人員計画シート参照",
         "=人員計画!D32", "=人員計画!E32", "=人員計画!F32"),
    (32, "採用費合計（万円）※人員計画シート参照",
         "=人員計画!D42", "=人員計画!E42", "=人員計画!F42"),
    (33, "B2Bマーケティング・CAC費（万円）",        120,  360,  600),
    (34, "人材紹介 候補者集客広告費（万円）",         30,  200,  500),
    (35, "地代家賃・オフィス費（万円）",              60,  180,  360),
    (36, "SaaSツール・システム費（万円）",            60,  120,  240),
    (37, "法務・許認可・顧問費（万円）",             200,  300,  400),
    (38, "その他固定費（万円）",                      60,  120,  200),
    (39, "SG&A合計（万円）",
         "=SUM(B31:B38)","=SUM(C31:C38)","=SUM(D31:D38)"),
    (40, "営業利益（万円）",     "=B26-B39","=C26-C39","=D26-D39"),
    (41, "営業利益率",           "=B40/B15","=C40/C15","=D40/D15"),
]
for row, lbl, v1, v2, v3 in A3:
    bold = ("合計" in lbl or "営業利益" in lbl)
    fc = C_SUB if "SG&A合計" in lbl else (C_TOT if "営業利益（万円）" in lbl else None)
    nf = NF_PCT if "率" in lbl else NF_MAN
    sc(ws_a, row, 1, lbl, bold=bold, fc=fc)
    for c, v in [(2,v1),(3,v2),(4,v3)]:
        if isinstance(v, str):
            sc(ws_a, row, c, f=v, bold=bold, fc=fc, nf=nf, align="center")
        else:
            sc(ws_a, row, c, v=v, bold=bold, fc=fc, nf=nf, align="center")
bdr(ws_a, 30, 41, 1, 4)

# ─ セクション④: 運転資本サイト（BS・CF連動）
sec_hdr(ws_a, 43, "④ 運転資本サイト（売掛金・買掛金）", 4)
col_hdrs(ws_a, 44, ["項目", "Year1", "Year2", "Year3"], fc_def=C_BHDR)

A4 = [
    (45, "売掛金 回収サイト（日）─ 月末締め翌月末払い", 30, 30, 30),
    (46, "買掛金 支払サイト（日）─ 月末締め翌月末払い", 30, 30, 30),
    (47, "期末売掛金残高（万円）=売上/12",
         "=ROUND(B15/12,0)","=ROUND(C15/12,0)","=ROUND(D15/12,0)"),
    (48, "期末買掛金残高（万円）=COGS/12",
         "=ROUND(B25/12,0)","=ROUND(C25/12,0)","=ROUND(D25/12,0)"),
]
for row, lbl, v1, v2, v3 in A4:
    nf = NF_MAN if "万円" in lbl else "#,##0"
    sc(ws_a, row, 1, lbl)
    for c, v in [(2,v1),(3,v2),(4,v3)]:
        if isinstance(v, str):
            sc(ws_a, row, c, f=v, nf=NF_MAN, align="center")
        else:
            sc(ws_a, row, c, v=v, nf=NF_MAN, align="center")
bdr(ws_a, 44, 48, 1, 4)

# ─ セクション⑤: 設備投資・ソフトウェア資産・償却
sec_hdr(ws_a, 50, "⑤ 設備投資・ソフトウェア資産・減価償却", 4)
col_hdrs(ws_a, 51, ["項目", "Year1", "Year2", "Year3"], fc_def=C_BHDR)

A5 = [
    (52, "自社システム開発費（資産計上・万円）",    100,  300,  800),
    (53, "償却期間（年）",                            3,    3,    3),
    (54, "当期償却費（万円）",
         "=ROUND(B52/B53,0)","=ROUND(B52/B53+C52/C53,0)","=ROUND(B52/B53+C52/C53+D52/D53,0)"),
    (55, "ソフトウェア資産残高（万円・期末）",
         "=B52-B54","=B52+C52-C54-B54","=B52+C52+D52-D54-C54-B54"),
]
for row, lbl, v1, v2, v3 in A5:
    nf = NF_MAN if "万円" in lbl else NF_MAN
    sc(ws_a, row, 1, lbl)
    for c, v in [(2,v1),(3,v2),(4,v3)]:
        if isinstance(v, str):
            sc(ws_a, row, c, f=v, nf=NF_MAN, align="center")
        else:
            sc(ws_a, row, c, v=v, nf=NF_MAN, align="center")
bdr(ws_a, 51, 55, 1, 4)

# ─ セクション⑥: 資金調達前提
sec_hdr(ws_a, 57, "⑥ 資金調達前提（エクイティ・デット）", 4)
col_hdrs(ws_a, 58, ["項目", "Year1", "Year2", "Year3"], fc_def=C_BHDR)

A6 = [
    (59, "期首資本金（払込済・万円）",              500,    0,    0),
    (60, "シード調達（エクイティ・万円）",            0, 3000,    0),
    (61, "シリーズA調達（万円）",                     0,    0, 10000),
    (62, "銀行借入（万円）",                        500,    0,    0),
    (63, "借入金利（年率）",                       0.02, 0.02, 0.02),
    (64, "支払利息（万円）",
         "=ROUND(B62*B63,0)","=ROUND(B62*C63,0)","=ROUND(B62*D63,0)"),
    (65, "税引前利益（万円）",  "=B40-B64","=C40-C64","=D40-D64"),
    (66, "法人税等（万円）30%", "=IF(B65>0,ROUND(B65*0.3,0),0)",
                                "=IF(C65>0,ROUND(C65*0.3,0),0)",
                                "=IF(D65>0,ROUND(D65*0.3,0),0)"),
    (67, "当期純利益（万円）",  "=B65-B66","=C65-C66","=D65-D66"),
]
for row, lbl, v1, v2, v3 in A6:
    bold = ("税引前利益" in lbl or "当期純利益" in lbl)
    fc = C_TOT if "当期純利益" in lbl else None
    nf = NF_PCT if "金利" in lbl else NF_MAN
    sc(ws_a, row, 1, lbl, bold=bold, fc=fc)
    for c, v in [(2,v1),(3,v2),(4,v3)]:
        if isinstance(v, str):
            sc(ws_a, row, c, f=v, bold=bold, fc=fc, nf=nf, align="center")
        else:
            sc(ws_a, row, c, v=v, bold=bold, fc=fc, nf=nf, align="center")
bdr(ws_a, 58, 67, 1, 4)

# ─ セクション⑦: 解約率（Churn）・CVR前提
sec_hdr(ws_a, 69, "⑦ 解約率（Churn）・成約率（CVR）前提  ─  感度分析シート連動", 4)
col_hdrs(ws_a, 70, ["項目", "Year1", "Year2", "Year3"], fc_def=C_BHDR)

A7 = [
    (71, "月次Churn率（B2B解約率）ベース",         0.02, 0.02, 0.015),
    (72, "年間解約予測社数（期初社数×年次換算）",
         "=ROUND(B6*B71*12,0)","=ROUND(C6*C71*12,0)","=ROUND(D6*D71*12,0)"),
    (73, "グロス新規獲得必要社数（純増＋解約補填）",
         "=B6+B72","=C6-B6+C72","=D6-C6+D72"),
    (74, "人材紹介 候補者充足率（CVR・ベース）",  0.15, 0.18, 0.20),
    (75, "プロジェクト案件 成約率（ベース）",     0.30, 0.35, 0.35),
]
for row, lbl, v1, v2, v3 in A7:
    nf = NF_PCT if "率" in lbl else NF_MAN
    sc(ws_a, row, 1, lbl)
    for c, v in [(2,v1),(3,v2),(4,v3)]:
        if isinstance(v, str):
            sc(ws_a, row, c, f=v, nf=nf, align="center")
        else:
            sc(ws_a, row, c, v=v, nf=nf, align="center")
bdr(ws_a, 70, 75, 1, 4)
note_row(ws_a, 76, "※ Churn率・CVRの感度分析（悲観/中央/楽観シナリオ別）は「感度分析」シート参照", 4)

# ══════════════════════════════════════════════
# SHEET 2: 人員計画
# ══════════════════════════════════════════════
ws_h = wb.create_sheet("人員計画")
ws_h.column_dimensions["A"].width = 28
for col in ["B","C","D","E","F","G","H","I"]:
    ws_h.column_dimensions[col].width = 14

title_row(ws_h, 1, "UNLID 人員計画（採用計画）v4.2  ─  採用数・人件費がPL・感度分析に自動連動", 8)
sc(ws_h, 2, 1,
   "★ この表の人件費合計・採用費合計が「前提条件」シートのSG&A（行31,32）に自動参照される",
   bold=True, fc=C_WARN, size=9)
mrg(ws_h, 2, 1, 8)

# ─ 役割定義テーブル
sec_hdr(ws_h, 4, "① 役職定義・単価テーブル（入力エリア）", 8)
col_hdrs(ws_h, 5,
    ["役職", "月額給与(万円)", "法定福利費率", "採用費単価(万円)", "備考"],
    fc_def=C_BHDR)

ROLES = [
    ("CEO / 創業者",           30,  0.15,   0, "役員報酬（Y1）、段階的引上げ"),
    ("FDE（フラクショナルFDE）",45,  0.15, 140, "正社員化。副業→内部登用でコスト圧縮"),
    ("CSM（カスタマーサクセス）",40, 0.15, 100, "1名が25〜30社を担当"),
    ("PM（プロジェクトMgr）",   50,  0.15, 150, "プロジェクト5件/人目安"),
    ("AIエンジニア",            60,  0.15, 200, "自社ツール・大型PJ担当"),
    ("Admin / BizDev",          35,  0.15,  80, "バックオフィス・採用支援"),
]
for i, (role, salary, welf, recruit, note) in enumerate(ROLES, 6):
    sc(ws_h, i, 1, role, bold=True)
    sc(ws_h, i, 2, salary, nf=NF_MAN, align="center")
    sc(ws_h, i, 3, welf, nf=NF_PCT, align="center")
    sc(ws_h, i, 4, recruit, nf=NF_MAN, align="center")
    sc(ws_h, i, 5, note, italic=True, fc=C_NOTE, size=9)
bdr(ws_h, 5, 11, 1, 5)

# ─ 採用スケジュールテーブル
sec_hdr(ws_h, 13, "② 採用スケジュール（期末時点ヘッドカウント）", 8)
col_hdrs(ws_h, 14,
    ["役職", "Y1期首", "Y1採用", "Y1期末", "Y2採用", "Y2期末", "Y3採用", "Y3期末"],
    fc_def=C_BHDR)

# (role, Y1_start, Y1_hire, Y2_hire, Y3_hire)
HIRE = [
    ("CEO / 創業者",           1, 0, 0, 0),
    ("FDE（フラクショナルFDE）",0, 2, 8,22),
    ("CSM（カスタマーサクセス）",0,0, 3, 5),
    ("PM（プロジェクトMgr）",   0, 0, 2, 4),
    ("AIエンジニア",            0, 1, 3, 5),
    ("Admin / BizDev",          0, 0, 1, 2),
]

# ヘッドカウントをsheet上の数式で計算
ROW_ROLE_START = 15
for i, (role, y1s, y1h, y2h, y3h) in enumerate(HIRE):
    r = ROW_ROLE_START + i
    y1e = y1s + y1h
    y2e = y1e + y2h
    y3e = y2e + y3h
    sc(ws_h, r, 1, role, bold=True)
    sc(ws_h, r, 2, y1s, nf=NF_MAN, align="center")      # Y1期首
    sc(ws_h, r, 3, y1h, nf=NF_MAN, align="center")      # Y1採用
    sc(ws_h, r, 4, y1e, bold=True, nf=NF_MAN, align="center", fc=C_GPFT)  # Y1期末
    sc(ws_h, r, 5, y2h, nf=NF_MAN, align="center")      # Y2採用
    sc(ws_h, r, 6, y2e, bold=True, nf=NF_MAN, align="center", fc=C_GPFT)  # Y2期末
    sc(ws_h, r, 7, y3h, nf=NF_MAN, align="center")      # Y3採用
    sc(ws_h, r, 8, y3e, bold=True, nf=NF_MAN, align="center", fc=C_GPFT)  # Y3期末

# 合計行
r_tot = ROW_ROLE_START + len(HIRE)
sc(ws_h, r_tot, 1, "合計（人）", bold=True, fc=C_NAVY, color="FFFFFF")
for c in range(2, 9):
    sc(ws_h, r_tot, c, f=f"=SUM({gl(c)}{ROW_ROLE_START}:{gl(c)}{r_tot-1})",
       bold=True, fc=C_TOT, nf=NF_MAN, align="center")
bdr(ws_h, 14, r_tot, 1, 8)

# ─ 年次人件費計算テーブル（→前提条件シートが参照するキーセル）
sec_hdr(ws_h, 24, "③ 年次人件費・採用費サマリー（→前提条件シートに自動連動）", 8)
col_hdrs(ws_h, 25, ["役職", "月額給与(万)", "法定福利率", "Y1人件費", "Y2人件費", "Y3人件費"], fc_def=C_BHDR)

# 各役職の年次人件費 = 給与 × (1+法定福利費率) × 12 × 期末人数
# 給与・福利費率はROLES、期末人数はHIREから
salary_map = {r[0]: (r[1], r[2]) for r in ROLES}
headcount_map = {
    role: (y1s+y1h, y1s+y1h+y2h, y1s+y1h+y2h+y3h)
    for role, y1s, y1h, y2h, y3h in HIRE
}

for i, (role, salary, welf, recruit, _) in enumerate(ROLES):
    r = 26 + i
    y1e, y2e, y3e = headcount_map[role]
    y1_cost = round(salary * (1 + welf) * 12 * y1e)
    y2_cost = round(salary * (1 + welf) * 12 * y2e)
    y3_cost = round(salary * (1 + welf) * 12 * y3e)
    sc(ws_h, r, 1, role)
    sc(ws_h, r, 2, salary, nf=NF_MAN, align="center")
    sc(ws_h, r, 3, welf, nf=NF_PCT, align="center")
    sc(ws_h, r, 4, y1_cost, nf=NF_MAN, align="center")
    sc(ws_h, r, 5, y2_cost, nf=NF_MAN, align="center")
    sc(ws_h, r, 6, y3_cost, nf=NF_MAN, align="center")

# 人件費合計行（Row 32 = ROW 26+6）
r_sal_tot = 26 + len(ROLES)
sc(ws_h, r_sal_tot, 1, "人件費合計（万円）", bold=True, fc=C_NAVY, color="FFFFFF")
for c in [4, 5, 6]:
    sc(ws_h, r_sal_tot, c, f=f"=SUM({gl(c)}26:{gl(c)}{r_sal_tot-1})",
       bold=True, fc=C_TOT, nf=NF_MAN, align="center")
bdr(ws_h, 25, r_sal_tot, 1, 6)

# 採用費計算
sec_hdr(ws_h, 34, "④ 採用費サマリー（採用人数 × 採用費単価）", 8)
col_hdrs(ws_h, 35, ["役職", "採用費単価(万)", "", "Y1採用費", "Y2採用費", "Y3採用費"], fc_def=C_BHDR)

recruit_map = {r[0]: r[3] for r in ROLES}
for i, (role, y1s, y1h, y2h, y3h) in enumerate(HIRE):
    r = 36 + i
    recruit_fee = recruit_map[role]
    sc(ws_h, r, 1, role)
    sc(ws_h, r, 2, recruit_fee, nf=NF_MAN, align="center")
    sc(ws_h, r, 4, y1h * recruit_fee, nf=NF_MAN, align="center")
    sc(ws_h, r, 5, y2h * recruit_fee, nf=NF_MAN, align="center")
    sc(ws_h, r, 6, y3h * recruit_fee, nf=NF_MAN, align="center")

r_rec_tot = 36 + len(HIRE)
sc(ws_h, r_rec_tot, 1, "採用費合計（万円）", bold=True, fc=C_NAVY, color="FFFFFF")
for c in [4, 5, 6]:
    sc(ws_h, r_rec_tot, c, f=f"=SUM({gl(c)}36:{gl(c)}{r_rec_tot-1})",
       bold=True, fc=C_ORANGE, nf=NF_MAN, align="center")
bdr(ws_h, 35, r_rec_tot, 1, 6)

# ★ キーセル定義（前提条件シートからの参照先）
# D36=人件費Y1, E36=人件費Y2, F36=人件費Y3
# D37=採用費Y1, E37=採用費Y2, F37=採用費Y3
# → 前提条件シート行31,32がこれを参照する
# 実際のセルアドレス確認用にコメント
note_row(ws_h, r_rec_tot+1,
    f"★ 前提条件シート参照セル: 人件費={gl(4)}{r_sal_tot}〜{gl(6)}{r_sal_tot} / 採用費={gl(4)}{r_rec_tot}〜{gl(6)}{r_rec_tot}", 8)

# 前提条件の参照式を確定セルアドレスに合わせて修正
# （人員計画シートのD36・E36・F36 が人件費合計行）
# Part1終了 ─ Part2に続く
