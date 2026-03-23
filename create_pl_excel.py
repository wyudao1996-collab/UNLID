# -*- coding: utf-8 -*-
"""UNLID 3年間PL を Excel（.xlsx）形式で生成するスクリプト"""

from openpyxl import Workbook
from openpyxl.styles import (
    Font, Alignment, PatternFill, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "3年間PL"

# ── ヘルパー関数 ──────────────────────────────────────
def set_cell(ws, row, col, value=None, bold=False, italic=False,
             font_size=10, fill_color=None, align="left",
             number_format=None, formula=None):
    cell = ws.cell(row=row, column=col)
    if formula is not None:
        cell.value = formula
    elif value is not None:
        cell.value = value
    cell.font = Font(name="Meiryo UI", bold=bold, italic=italic, size=font_size)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    if fill_color:
        cell.fill = PatternFill("solid", fgColor=fill_color)
    if number_format:
        cell.number_format = number_format
    return cell

def thin_border():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

def apply_border_range(ws, min_row, max_row, min_col, max_col):
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            ws.cell(row=r, column=c).border = thin_border()

# カラー定義
C_TITLE   = "1F3864"  # 濃紺（タイトル行）
C_SECTION = "2E75B6"  # 青（セクション見出し）
C_HEADER  = "BDD7EE"  # 薄青（ヘッダー行）
C_SUBTOT  = "D9E1F2"  # 薄紫（小計行）
C_TOTAL   = "FFF2CC"  # 薄黄（合計行）
C_PROFIT  = "E2EFDA"  # 薄緑（利益行）
C_LOSS    = "FCE4D6"  # 薄橙（損失行）
C_NOTE    = "F2F2F2"  # 薄灰（根拠行）

# 列幅設定
ws.column_dimensions["A"].width = 36
ws.column_dimensions["B"].width = 16
ws.column_dimensions["C"].width = 16
ws.column_dimensions["D"].width = 16
ws.column_dimensions["E"].width = 16
ws.column_dimensions["F"].width = 16

# ── 行の高さ設定ヘルパー ──────────────────────────────
def rh(ws, row, h): ws.row_dimensions[row].height = h

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 1. タイトル（行1〜3）
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
set_cell(ws, 1, 1, "UNLID 3年間損益計算書（P&L）",
         bold=True, font_size=14, fill_color=C_TITLE, align="center")
ws.cell(row=1, column=1).font = Font(name="Meiryo UI", bold=True, size=14, color="FFFFFF")
ws.merge_cells("A1:F1")
rh(ws, 1, 28)

set_cell(ws, 2, 1, "作成日：2026年3月　｜　前提：v3.0事業戦略レポートに基づく",
         font_size=9, fill_color="D6E4F0", align="center")
ws.merge_cells("A2:F2")

set_cell(ws, 3, 1, "単位：万円（税引前）",
         font_size=9, italic=True, align="right")
ws.merge_cells("A3:F3")

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 2. 前提条件（行5〜13）
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
set_cell(ws, 5, 1, "前提条件（シナリオ：中央値ベース）",
         bold=True, fill_color=C_SECTION, align="center")
ws.cell(row=5, column=1).font = Font(name="Meiryo UI", bold=True, size=10, color="FFFFFF")
ws.merge_cells("A5:F5")

for col, label in [(1, ""), (2, "Year1（フェーズ1）"),
                   (3, "Year2（フェーズ2）"), (4, "Year3（フェーズ3）")]:
    set_cell(ws, 6, col, label, bold=True, fill_color=C_HEADER, align="center")

assumptions = [
    (7,  "FDE正社員数（期末）",          "2名",  "10名",  "32名"),
    (8,  "B2B契約企業数（期末）",        "10社", "50社",  "160社"),
    (9,  "FDE受託平均月額",              "15万円","18万円","20万円"),
    (10, "副業ワーカー数（累計）",       "30名", "300名", "3000名"),
    (11, "完遂ログ累計件数",             "60件", "400件", "1500件"),
    (12, "有料職業紹介業",               "未取得→年中取得", "本格運用", "本格運用"),
]
for row, label, y1, y2, y3 in assumptions:
    set_cell(ws, row, 1, label)
    set_cell(ws, row, 2, y1, align="center")
    set_cell(ws, row, 3, y2, align="center")
    set_cell(ws, row, 4, y3, align="center")

apply_border_range(ws, 6, 12, 1, 4)

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 3. 損益計算書（行15〜）
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
set_cell(ws, 14, 1, "損益計算書",
         bold=True, fill_color=C_SECTION, align="center")
ws.cell(row=14, column=1).font = Font(name="Meiryo UI", bold=True, size=10, color="FFFFFF")
ws.merge_cells("A14:F14")

for col, label in [(1, "項目"), (2, "Year1"), (3, "Year2"), (4, "Year3")]:
    set_cell(ws, 15, col, label, bold=True, fill_color=C_HEADER, align="center")
apply_border_range(ws, 15, 15, 1, 4)

# ─── 売上高 ───
set_cell(ws, 16, 1, "【売上高】", bold=True, fill_color="DAEEF3")
ws.merge_cells("A16:D16")

# B2B受託収益（フラクショナルFDE）
ROW_B2B = 17
set_cell(ws, ROW_B2B, 1, "  B2B受託収益（フラクショナルFDE）")
set_cell(ws, ROW_B2B, 2, 1800, number_format="#,##0")
set_cell(ws, ROW_B2B, 3, 10800, number_format="#,##0")
set_cell(ws, ROW_B2B, 4, 38400, number_format="#,##0")

set_cell(ws, 18, 1, "    └ 計算根拠", italic=True, fill_color=C_NOTE)
set_cell(ws, 18, 2, "10社×15万×12ヶ月", italic=True, fill_color=C_NOTE, align="center")
set_cell(ws, 18, 3, "50社×18万×12ヶ月", italic=True, fill_color=C_NOTE, align="center")
set_cell(ws, 18, 4, "160社×20万×12ヶ月", italic=True, fill_color=C_NOTE, align="center")

# 人材紹介フィー
ROW_JINZAI = 19
set_cell(ws, ROW_JINZAI, 1, "  人材紹介フィー（スカウト採用成功報酬）")
set_cell(ws, ROW_JINZAI, 2, 420, number_format="#,##0")
set_cell(ws, ROW_JINZAI, 3, 2800, number_format="#,##0")
set_cell(ws, ROW_JINZAI, 4, 7000, number_format="#,##0")

set_cell(ws, 20, 1, "    └ 計算根拠", italic=True, fill_color=C_NOTE)
set_cell(ws, 20, 2, "3件×140万円", italic=True, fill_color=C_NOTE, align="center")
set_cell(ws, 20, 3, "20件×140万円", italic=True, fill_color=C_NOTE, align="center")
set_cell(ws, 20, 4, "50件×140万円", italic=True, fill_color=C_NOTE, align="center")

# プロジェクト型受託
ROW_PROJ = 21
set_cell(ws, ROW_PROJ, 1, "  プロジェクト型受託（スポット案件）")
set_cell(ws, ROW_PROJ, 2, 0, number_format="#,##0")
set_cell(ws, ROW_PROJ, 3, 800, number_format="#,##0")
set_cell(ws, ROW_PROJ, 4, 2400, number_format="#,##0")

set_cell(ws, 22, 1, "    └ 計算根拠", italic=True, fill_color=C_NOTE)
set_cell(ws, 22, 2, "実績なし", italic=True, fill_color=C_NOTE, align="center")
set_cell(ws, 22, 3, "5件×160万円", italic=True, fill_color=C_NOTE, align="center")
set_cell(ws, 22, 4, "15件×160万円", italic=True, fill_color=C_NOTE, align="center")

# 売上合計
ROW_URIAGE = 23
set_cell(ws, ROW_URIAGE, 1, "売上合計", bold=True, fill_color=C_TOTAL)
set_cell(ws, ROW_URIAGE, 2, formula=f"=B{ROW_B2B}+B{ROW_JINZAI}+B{ROW_PROJ}",
         bold=True, fill_color=C_TOTAL, number_format="#,##0")
set_cell(ws, ROW_URIAGE, 3, formula=f"=C{ROW_B2B}+C{ROW_JINZAI}+C{ROW_PROJ}",
         bold=True, fill_color=C_TOTAL, number_format="#,##0")
set_cell(ws, ROW_URIAGE, 4, formula=f"=D{ROW_B2B}+D{ROW_JINZAI}+D{ROW_PROJ}",
         bold=True, fill_color=C_TOTAL, number_format="#,##0")

# 前年比成長率
ROW_GROWTH = 24
set_cell(ws, ROW_GROWTH, 1, "前年比成長率", italic=True)
set_cell(ws, ROW_GROWTH, 2, "—", align="center", italic=True)
set_cell(ws, ROW_GROWTH, 3,
         formula=f"=(C{ROW_URIAGE}-B{ROW_URIAGE})/B{ROW_URIAGE}",
         number_format="0%", italic=True, align="center")
set_cell(ws, ROW_GROWTH, 4,
         formula=f"=(D{ROW_URIAGE}-C{ROW_URIAGE})/C{ROW_URIAGE}",
         number_format="0%", italic=True, align="center")

# ─── 売上原価 ───
set_cell(ws, 26, 1, "【売上原価（変動費）】", bold=True, fill_color="DAEEF3")
ws.merge_cells("A26:D26")

ROW_WORKER = 27
set_cell(ws, ROW_WORKER, 1, "  副業ワーカー報酬")
set_cell(ws, ROW_WORKER, 2, 200, number_format="#,##0")
set_cell(ws, ROW_WORKER, 3, 1200, number_format="#,##0")
set_cell(ws, ROW_WORKER, 4, 4000, number_format="#,##0")

set_cell(ws, 28, 1, "    └ 計算根拠", italic=True, fill_color=C_NOTE)
set_cell(ws, 28, 2, "売上の9%", italic=True, fill_color=C_NOTE, align="center")
set_cell(ws, 28, 3, "売上の8%", italic=True, fill_color=C_NOTE, align="center")
set_cell(ws, 28, 4, "売上の8%", italic=True, fill_color=C_NOTE, align="center")

ROW_FDE_OUT = 29
set_cell(ws, ROW_FDE_OUT, 1, "  FDE外注・業務委託費（スポット補完）")
set_cell(ws, ROW_FDE_OUT, 2, 0, number_format="#,##0")
set_cell(ws, ROW_FDE_OUT, 3, 200, number_format="#,##0")
set_cell(ws, ROW_FDE_OUT, 4, 600, number_format="#,##0")

set_cell(ws, 30, 1, "    └ 計算根拠", italic=True, fill_color=C_NOTE)
set_cell(ws, 30, 2, "—", italic=True, fill_color=C_NOTE, align="center")
set_cell(ws, 30, 3, "スポット案件の外注", italic=True, fill_color=C_NOTE, align="center")
set_cell(ws, 30, 4, "スポット案件の外注", italic=True, fill_color=C_NOTE, align="center")

# 売上原価合計
ROW_COGS = 31
set_cell(ws, ROW_COGS, 1, "売上原価合計", bold=True, fill_color=C_SUBTOT)
set_cell(ws, ROW_COGS, 2, formula=f"=B{ROW_WORKER}+B{ROW_FDE_OUT}",
         bold=True, fill_color=C_SUBTOT, number_format="#,##0")
set_cell(ws, ROW_COGS, 3, formula=f"=C{ROW_WORKER}+C{ROW_FDE_OUT}",
         bold=True, fill_color=C_SUBTOT, number_format="#,##0")
set_cell(ws, ROW_COGS, 4, formula=f"=D{ROW_WORKER}+D{ROW_FDE_OUT}",
         bold=True, fill_color=C_SUBTOT, number_format="#,##0")

# 売上総利益（粗利）
ROW_GROSS = 33
set_cell(ws, ROW_GROSS, 1, "売上総利益（粗利）", bold=True, fill_color=C_PROFIT)
set_cell(ws, ROW_GROSS, 2, formula=f"=B{ROW_URIAGE}-B{ROW_COGS}",
         bold=True, fill_color=C_PROFIT, number_format="#,##0")
set_cell(ws, ROW_GROSS, 3, formula=f"=C{ROW_URIAGE}-C{ROW_COGS}",
         bold=True, fill_color=C_PROFIT, number_format="#,##0")
set_cell(ws, ROW_GROSS, 4, formula=f"=D{ROW_URIAGE}-D{ROW_COGS}",
         bold=True, fill_color=C_PROFIT, number_format="#,##0")

# 粗利率
ROW_GROSS_RATE = 34
set_cell(ws, ROW_GROSS_RATE, 1, "粗利率", italic=True)
set_cell(ws, ROW_GROSS_RATE, 2, formula=f"=B{ROW_GROSS}/B{ROW_URIAGE}",
         number_format="0%", italic=True, align="center")
set_cell(ws, ROW_GROSS_RATE, 3, formula=f"=C{ROW_GROSS}/C{ROW_URIAGE}",
         number_format="0%", italic=True, align="center")
set_cell(ws, ROW_GROSS_RATE, 4, formula=f"=D{ROW_GROSS}/D{ROW_URIAGE}",
         number_format="0%", italic=True, align="center")

# ─── 販管費 ───
set_cell(ws, 36, 1, "【販売費及び一般管理費（固定費）】", bold=True, fill_color="DAEEF3")
ws.merge_cells("A36:D36")

ROW_FDE_SAL = 37
set_cell(ws, ROW_FDE_SAL, 1, "  FDE正社員人件費（給与・社保）")
set_cell(ws, ROW_FDE_SAL, 2, 1080, number_format="#,##0")
set_cell(ws, ROW_FDE_SAL, 3, 4800, number_format="#,##0")
set_cell(ws, ROW_FDE_SAL, 4, 16200, number_format="#,##0")

set_cell(ws, 38, 1, "    └ 計算根拠", italic=True, fill_color=C_NOTE)
set_cell(ws, 38, 2, "2名×月45万×12ヶ月", italic=True, fill_color=C_NOTE, align="center")
set_cell(ws, 38, 3, "10名×月40万×12ヶ月", italic=True, fill_color=C_NOTE, align="center")
set_cell(ws, 38, 4, "32名×月42万×12ヶ月", italic=True, fill_color=C_NOTE, align="center")

ROW_FOUNDER = 39
set_cell(ws, ROW_FOUNDER, 1, "  創業者報酬")
set_cell(ws, ROW_FOUNDER, 2, 360, number_format="#,##0")
set_cell(ws, ROW_FOUNDER, 3, 600, number_format="#,##0")
set_cell(ws, ROW_FOUNDER, 4, 840, number_format="#,##0")

set_cell(ws, 40, 1, "    └ 計算根拠", italic=True, fill_color=C_NOTE)
set_cell(ws, 40, 2, "月30万×12ヶ月", italic=True, fill_color=C_NOTE, align="center")
set_cell(ws, 40, 3, "月50万×12ヶ月", italic=True, fill_color=C_NOTE, align="center")
set_cell(ws, 40, 4, "月70万×12ヶ月", italic=True, fill_color=C_NOTE, align="center")

ROW_ADMIN = 41
set_cell(ws, ROW_ADMIN, 1, "  管理・営業スタッフ人件費")
set_cell(ws, ROW_ADMIN, 2, 0, number_format="#,##0")
set_cell(ws, ROW_ADMIN, 3, 1080, number_format="#,##0")
set_cell(ws, ROW_ADMIN, 4, 3360, number_format="#,##0")

set_cell(ws, 42, 1, "    └ 計算根拠", italic=True, fill_color=C_NOTE)
set_cell(ws, 42, 2, "なし（創業者が担当）", italic=True, fill_color=C_NOTE, align="center")
set_cell(ws, 42, 3, "3名×月30万×12", italic=True, fill_color=C_NOTE, align="center")
set_cell(ws, 42, 4, "8名×月35万×12", italic=True, fill_color=C_NOTE, align="center")

ROW_OFFICE = 43
set_cell(ws, ROW_OFFICE, 1, "  オフィス・ツール・インフラ費")
set_cell(ws, ROW_OFFICE, 2, 180, number_format="#,##0")
set_cell(ws, ROW_OFFICE, 3, 360, number_format="#,##0")
set_cell(ws, ROW_OFFICE, 4, 600, number_format="#,##0")

set_cell(ws, 44, 1, "    └ 計算根拠", italic=True, fill_color=C_NOTE)
set_cell(ws, 44, 2, "月15万×12ヶ月", italic=True, fill_color=C_NOTE, align="center")
set_cell(ws, 44, 3, "月30万×12ヶ月", italic=True, fill_color=C_NOTE, align="center")
set_cell(ws, 44, 4, "月50万×12ヶ月", italic=True, fill_color=C_NOTE, align="center")

ROW_MKT = 45
set_cell(ws, ROW_MKT, 1, "  マーケティング・営業費")
set_cell(ws, ROW_MKT, 2, 120, number_format="#,##0")
set_cell(ws, ROW_MKT, 3, 360, number_format="#,##0")
set_cell(ws, ROW_MKT, 4, 600, number_format="#,##0")

set_cell(ws, 46, 1, "    └ 計算根拠", italic=True, fill_color=C_NOTE)
set_cell(ws, 46, 2, "月10万×12ヶ月", italic=True, fill_color=C_NOTE, align="center")
set_cell(ws, 46, 3, "月30万×12ヶ月", italic=True, fill_color=C_NOTE, align="center")
set_cell(ws, 46, 4, "月50万×12ヶ月", italic=True, fill_color=C_NOTE, align="center")

ROW_LEGAL = 47
set_cell(ws, ROW_LEGAL, 1, "  法務・許認可・顧問弁護士")
set_cell(ws, ROW_LEGAL, 2, 200, number_format="#,##0")
set_cell(ws, ROW_LEGAL, 3, 300, number_format="#,##0")
set_cell(ws, ROW_LEGAL, 4, 400, number_format="#,##0")

set_cell(ws, 48, 1, "    └ 計算根拠", italic=True, fill_color=C_NOTE)
set_cell(ws, 48, 2, "労働法精査・職業紹介許可申請", italic=True, fill_color=C_NOTE, align="center")
set_cell(ws, 48, 3, "継続顧問", italic=True, fill_color=C_NOTE, align="center")
set_cell(ws, 48, 4, "継続顧問", italic=True, fill_color=C_NOTE, align="center")

ROW_RECRUIT = 49
set_cell(ws, ROW_RECRUIT, 1, "  採用費（FDE・スタッフ）")
set_cell(ws, ROW_RECRUIT, 2, 0, number_format="#,##0")
set_cell(ws, ROW_RECRUIT, 3, 500, number_format="#,##0")
set_cell(ws, ROW_RECRUIT, 4, 1500, number_format="#,##0")

set_cell(ws, 50, 1, "    └ 計算根拠", italic=True, fill_color=C_NOTE)
set_cell(ws, 50, 2, "リファラルのみ", italic=True, fill_color=C_NOTE, align="center")
set_cell(ws, 50, 3, "FDE採用8名分", italic=True, fill_color=C_NOTE, align="center")
set_cell(ws, 50, 4, "FDE採用22名分", italic=True, fill_color=C_NOTE, align="center")

ROW_SYS = 51
set_cell(ws, ROW_SYS, 1, "  システム開発・完遂ログプラットフォーム")
set_cell(ws, ROW_SYS, 2, 100, number_format="#,##0")
set_cell(ws, ROW_SYS, 3, 300, number_format="#,##0")
set_cell(ws, ROW_SYS, 4, 800, number_format="#,##0")

set_cell(ws, 52, 1, "    └ 計算根拠", italic=True, fill_color=C_NOTE)
set_cell(ws, 52, 2, "PoC・ノーコード", italic=True, fill_color=C_NOTE, align="center")
set_cell(ws, 52, 3, "本格開発", italic=True, fill_color=C_NOTE, align="center")
set_cell(ws, 52, 4, "スケール対応", italic=True, fill_color=C_NOTE, align="center")

ROW_MISC = 53
set_cell(ws, ROW_MISC, 1, "  その他（交通・研修・保険等）")
set_cell(ws, ROW_MISC, 2, 100, number_format="#,##0")
set_cell(ws, ROW_MISC, 3, 200, number_format="#,##0")
set_cell(ws, ROW_MISC, 4, 400, number_format="#,##0")

# 販管費合計
ROW_SGA = 54
set_cell(ws, ROW_SGA, 1, "販管費合計", bold=True, fill_color=C_SUBTOT)
sga_rows = [ROW_FDE_SAL, ROW_FOUNDER, ROW_ADMIN, ROW_OFFICE,
            ROW_MKT, ROW_LEGAL, ROW_RECRUIT, ROW_SYS, ROW_MISC]
for col_letter in ["B", "C", "D"]:
    refs = "+".join([f"{col_letter}{r}" for r in sga_rows])
    set_cell(ws, ROW_SGA, ord(col_letter) - ord("A") + 1,
             formula=f"={refs}", bold=True, fill_color=C_SUBTOT, number_format="#,##0")

# 営業利益（EBIT）
ROW_EBIT = 56
set_cell(ws, ROW_EBIT, 1, "営業利益（EBIT）", bold=True, fill_color=C_PROFIT)
for i, col_letter in enumerate(["B", "C", "D"]):
    fill = C_PROFIT if i > 0 else C_LOSS  # Year1は赤字
    set_cell(ws, ROW_EBIT, ord(col_letter) - ord("A") + 1,
             formula=f"={col_letter}{ROW_GROSS}-{col_letter}{ROW_SGA}",
             bold=True, fill_color=fill, number_format="#,##0")
# Year1は赤字なので色を変える（実際の値に依存するため、ここではロジックなし）

# 営業利益率
ROW_EBIT_RATE = 57
set_cell(ws, ROW_EBIT_RATE, 1, "営業利益率", italic=True)
for col_letter in ["B", "C", "D"]:
    set_cell(ws, ROW_EBIT_RATE, ord(col_letter) - ord("A") + 1,
             formula=f"={col_letter}{ROW_EBIT}/{col_letter}{ROW_URIAGE}",
             number_format="0%", italic=True, align="center")

# 減価償却費
ROW_DA = 59
set_cell(ws, ROW_DA, 1, "  減価償却費（参考値）", italic=True)
set_cell(ws, ROW_DA, 2, 20, number_format="#,##0", italic=True)
set_cell(ws, ROW_DA, 3, 80, number_format="#,##0", italic=True)
set_cell(ws, ROW_DA, 4, 200, number_format="#,##0", italic=True)

# EBITDA
ROW_EBITDA = 60
set_cell(ws, ROW_EBITDA, 1, "EBITDA", bold=True, fill_color=C_PROFIT)
for col_letter in ["B", "C", "D"]:
    set_cell(ws, ROW_EBITDA, ord(col_letter) - ord("A") + 1,
             formula=f"={col_letter}{ROW_EBIT}+{col_letter}{ROW_DA}",
             bold=True, fill_color=C_PROFIT, number_format="#,##0")

# EBITDAマージン
ROW_EBITDA_RATE = 61
set_cell(ws, ROW_EBITDA_RATE, 1, "EBITDAマージン", italic=True)
for col_letter in ["B", "C", "D"]:
    set_cell(ws, ROW_EBITDA_RATE, ord(col_letter) - ord("A") + 1,
             formula=f"={col_letter}{ROW_EBITDA}/{col_letter}{ROW_URIAGE}",
             number_format="0%", italic=True, align="center")

# 法人税等
ROW_TAX = 63
set_cell(ws, ROW_TAX, 1, "  法人税等（実効税率30%想定）")
for col_letter in ["B", "C", "D"]:
    set_cell(ws, ROW_TAX, ord(col_letter) - ord("A") + 1,
             formula=f"=IF({col_letter}{ROW_EBIT}>0,ROUND({col_letter}{ROW_EBIT}*0.3,0),0)",
             number_format="#,##0")

set_cell(ws, 64, 1, "    └ 計算根拠", italic=True, fill_color=C_NOTE)
set_cell(ws, 64, 2, "赤字のため税なし", italic=True, fill_color=C_NOTE, align="center")
set_cell(ws, 64, 3, "EBIT×30%", italic=True, fill_color=C_NOTE, align="center")
set_cell(ws, 64, 4, "EBIT×30%", italic=True, fill_color=C_NOTE, align="center")

# 当期純利益
ROW_NET = 65
set_cell(ws, ROW_NET, 1, "当期純利益", bold=True, fill_color=C_PROFIT)
for i, col_letter in enumerate(["B", "C", "D"]):
    fill = C_PROFIT if i > 0 else C_LOSS
    set_cell(ws, ROW_NET, ord(col_letter) - ord("A") + 1,
             formula=f"={col_letter}{ROW_EBIT}-{col_letter}{ROW_TAX}",
             bold=True, fill_color=fill, number_format="#,##0")

# 純利益率
ROW_NET_RATE = 66
set_cell(ws, ROW_NET_RATE, 1, "純利益率", italic=True)
for col_letter in ["B", "C", "D"]:
    set_cell(ws, ROW_NET_RATE, ord(col_letter) - ord("A") + 1,
             formula=f"={col_letter}{ROW_NET}/{col_letter}{ROW_URIAGE}",
             number_format="0%", italic=True, align="center")

apply_border_range(ws, 15, ROW_NET_RATE, 1, 4)

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 4. 四半期別売上推移（Year1）（行68〜73）
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
set_cell(ws, 68, 1, "四半期別売上推移（Year1）",
         bold=True, fill_color=C_SECTION, align="center")
ws.cell(row=68, column=1).font = Font(name="Meiryo UI", bold=True, size=10, color="FFFFFF")
ws.merge_cells("A68:F68")

for col, label in [(1,""), (2,"Q1（月1〜3）"), (3,"Q2（月4〜6）"),
                   (4,"Q3（月7〜9）"), (5,"Q4（月10〜12）"), (6,"Year1合計")]:
    set_cell(ws, 69, col, label, bold=True, fill_color=C_HEADER, align="center")

quarterly = [
    (70, "B2B受託",       225, 360, 540, 675),
    (71, "人材紹介フィー",  0,   0,  140, 280),
    (72, "プロジェクト型",  0,   0,    0,   0),
]
for row, label, q1, q2, q3, q4 in quarterly:
    set_cell(ws, row, 1, label)
    set_cell(ws, row, 2, q1, number_format="#,##0")
    set_cell(ws, row, 3, q2, number_format="#,##0")
    set_cell(ws, row, 4, q3, number_format="#,##0")
    set_cell(ws, row, 5, q4, number_format="#,##0")
    set_cell(ws, row, 6, formula=f"=SUM(B{row}:E{row})",
             bold=True, fill_color=C_TOTAL, number_format="#,##0")

ROW_Q_TOTAL = 73
set_cell(ws, ROW_Q_TOTAL, 1, "売上合計", bold=True, fill_color=C_TOTAL)
for col in range(2, 7):
    col_letter = get_column_letter(col)
    set_cell(ws, ROW_Q_TOTAL, col,
             formula=f"=SUM({col_letter}70:{col_letter}72)",
             bold=True, fill_color=C_TOTAL, number_format="#,##0")

apply_border_range(ws, 69, ROW_Q_TOTAL, 1, 6)

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 5. 月次キャッシュフロー概算（Year1）（行75〜89）
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
set_cell(ws, 75, 1, "月次キャッシュフロー概算（Year1）",
         bold=True, fill_color=C_SECTION, align="center")
ws.cell(row=75, column=1).font = Font(name="Meiryo UI", bold=True, size=10, color="FFFFFF")
ws.merge_cells("A75:F75")

for col, label in [(1,"月"), (2,"売上（万円）"), (3,"人件費（万円）"),
                   (4,"その他コスト（万円）"), (5,"月次CF（万円）"), (6,"累計CF（万円）")]:
    set_cell(ws, 76, col, label, bold=True, fill_color=C_HEADER, align="center")

monthly_data = [
    (77,  1,  75, 120, 30),
    (78,  2,  75, 120, 30),
    (79,  3,  75, 120, 30),
    (80,  4, 120, 120, 30),
    (81,  5, 120, 120, 30),
    (82,  6, 120, 120, 30),
    (83,  7, 150, 160, 40),
    (84,  8, 165, 160, 40),
    (85,  9, 195, 160, 40),
    (86, 10, 255, 180, 50),
    (87, 11, 330, 180, 50),
    (88, 12, 370, 180, 50),
]
for row, month, sales, labor, other in monthly_data:
    set_cell(ws, row, 1, month, align="center")
    set_cell(ws, row, 2, sales, number_format="#,##0")
    set_cell(ws, row, 3, labor, number_format="#,##0")
    set_cell(ws, row, 4, other, number_format="#,##0")
    # 月次CF = 売上 - 人件費 - その他コスト
    set_cell(ws, row, 5, formula=f"=B{row}-C{row}-D{row}", number_format="#,##0")
    # 累計CF
    if row == 77:
        set_cell(ws, row, 6, formula=f"=E{row}", number_format="#,##0")
    else:
        set_cell(ws, row, 6, formula=f"=F{row-1}+E{row}", number_format="#,##0")

set_cell(ws, 89, 1,
         "（注：Q1-Q2は初期クライアント開拓中。Q3以降から人材紹介フィーが発生）",
         italic=True, fill_color=C_NOTE)
ws.merge_cells("A89:F89")

apply_border_range(ws, 76, 88, 1, 6)

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 6. KPI・マイルストーン（行91〜100）
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
set_cell(ws, 91, 1, "KPI・マイルストーン",
         bold=True, fill_color=C_SECTION, align="center")
ws.cell(row=91, column=1).font = Font(name="Meiryo UI", bold=True, size=10, color="FFFFFF")
ws.merge_cells("A91:D91")

for col, label in [(1,"指標"), (2,"Year1目標"), (3,"Year2目標"), (4,"Year3目標")]:
    set_cell(ws, 92, col, label, bold=True, fill_color=C_HEADER, align="center")

kpi_data = [
    (93,  "B2B契約企業数",          "10社",   "50社",   "160社"),
    (94,  "B2B継続率（3ヶ月以上）",  "60%以上", "70%以上", "80%以上"),
    (95,  "FDE正社員数",             "2名",    "10名",   "32名"),
    (96,  "副業ワーカー月間アクティブ数", "20名", "150名", "1000名"),
    (97,  "月間完遂件数",            "5件",    "30件",   "100件"),
    (98,  "完遂ログ累計",            "60件",   "400件",  "1500件"),
    (99,  "人材紹介成立件数",        "3件",    "20件",   "50件"),
    (100, "3軸平均スコア",           "3.5以上", "3.8以上", "4.0以上"),
]
for row, label, y1, y2, y3 in kpi_data:
    set_cell(ws, row, 1, label)
    set_cell(ws, row, 2, y1, align="center")
    set_cell(ws, row, 3, y2, align="center")
    set_cell(ws, row, 4, y3, align="center")

apply_border_range(ws, 92, 100, 1, 4)

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 7. 損益感度分析（Year2）（行102〜106）
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
set_cell(ws, 102, 1, "損益感度分析（Year2）",
         bold=True, fill_color=C_SECTION, align="center")
ws.cell(row=102, column=1).font = Font(name="Meiryo UI", bold=True, size=10, color="FFFFFF")
ws.merge_cells("A102:F102")

for col, label in [(1,"シナリオ"), (2,"B2B社数"), (3,"月額"),
                   (4,"人材紹介件数"), (5,"Year2売上"), (6,"Year2営業利益")]:
    set_cell(ws, 103, col, label, bold=True, fill_color=C_HEADER, align="center")

sensitivity = [
    (104, "楽観シナリオ", "65社", "20万円", "30件", 19800, 9600),
    (105, "中央値シナリオ", "50社", "18万円", "20件", 14400, 4500),
    (106, "保守シナリオ", "30社", "15万円", "10件", 6800, -1000),
]
for row, label, companies, price, intro, sales, profit in sensitivity:
    fill = "E2EFDA" if profit >= 0 else "FCE4D6"
    set_cell(ws, row, 1, label, fill_color=fill)
    set_cell(ws, row, 2, companies, align="center", fill_color=fill)
    set_cell(ws, row, 3, price, align="center", fill_color=fill)
    set_cell(ws, row, 4, intro, align="center", fill_color=fill)
    set_cell(ws, row, 5, sales, number_format="#,##0万円", fill_color=fill)
    set_cell(ws, row, 6, profit, number_format="#,##0万円", fill_color=fill)

apply_border_range(ws, 103, 106, 1, 6)

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 8. 資金調達シナリオ（行108〜112）
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
set_cell(ws, 108, 1, "資金調達シナリオ",
         bold=True, fill_color=C_SECTION, align="center")
ws.cell(row=108, column=1).font = Font(name="Meiryo UI", bold=True, size=10, color="FFFFFF")
ws.merge_cells("A108:D108")

for col, label in [(1,"フェーズ"), (2,"必要資金"), (3,"調達方法"), (4,"用途")]:
    set_cell(ws, 109, col, label, bold=True, fill_color=C_HEADER, align="center")

funding = [
    (110, "フェーズ0〜1（〜12ヶ月）", "2,000万円",
     "エンジェル・自己資金", "初期人件費・法的精査・システム開発"),
    (111, "フェーズ2（1〜3年）", "1億円（シリーズA相当）",
     "VC調達", "FDE採用加速・マーケティング・SaaS開発"),
    (112, "フェーズ3（3〜5年）", "5億円（シリーズB相当）",
     "VC調達", "完遂ログSaaS化・地方展開"),
]
for row, phase, amount, method, purpose in funding:
    set_cell(ws, row, 1, phase)
    set_cell(ws, row, 2, amount, align="center")
    set_cell(ws, row, 3, method, align="center")
    set_cell(ws, row, 4, purpose)

apply_border_range(ws, 109, 112, 1, 4)

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 9. 重要前提・注意事項（行114〜120）
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
set_cell(ws, 114, 1, "重要前提・注意事項",
         bold=True, fill_color=C_SECTION, align="center")
ws.cell(row=114, column=1).font = Font(name="Meiryo UI", bold=True, size=10, color="FFFFFF")
ws.merge_cells("A114:F114")

notes = [
    (115, "1. 本PLは中央値シナリオ（楽観でも悲観でもない）に基づく"),
    (116, "2. 有料職業紹介業許可の取得をYear1中に完了すると仮定（遅延リスクあり）"),
    (117, "3. FDE採用が計画通り進まない場合、Year2以降の売上は大幅下振れのリスクあり"),
    (118, "4. 消費税・インボイス対応は別途計上（本PLは税引前）"),
    (119, "5. 完遂ログSaaS化は本PLに含まない（Year4以降の別途試算が必要）"),
    (120, "6. 資金調達なしのブートストラップ場合、Year1末の累計CF▲140万円は自己資金でカバーする必要がある"),
]
for row, text in notes:
    set_cell(ws, row, 1, text, fill_color=C_NOTE)
    ws.merge_cells(f"A{row}:F{row}")

# ── 印刷設定 ──────────────────────────────────
ws.print_title_rows = "1:3"
ws.page_setup.fitToPage = True
ws.page_setup.fitToWidth = 1
ws.freeze_panes = "B16"  # 項目列と損益計算書ヘッダーを固定

# ── 保存 ──────────────────────────────────────
output_path = "C:/Users/user/Desktop/UNLID/UNLID_3年間PL.xlsx"
wb.save(output_path)
print(f"Excel ファイルを作成しました: {output_path}")
