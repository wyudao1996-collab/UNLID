# -*- coding: utf-8 -*-
"""UNLID 財務三表（PL/BS/CF）+ 感度分析 Excel生成スクリプト"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from openpyxl.styles import PatternFill as PF

wb = Workbook()

# ══════════════════════════════════════════════════
# ヘルパー関数
# ══════════════════════════════════════════════════
def sc(ws, row, col, value=None, formula=None, bold=False, italic=False,
       size=10, fc=None, align="left", nf=None, color="000000"):
    cell = ws.cell(row=row, column=col)
    cell.value = formula if formula else value
    cell.font = Font(name="Meiryo UI", bold=bold, italic=italic,
                     size=size, color=color)
    cell.alignment = Alignment(horizontal=align, vertical="center",
                               wrap_text=True)
    if fc:
        cell.fill = PatternFill("solid", fgColor=fc)
    if nf:
        cell.number_format = nf
    return cell

def border_range(ws, r1, r2, c1, c2, style="thin"):
    s = Side(style=style)
    b = Border(left=s, right=s, top=s, bottom=s)
    for r in range(r1, r2+1):
        for c in range(c1, c2+1):
            ws.cell(row=r, column=c).border = b

def thick_bottom(ws, row, c1, c2):
    for c in range(c1, c2+1):
        cell = ws.cell(row=row, column=c)
        existing = cell.border
        s_thin = Side(style="thin")
        s_thick = Side(style="medium")
        cell.border = Border(left=s_thin, right=s_thin,
                             top=s_thin, bottom=s_thick)

def hdr(ws, row, labels, fills=None, bold=True, size=10):
    for i, lbl in enumerate(labels):
        col = i + 1
        f = fills[i] if fills else None
        sc(ws, row, col, lbl, bold=bold, size=size,
           fc=f, align="center", color="FFFFFF" if f in
           ["1F3864","2E75B6","1F4E79"] else "000000")

def rh(ws, row, h): ws.row_dimensions[row].height = h

# カラーパレット
C_DARK   = "1F3864"
C_BLUE   = "2E75B6"
C_LBLUE  = "BDD7EE"
C_DBLUE  = "1F4E79"
C_SUB    = "D9E1F2"
C_TOT    = "FFF2CC"
C_PROFT  = "E2EFDA"
C_LOSS   = "FCE4D6"
C_NOTE   = "F2F2F2"
C_ORANGE = "F4B942"
C_GREEN  = "70AD47"
C_GRAY   = "808080"

NF_MAN = '#,##0'
NF_PCT = '0.0%'

# ══════════════════════════════════════════════════
# SHEET 1: 前提条件（Assumptions）
# ══════════════════════════════════════════════════
ws_a = wb.active
ws_a.title = "前提条件"

ws_a.column_dimensions["A"].width = 32
for col in ["B","C","D"]:
    ws_a.column_dimensions[col].width = 16

# タイトル
sc(ws_a, 1, 1, "UNLID 財務モデル 前提条件", bold=True, size=14,
   fc=C_DARK, color="FFFFFF", align="center")
ws_a.merge_cells("A1:D1"); rh(ws_a, 1, 28)
sc(ws_a, 2, 1, "作成日：2026年3月  |  単位：万円  |  中央値シナリオ",
   fc="D6E4F0", align="center", size=9, italic=True)
ws_a.merge_cells("A2:D2")

# ─ セクション①: 売上ドライバー
sc(ws_a, 4, 1, "① 売上ドライバー", bold=True, fc=C_BLUE,
   color="FFFFFF", align="center")
ws_a.merge_cells("A4:D4")
for c,l in [(1,"項目"),(2,"Year1"),(3,"Year2"),(4,"Year3")]:
    sc(ws_a, 5, c, l, bold=True, fc=C_LBLUE, align="center")

rows_a = [
    (6,  "B2B契約企業数（期末）",      10,    50,   160),
    (7,  "FDE受託平均月額（万円）",     15,    18,    20),
    (8,  "B2B受託収益（万円）",    "=B6*B7*12","=C6*C7*12","=D6*D7*12"),
    (9,  "人材紹介件数",                3,    20,    50),
    (10, "紹介フィー単価（万円）",     140,   140,   140),
    (11, "人材紹介フィー（万円）",  "=B9*B10","=C9*C10","=D9*D10"),
    (12, "プロジェクト件数",             0,     5,    15),
    (13, "プロジェクト単価（万円）",     0,   160,   160),
    (14, "プロジェクト収益（万円）", "=B12*B13","=C12*C13","=D12*D13"),
    (15, "売上合計（万円）","=B8+B11+B14","=C8+C11+C14","=D8+D11+D14"),
]
for row, label, v1, v2, v3 in rows_a:
    bold = label in ["B2B受託収益（万円）","人材紹介フィー（万円）",
                     "プロジェクト収益（万円）","売上合計（万円）"]
    fc = C_TOT if "売上合計" in label else (C_SUB if bold else None)
    def fill_v(v):
        return (None, v) if isinstance(v, str) else (v, None)
    for c, v in [(2,v1),(3,v2),(4,v3)]:
        val, frm = (None, v) if isinstance(v, str) else (v, None)
        sc(ws_a, row, c, value=val, formula=frm,
           bold=bold, fc=fc, nf=NF_MAN, align="center")
    sc(ws_a, row, 1, label, bold=bold, fc=fc)
border_range(ws_a, 5, 15, 1, 4)

# ─ セクション②: コストドライバー
sc(ws_a, 17, 1, "② コストドライバー", bold=True, fc=C_BLUE,
   color="FFFFFF", align="center")
ws_a.merge_cells("A17:D17")
for c,l in [(1,"項目"),(2,"Year1"),(3,"Year2"),(4,"Year3")]:
    sc(ws_a, 18, c, l, bold=True, fc=C_LBLUE, align="center")

cost_rows = [
    (19, "FDE正社員数",            2,    10,   32),
    (20, "FDE月額人件費（万円）",  45,    40,   42),
    (21, "FDE人件費合計（万円）","=B19*B20*12","=C19*C20*12","=D19*D20*12"),
    (22, "創業者報酬（万円）",    360,   600,  840),
    (23, "管理スタッフ人件費",      0,  1080, 3360),
    (24, "人件費合計（万円）","=B21+B22+B23","=C21+C22+C23","=D21+D22+D23"),
    (25, "ワーカー報酬率（%）",   0.09,  0.08, 0.08),
    (26, "ワーカー報酬（万円）","=前提条件!B15*B25","=前提条件!C15*C25","=前提条件!D15*D25"),
    (27, "オフィス・インフラ費",  180,   360,  600),
    (28, "マーケティング費",       120,   360,  600),
    (29, "法務・許認可費",         200,   300,  400),
    (30, "採用費",                   0,   500, 1500),
    (31, "システム開発費",          100,   300,  800),
    (32, "その他",                  100,   200,  400),
    (33, "販管費合計（万円）","=B24+B27+B28+B29+B30+B31+B32",
                              "=C24+C27+C28+C29+C30+C31+C32",
                              "=D24+D27+D28+D29+D30+D31+D32"),
]
for row, label, v1, v2, v3 in cost_rows:
    bold = "合計" in label
    fc = C_SUB if bold else None
    nf = NF_PCT if "率" in label else NF_MAN
    for c, v in [(2,v1),(3,v2),(4,v3)]:
        val, frm = (None, v) if isinstance(v, str) else (v, None)
        sc(ws_a, row, c, value=val, formula=frm,
           bold=bold, fc=fc, nf=nf, align="center")
    sc(ws_a, row, 1, label, bold=bold, fc=fc)
border_range(ws_a, 18, 33, 1, 4)

# ─ セクション③: その他前提
sc(ws_a, 35, 1, "③ その他前提", bold=True, fc=C_BLUE, color="FFFFFF", align="center")
ws_a.merge_cells("A35:D35")
for c,l in [(1,"項目"),(2,"Year1"),(3,"Year2"),(4,"Year3")]:
    sc(ws_a, 36, c, l, bold=True, fc=C_LBLUE, align="center")
other = [
    (37, "実効税率",                   0.30,  0.30,  0.30),
    (38, "減価償却費（万円）",            20,    80,   200),
    (39, "設備投資（Capex）（万円）",    100,   300,   800),
    (40, "初期調達資本（万円）",        2000,     0,     0),
    (41, "運転資本回転日数（売掛）",      30,    30,    30),
    (42, "買掛金回転日数",               15,    15,    15),
]
for row, label, v1, v2, v3 in other:
    nf = NF_PCT if "率" in label else NF_MAN
    for c, v in [(2,v1),(3,v2),(4,v3)]:
        sc(ws_a, row, c, value=v, nf=nf, align="center")
    sc(ws_a, row, 1, label)
border_range(ws_a, 36, 42, 1, 4)

ws_a.freeze_panes = "B5"


# ══════════════════════════════════════════════════
# SHEET 2: PL（損益計算書）
# ══════════════════════════════════════════════════
ws_pl = wb.create_sheet("PL（損益計算書）")
ws_pl.column_dimensions["A"].width = 36
for col in ["B","C","D","E"]:
    ws_pl.column_dimensions[col].width = 16

# タイトル
sc(ws_pl, 1, 1, "UNLID 損益計算書（P&L）3ヶ年", bold=True, size=14,
   fc=C_DARK, color="FFFFFF", align="center")
ws_pl.merge_cells("A1:E1"); rh(ws_pl, 1, 28)
sc(ws_pl, 2, 1, "単位：万円（税引前）  |  前提条件シートの数値と連動",
   fc="D6E4F0", align="center", size=9, italic=True)
ws_pl.merge_cells("A1:E1")
ws_pl.merge_cells("A2:E2")

# ヘッダー
for c,l in [(1,"項目"),(2,"Year1"),(3,"Year2"),(4,"Year3"),(5,"3年累計")]:
    sc(ws_pl, 3, c, l, bold=True, fc=C_DBLUE, align="center", color="FFFFFF")

# ─ 売上高
sc(ws_pl, 4, 1, "【売上高】", bold=True, fc="DAEEF3")
ws_pl.merge_cells("A4:E4")

PL = {}  # 行番号を記録

PL["b2b"]   = 5
PL["jinzai"]= 6
PL["proj"]  = 7
PL["uriage"]= 8
PL["growth"]= 9

rows_pl_rev = [
    (PL["b2b"],    "  B2B受託収益",           "=前提条件!B8",  "=前提条件!C8",  "=前提条件!D8"),
    (PL["jinzai"], "  人材紹介フィー",          "=前提条件!B11", "=前提条件!C11", "=前提条件!D11"),
    (PL["proj"],   "  プロジェクト型受託",      "=前提条件!B14", "=前提条件!C14", "=前提条件!D14"),
]
for row, label, f1, f2, f3 in rows_pl_rev:
    sc(ws_pl, row, 1, label)
    sc(ws_pl, row, 2, formula=f1, nf=NF_MAN)
    sc(ws_pl, row, 3, formula=f2, nf=NF_MAN)
    sc(ws_pl, row, 4, formula=f3, nf=NF_MAN)
    sc(ws_pl, row, 5, formula=f"=SUM(B{row}:D{row})", nf=NF_MAN, fc=C_SUB)

sc(ws_pl, PL["uriage"], 1, "売上合計", bold=True, fc=C_TOT)
for c, col in [(2,"B"),(3,"C"),(4,"D")]:
    sc(ws_pl, PL["uriage"], c,
       formula=f"=SUM({col}{PL['b2b']}:{col}{PL['proj']})",
       bold=True, fc=C_TOT, nf=NF_MAN)
sc(ws_pl, PL["uriage"], 5,
   formula=f"=SUM(B{PL['uriage']}:D{PL['uriage']})",
   bold=True, fc=C_TOT, nf=NF_MAN)

sc(ws_pl, PL["growth"], 1, "前年比成長率", italic=True)
sc(ws_pl, PL["growth"], 2, "—", align="center", italic=True)
sc(ws_pl, PL["growth"], 3,
   formula=f"=(C{PL['uriage']}-B{PL['uriage']})/B{PL['uriage']}",
   nf=NF_PCT, align="center", italic=True)
sc(ws_pl, PL["growth"], 4,
   formula=f"=(D{PL['uriage']}-C{PL['uriage']})/C{PL['uriage']}",
   nf=NF_PCT, align="center", italic=True)
sc(ws_pl, PL["growth"], 5, "—", align="center", italic=True)

# ─ 売上原価
sc(ws_pl, 11, 1, "【売上原価（変動費）】", bold=True, fc="DAEEF3")
ws_pl.merge_cells("A11:E11")

PL["worker"] = 12
PL["fde_out"]= 13
PL["cogs"]   = 14

rows_pl_cogs = [
    (PL["worker"],  "  副業ワーカー報酬",       "=前提条件!B26", "=前提条件!C26", "=前提条件!D26"),
    (PL["fde_out"], "  FDE外注費",               "=0",            "=200",          "=600"),
]
for row, label, f1, f2, f3 in rows_pl_cogs:
    sc(ws_pl, row, 1, label)
    sc(ws_pl, row, 2, formula=f1, nf=NF_MAN)
    sc(ws_pl, row, 3, formula=f2, nf=NF_MAN)
    sc(ws_pl, row, 4, formula=f3, nf=NF_MAN)
    sc(ws_pl, row, 5, formula=f"=SUM(B{row}:D{row})", nf=NF_MAN, fc=C_SUB)

sc(ws_pl, PL["cogs"], 1, "売上原価合計", bold=True, fc=C_SUB)
for c, col in [(2,"B"),(3,"C"),(4,"D")]:
    sc(ws_pl, PL["cogs"], c,
       formula=f"=SUM({col}{PL['worker']}:{col}{PL['fde_out']})",
       bold=True, fc=C_SUB, nf=NF_MAN)
sc(ws_pl, PL["cogs"], 5,
   formula=f"=SUM(B{PL['cogs']}:D{PL['cogs']})",
   bold=True, fc=C_SUB, nf=NF_MAN)

# ─ 売上総利益
PL["gross"] = 16
PL["gross_r"]= 17
sc(ws_pl, 15, 1, ""); ws_pl.merge_cells("A15:E15")
sc(ws_pl, PL["gross"], 1, "売上総利益（粗利）", bold=True, fc=C_PROFT)
for c, col in [(2,"B"),(3,"C"),(4,"D")]:
    sc(ws_pl, PL["gross"], c,
       formula=f"={col}{PL['uriage']}-{col}{PL['cogs']}",
       bold=True, fc=C_PROFT, nf=NF_MAN)
sc(ws_pl, PL["gross"], 5,
   formula=f"=SUM(B{PL['gross']}:D{PL['gross']})",
   bold=True, fc=C_PROFT, nf=NF_MAN)

sc(ws_pl, PL["gross_r"], 1, "粗利率", italic=True)
for c, col in [(2,"B"),(3,"C"),(4,"D")]:
    sc(ws_pl, PL["gross_r"], c,
       formula=f"={col}{PL['gross']}/{col}{PL['uriage']}",
       nf=NF_PCT, italic=True, align="center")

# ─ 販管費
sc(ws_pl, 19, 1, "【販売費及び一般管理費】", bold=True, fc="DAEEF3")
ws_pl.merge_cells("A19:E19")

PL["fde_sal"] = 20
PL["founder"] = 21
PL["admin"]   = 22
PL["office"]  = 23
PL["mkt"]     = 24
PL["legal"]   = 25
PL["recruit"] = 26
PL["sys"]     = 27
PL["misc"]    = 28
PL["sga"]     = 29
PL["sga_r"]   = 30

rows_sga = [
    (PL["fde_sal"], "  FDE人件費",           "=前提条件!B21", "=前提条件!C21", "=前提条件!D21"),
    (PL["founder"], "  創業者報酬",           "=前提条件!B22", "=前提条件!C22", "=前提条件!D22"),
    (PL["admin"],   "  管理・営業スタッフ費", "=前提条件!B23", "=前提条件!C23", "=前提条件!D23"),
    (PL["office"],  "  オフィス・インフラ費", "=前提条件!B27", "=前提条件!C27", "=前提条件!D27"),
    (PL["mkt"],     "  マーケティング費",     "=前提条件!B28", "=前提条件!C28", "=前提条件!D28"),
    (PL["legal"],   "  法務・許認可費",       "=前提条件!B29", "=前提条件!C29", "=前提条件!D29"),
    (PL["recruit"], "  採用費",               "=前提条件!B30", "=前提条件!C30", "=前提条件!D30"),
    (PL["sys"],     "  システム開発費",       "=前提条件!B31", "=前提条件!C31", "=前提条件!D31"),
    (PL["misc"],    "  その他",               "=前提条件!B32", "=前提条件!C32", "=前提条件!D32"),
]
for row, label, f1, f2, f3 in rows_sga:
    sc(ws_pl, row, 1, label)
    sc(ws_pl, row, 2, formula=f1, nf=NF_MAN)
    sc(ws_pl, row, 3, formula=f2, nf=NF_MAN)
    sc(ws_pl, row, 4, formula=f3, nf=NF_MAN)
    sc(ws_pl, row, 5, formula=f"=SUM(B{row}:D{row})", nf=NF_MAN, fc=C_SUB)

sc(ws_pl, PL["sga"], 1, "販管費合計", bold=True, fc=C_SUB)
for c, col in [(2,"B"),(3,"C"),(4,"D")]:
    sc(ws_pl, PL["sga"], c,
       formula=f"=SUM({col}{PL['fde_sal']}:{col}{PL['misc']})",
       bold=True, fc=C_SUB, nf=NF_MAN)
sc(ws_pl, PL["sga"], 5,
   formula=f"=SUM(B{PL['sga']}:D{PL['sga']})",
   bold=True, fc=C_SUB, nf=NF_MAN)

sc(ws_pl, PL["sga_r"], 1, "販管費率", italic=True)
for c, col in [(2,"B"),(3,"C"),(4,"D")]:
    sc(ws_pl, PL["sga_r"], c,
       formula=f"={col}{PL['sga']}/{col}{PL['uriage']}",
       nf=NF_PCT, italic=True, align="center")

# ─ 営業利益
PL["ebit"]   = 32
PL["ebit_r"] = 33
PL["da"]     = 34
PL["ebitda"] = 35
PL["ebitda_r"]= 36
PL["tax"]    = 37
PL["net"]    = 38
PL["net_r"]  = 39

sc(ws_pl, 31, 1, ""); ws_pl.merge_cells("A31:E31")

sc(ws_pl, PL["ebit"], 1, "営業利益（EBIT）", bold=True, fc=C_PROFT)
for c, col in [(2,"B"),(3,"C"),(4,"D")]:
    sc(ws_pl, PL["ebit"], c,
       formula=f"={col}{PL['gross']}-{col}{PL['sga']}",
       bold=True, fc=C_PROFT, nf=NF_MAN)
sc(ws_pl, PL["ebit"], 5,
   formula=f"=SUM(B{PL['ebit']}:D{PL['ebit']})",
   bold=True, fc=C_PROFT, nf=NF_MAN)

sc(ws_pl, PL["ebit_r"], 1, "営業利益率", italic=True)
for c, col in [(2,"B"),(3,"C"),(4,"D")]:
    sc(ws_pl, PL["ebit_r"], c,
       formula=f"={col}{PL['ebit']}/{col}{PL['uriage']}",
       nf=NF_PCT, italic=True, align="center")

sc(ws_pl, PL["da"], 1, "  (+) 減価償却費", italic=True)
for c, ref in [(2,"B37"),(3,"C37"),(4,"D37")]:
    # 前提条件シートの減価償却費行（37+1=38行目、ヘッダーずれあり）
    sc(ws_pl, PL["da"], c, formula=f"=前提条件!{ref[0]}38", nf=NF_MAN, italic=True)

sc(ws_pl, PL["ebitda"], 1, "EBITDA", bold=True, fc=C_PROFT)
for c, col in [(2,"B"),(3,"C"),(4,"D")]:
    sc(ws_pl, PL["ebitda"], c,
       formula=f"={col}{PL['ebit']}+{col}{PL['da']}",
       bold=True, fc=C_PROFT, nf=NF_MAN)
sc(ws_pl, PL["ebitda"], 5,
   formula=f"=SUM(B{PL['ebitda']}:D{PL['ebitda']})",
   bold=True, fc=C_PROFT, nf=NF_MAN)

sc(ws_pl, PL["ebitda_r"], 1, "EBITDAマージン", italic=True)
for c, col in [(2,"B"),(3,"C"),(4,"D")]:
    sc(ws_pl, PL["ebitda_r"], c,
       formula=f"={col}{PL['ebitda']}/{col}{PL['uriage']}",
       nf=NF_PCT, italic=True, align="center")

sc(ws_pl, PL["tax"], 1, "  法人税等（実効税率30%）")
for c, col in [(2,"B"),(3,"C"),(4,"D")]:
    sc(ws_pl, PL["tax"], c,
       formula=f"=IF({col}{PL['ebit']}>0,ROUND({col}{PL['ebit']}*前提条件!B37,0),0)",
       nf=NF_MAN)

sc(ws_pl, PL["net"], 1, "当期純利益", bold=True, fc=C_PROFT)
for c, col in [(2,"B"),(3,"C"),(4,"D")]:
    sc(ws_pl, PL["net"], c,
       formula=f"={col}{PL['ebit']}-{col}{PL['tax']}",
       bold=True, fc=C_PROFT, nf=NF_MAN)
sc(ws_pl, PL["net"], 5,
   formula=f"=SUM(B{PL['net']}:D{PL['net']})",
   bold=True, fc=C_PROFT, nf=NF_MAN)

sc(ws_pl, PL["net_r"], 1, "純利益率", italic=True)
for c, col in [(2,"B"),(3,"C"),(4,"D")]:
    sc(ws_pl, PL["net_r"], c,
       formula=f"={col}{PL['net']}/{col}{PL['uriage']}",
       nf=NF_PCT, italic=True, align="center")

border_range(ws_pl, 3, PL["net_r"], 1, 5)
ws_pl.freeze_panes = "B4"


# ══════════════════════════════════════════════════
# SHEET 3: BS（貸借対照表）
# ══════════════════════════════════════════════════
ws_bs = wb.create_sheet("BS（貸借対照表）")
ws_bs.column_dimensions["A"].width = 34
for col in ["B","C","D","E"]:
    ws_bs.column_dimensions[col].width = 16

sc(ws_bs, 1, 1, "UNLID 貸借対照表（B/S）3ヶ年", bold=True, size=14,
   fc=C_DARK, color="FFFFFF", align="center")
ws_bs.merge_cells("A1:E1"); rh(ws_bs, 1, 28)
sc(ws_bs, 2, 1, "単位：万円  |  期末残高  |  PL・CF連動",
   fc="D6E4F0", align="center", size=9, italic=True)
ws_bs.merge_cells("A2:E2")

for c,l in [(1,"項目"),(2,"期首（Year0）"),(3,"Year1末"),(4,"Year2末"),(5,"Year3末")]:
    sc(ws_bs, 3, c, l, bold=True, fc=C_DBLUE, align="center", color="FFFFFF")

# ── 資産の部
sc(ws_bs, 4, 1, "【資産の部】", bold=True, fc="DAEEF3")
ws_bs.merge_cells("A4:E4")

BS = {}
BS["cash"]   = 5   # 現金・預金
BS["ar"]     = 6   # 売掛金
BS["prepaid"]= 7   # 前払費用
BS["cur_tot"]= 8   # 流動資産合計
BS["fa"]     = 9   # 固定資産（純額）
BS["asset_tot"]=10 # 資産合計

# 期首（Year0）：調達資本2,000万を現金とみなす
sc(ws_bs, BS["cash"], 1, "  現金・預金")
sc(ws_bs, BS["cash"], 2, value=2000, nf=NF_MAN)
# Year1末：期首+CF営業+CF財務-CF投資 → CFシートから参照（後で更新）
# 簡易計算：売上×(1-売掛日数/365)と純利益から推計
sc(ws_bs, BS["cash"], 3,
   formula="=B5+CF（キャッシュフロー計算書）!B32", nf=NF_MAN)
sc(ws_bs, BS["cash"], 4,
   formula="=C5+CF（キャッシュフロー計算書）!C32", nf=NF_MAN)
sc(ws_bs, BS["cash"], 5,
   formula="=D5+CF（キャッシュフロー計算書）!D32", nf=NF_MAN)

sc(ws_bs, BS["ar"], 1, "  売掛金")
sc(ws_bs, BS["ar"], 2, value=0, nf=NF_MAN)
# 売掛金 = 売上 × 売掛回転日数/365
sc(ws_bs, BS["ar"], 3,
   formula=f"=ROUND(PL（損益計算書）!B{PL['uriage']}*前提条件!B41/365,0)", nf=NF_MAN)
sc(ws_bs, BS["ar"], 4,
   formula=f"=ROUND(PL（損益計算書）!C{PL['uriage']}*前提条件!C41/365,0)", nf=NF_MAN)
sc(ws_bs, BS["ar"], 5,
   formula=f"=ROUND(PL（損益計算書）!D{PL['uriage']}*前提条件!D41/365,0)", nf=NF_MAN)

sc(ws_bs, BS["prepaid"], 1, "  前払費用・その他流動資産")
for c in [2,3,4,5]:
    sc(ws_bs, BS["prepaid"], c, value=50, nf=NF_MAN)

sc(ws_bs, BS["cur_tot"], 1, "流動資産合計", bold=True, fc=C_SUB)
for c, col in [(2,"B"),(3,"C"),(4,"D"),(5,"E")]:
    sc(ws_bs, BS["cur_tot"], c,
       formula=f"=SUM({col}{BS['cash']}:{col}{BS['prepaid']})",
       bold=True, fc=C_SUB, nf=NF_MAN)

sc(ws_bs, BS["fa"], 1, "  固定資産（純額）")
sc(ws_bs, BS["fa"], 2, value=0, nf=NF_MAN)
# 固定資産 = 前期末 + Capex - 減価償却
sc(ws_bs, BS["fa"], 3,
   formula=f"=B{BS['fa']}+前提条件!B39-前提条件!B38", nf=NF_MAN)
sc(ws_bs, BS["fa"], 4,
   formula=f"=C{BS['fa']}+前提条件!C39-前提条件!C38", nf=NF_MAN)
sc(ws_bs, BS["fa"], 5,
   formula=f"=D{BS['fa']}+前提条件!D39-前提条件!D38", nf=NF_MAN)

sc(ws_bs, BS["asset_tot"], 1, "資産合計", bold=True, fc=C_TOT)
for c, col in [(2,"B"),(3,"C"),(4,"D"),(5,"E")]:
    sc(ws_bs, BS["asset_tot"], c,
       formula=f"={col}{BS['cur_tot']}+{col}{BS['fa']}",
       bold=True, fc=C_TOT, nf=NF_MAN)

# ── 負債の部
sc(ws_bs, 12, 1, "【負債の部】", bold=True, fc="DAEEF3")
ws_bs.merge_cells("A12:E12")

BS["ap"]      = 13  # 買掛金
BS["accrued"] = 14  # 未払費用
BS["cur_liab"]= 15  # 流動負債合計
BS["loan"]    = 16  # 借入金
BS["liab_tot"]= 17  # 負債合計

sc(ws_bs, BS["ap"], 1, "  買掛金")
sc(ws_bs, BS["ap"], 2, value=0, nf=NF_MAN)
sc(ws_bs, BS["ap"], 3,
   formula=f"=ROUND(PL（損益計算書）!B{PL['cogs']}*前提条件!B42/365,0)", nf=NF_MAN)
sc(ws_bs, BS["ap"], 4,
   formula=f"=ROUND(PL（損益計算書）!C{PL['cogs']}*前提条件!C42/365,0)", nf=NF_MAN)
sc(ws_bs, BS["ap"], 5,
   formula=f"=ROUND(PL（損益計算書）!D{PL['cogs']}*前提条件!D42/365,0)", nf=NF_MAN)

sc(ws_bs, BS["accrued"], 1, "  未払費用（人件費等）")
sc(ws_bs, BS["accrued"], 2, value=0, nf=NF_MAN)
sc(ws_bs, BS["accrued"], 3,
   formula=f"=ROUND(前提条件!B24/12,0)", nf=NF_MAN)
sc(ws_bs, BS["accrued"], 4,
   formula=f"=ROUND(前提条件!C24/12,0)", nf=NF_MAN)
sc(ws_bs, BS["accrued"], 5,
   formula=f"=ROUND(前提条件!D24/12,0)", nf=NF_MAN)

sc(ws_bs, BS["cur_liab"], 1, "流動負債合計", bold=True, fc=C_SUB)
for c, col in [(2,"B"),(3,"C"),(4,"D"),(5,"E")]:
    sc(ws_bs, BS["cur_liab"], c,
       formula=f"=SUM({col}{BS['ap']}:{col}{BS['accrued']})",
       bold=True, fc=C_SUB, nf=NF_MAN)

sc(ws_bs, BS["loan"], 1, "  借入金・長期負債")
for c in [2,3,4,5]:
    sc(ws_bs, BS["loan"], c, value=0, nf=NF_MAN)

sc(ws_bs, BS["liab_tot"], 1, "負債合計", bold=True, fc=C_SUB)
for c, col in [(2,"B"),(3,"C"),(4,"D"),(5,"E")]:
    sc(ws_bs, BS["liab_tot"], c,
       formula=f"={col}{BS['cur_liab']}+{col}{BS['loan']}",
       bold=True, fc=C_SUB, nf=NF_MAN)

# ── 純資産の部
sc(ws_bs, 19, 1, "【純資産の部】", bold=True, fc="DAEEF3")
ws_bs.merge_cells("A19:E19")

BS["capital"]   = 20
BS["retained"]  = 21
BS["equity_tot"]= 22
BS["liab_eq"]   = 23
BS["check"]     = 24

sc(ws_bs, BS["capital"], 1, "  資本金・資本剰余金")
sc(ws_bs, BS["capital"], 2, value=2000, nf=NF_MAN)
for c in [3,4,5]:
    sc(ws_bs, BS["capital"], c,
       formula=f"=B{BS['capital']}", nf=NF_MAN)

sc(ws_bs, BS["retained"], 1, "  利益剰余金（累積）")
sc(ws_bs, BS["retained"], 2, value=0, nf=NF_MAN)
sc(ws_bs, BS["retained"], 3,
   formula=f"=B{BS['retained']}+PL（損益計算書）!B{PL['net']}", nf=NF_MAN)
sc(ws_bs, BS["retained"], 4,
   formula=f"=C{BS['retained']}+PL（損益計算書）!C{PL['net']}", nf=NF_MAN)
sc(ws_bs, BS["retained"], 5,
   formula=f"=D{BS['retained']}+PL（損益計算書）!D{PL['net']}", nf=NF_MAN)

sc(ws_bs, BS["equity_tot"], 1, "純資産合計", bold=True, fc=C_PROFT)
for c, col in [(2,"B"),(3,"C"),(4,"D"),(5,"E")]:
    sc(ws_bs, BS["equity_tot"], c,
       formula=f"={col}{BS['capital']}+{col}{BS['retained']}",
       bold=True, fc=C_PROFT, nf=NF_MAN)

sc(ws_bs, BS["liab_eq"], 1, "負債・純資産合計", bold=True, fc=C_TOT)
for c, col in [(2,"B"),(3,"C"),(4,"D"),(5,"E")]:
    sc(ws_bs, BS["liab_eq"], c,
       formula=f"={col}{BS['liab_tot']}+{col}{BS['equity_tot']}",
       bold=True, fc=C_TOT, nf=NF_MAN)

sc(ws_bs, BS["check"], 1, "  ✓ バランスチェック（資産-負債純資産）", italic=True, fc=C_NOTE)
for c, col in [(2,"B"),(3,"C"),(4,"D"),(5,"E")]:
    sc(ws_bs, BS["check"], c,
       formula=f"={col}{BS['asset_tot']}-{col}{BS['liab_eq']}",
       italic=True, fc=C_NOTE, nf=NF_MAN, align="center")

border_range(ws_bs, 3, BS["check"], 1, 5)
ws_bs.freeze_panes = "B4"


# ══════════════════════════════════════════════════
# SHEET 4: CF（キャッシュフロー計算書）
# ══════════════════════════════════════════════════
ws_cf = wb.create_sheet("CF（キャッシュフロー計算書）")
ws_cf.column_dimensions["A"].width = 36
for col in ["B","C","D","E"]:
    ws_cf.column_dimensions[col].width = 16

sc(ws_cf, 1, 1, "UNLID キャッシュフロー計算書（間接法）3ヶ年", bold=True, size=14,
   fc=C_DARK, color="FFFFFF", align="center")
ws_cf.merge_cells("A1:E1"); rh(ws_cf, 1, 28)
sc(ws_cf, 2, 1, "単位：万円  |  PL・BS連動  ※CF合計 = 現金増減に一致",
   fc="D6E4F0", align="center", size=9, italic=True)
ws_cf.merge_cells("A2:E2")

for c,l in [(1,"項目"),(2,"Year1"),(3,"Year2"),(4,"Year3"),(5,"3年累計")]:
    sc(ws_cf, 3, c, l, bold=True, fc=C_DBLUE, align="center", color="FFFFFF")

CF = {}

# ─ 営業活動CF
sc(ws_cf, 4, 1, "【営業活動によるCF】", bold=True, fc="DAEEF3")
ws_cf.merge_cells("A4:E4")

CF["net"]     = 5
CF["da_add"]  = 6
CF["ar_chg"]  = 7
CF["ap_chg"]  = 8
CF["tax_pay"] = 9
CF["op_cf"]   = 10

sc(ws_cf, CF["net"], 1, "  当期純利益")
for c, col in [(2,"B"),(3,"C"),(4,"D")]:
    sc(ws_cf, CF["net"], c,
       formula=f"=PL（損益計算書）!{col}{PL['net']}", nf=NF_MAN)
sc(ws_cf, CF["net"], 5,
   formula=f"=SUM(B{CF['net']}:D{CF['net']})", nf=NF_MAN, fc=C_SUB)

sc(ws_cf, CF["da_add"], 1, "  (+) 減価償却費（非資金費用）")
for c, ref in [(2,"B38"),(3,"C38"),(4,"D38")]:
    sc(ws_cf, CF["da_add"], c, formula=f"=前提条件!{ref}", nf=NF_MAN)
sc(ws_cf, CF["da_add"], 5,
   formula=f"=SUM(B{CF['da_add']}:D{CF['da_add']})", nf=NF_MAN, fc=C_SUB)

sc(ws_cf, CF["ar_chg"], 1, "  (-) 売掛金増加")
sc(ws_cf, CF["ar_chg"], 2,
   formula=f"=-(BS（貸借対照表）!C{BS['ar']}-BS（貸借対照表）!B{BS['ar']})", nf=NF_MAN)
sc(ws_cf, CF["ar_chg"], 3,
   formula=f"=-(BS（貸借対照表）!D{BS['ar']}-BS（貸借対照表）!C{BS['ar']})", nf=NF_MAN)
sc(ws_cf, CF["ar_chg"], 4,
   formula=f"=-(BS（貸借対照表）!E{BS['ar']}-BS（貸借対照表）!D{BS['ar']})", nf=NF_MAN)
sc(ws_cf, CF["ar_chg"], 5,
   formula=f"=SUM(B{CF['ar_chg']}:D{CF['ar_chg']})", nf=NF_MAN, fc=C_SUB)

sc(ws_cf, CF["ap_chg"], 1, "  (+) 買掛金増加")
sc(ws_cf, CF["ap_chg"], 2,
   formula=f"=BS（貸借対照表）!C{BS['ap']}-BS（貸借対照表）!B{BS['ap']}", nf=NF_MAN)
sc(ws_cf, CF["ap_chg"], 3,
   formula=f"=BS（貸借対照表）!D{BS['ap']}-BS（貸借対照表）!C{BS['ap']}", nf=NF_MAN)
sc(ws_cf, CF["ap_chg"], 4,
   formula=f"=BS（貸借対照表）!E{BS['ap']}-BS（貸借対照表）!D{BS['ap']}", nf=NF_MAN)
sc(ws_cf, CF["ap_chg"], 5,
   formula=f"=SUM(B{CF['ap_chg']}:D{CF['ap_chg']})", nf=NF_MAN, fc=C_SUB)

sc(ws_cf, CF["tax_pay"], 1, "  (-) 法人税等支払い")
for c, col in [(2,"B"),(3,"C"),(4,"D")]:
    sc(ws_cf, CF["tax_pay"], c,
       formula=f"=-PL（損益計算書）!{col}{PL['tax']}", nf=NF_MAN)
sc(ws_cf, CF["tax_pay"], 5,
   formula=f"=SUM(B{CF['tax_pay']}:D{CF['tax_pay']})", nf=NF_MAN, fc=C_SUB)

sc(ws_cf, CF["op_cf"], 1, "営業活動CF合計", bold=True, fc=C_PROFT)
for c, col in [(2,"B"),(3,"C"),(4,"D")]:
    sc(ws_cf, CF["op_cf"], c,
       formula=f"=SUM({col}{CF['net']}:{col}{CF['tax_pay']})",
       bold=True, fc=C_PROFT, nf=NF_MAN)
sc(ws_cf, CF["op_cf"], 5,
   formula=f"=SUM(B{CF['op_cf']}:D{CF['op_cf']})",
   bold=True, fc=C_PROFT, nf=NF_MAN)

# ─ 投資活動CF
sc(ws_cf, 12, 1, "【投資活動によるCF】", bold=True, fc="DAEEF3")
ws_cf.merge_cells("A12:E12")

CF["capex"]  = 13
CF["inv_cf"] = 14

sc(ws_cf, CF["capex"], 1, "  (-) 設備投資（Capex）")
for c, ref in [(2,"B39"),(3,"C39"),(4,"D39")]:
    sc(ws_cf, CF["capex"], c, formula=f"=-前提条件!{ref}", nf=NF_MAN)
sc(ws_cf, CF["capex"], 5,
   formula=f"=SUM(B{CF['capex']}:D{CF['capex']})", nf=NF_MAN, fc=C_SUB)

sc(ws_cf, CF["inv_cf"], 1, "投資活動CF合計", bold=True, fc=C_LOSS)
for c, col in [(2,"B"),(3,"C"),(4,"D")]:
    sc(ws_cf, CF["inv_cf"], c,
       formula=f"={col}{CF['capex']}",
       bold=True, fc=C_LOSS, nf=NF_MAN)
sc(ws_cf, CF["inv_cf"], 5,
   formula=f"=SUM(B{CF['inv_cf']}:D{CF['inv_cf']})",
   bold=True, fc=C_LOSS, nf=NF_MAN)

# ─ 財務活動CF
sc(ws_cf, 16, 1, "【財務活動によるCF】", bold=True, fc="DAEEF3")
ws_cf.merge_cells("A16:E16")

CF["capital_in"] = 17
CF["fin_cf"]     = 18

sc(ws_cf, CF["capital_in"], 1, "  (+) 増資・資本調達")
for c, ref in [(2,"B40"),(3,"C40"),(4,"D40")]:
    sc(ws_cf, CF["capital_in"], c, formula=f"=前提条件!{ref}", nf=NF_MAN)
sc(ws_cf, CF["capital_in"], 5,
   formula=f"=SUM(B{CF['capital_in']}:D{CF['capital_in']})", nf=NF_MAN, fc=C_SUB)

sc(ws_cf, CF["fin_cf"], 1, "財務活動CF合計", bold=True, fc=C_SUB)
for c, col in [(2,"B"),(3,"C"),(4,"D")]:
    sc(ws_cf, CF["fin_cf"], c,
       formula=f"={col}{CF['capital_in']}",
       bold=True, fc=C_SUB, nf=NF_MAN)
sc(ws_cf, CF["fin_cf"], 5,
   formula=f"=SUM(B{CF['fin_cf']}:D{CF['fin_cf']})",
   bold=True, fc=C_SUB, nf=NF_MAN)

# ─ CF合計・現金残高
CF["total_cf"] = 20
CF["begin_cash"]= 21
CF["end_cash"]  = 22
CF["check"]     = 23

sc(ws_cf, 19, 1, ""); ws_cf.merge_cells("A19:E19")

sc(ws_cf, CF["total_cf"], 1, "当期CF合計（純増減）", bold=True, fc=C_TOT)
for c, col in [(2,"B"),(3,"C"),(4,"D")]:
    sc(ws_cf, CF["total_cf"], c,
       formula=f"={col}{CF['op_cf']}+{col}{CF['inv_cf']}+{col}{CF['fin_cf']}",
       bold=True, fc=C_TOT, nf=NF_MAN)
sc(ws_cf, CF["total_cf"], 5,
   formula=f"=SUM(B{CF['total_cf']}:D{CF['total_cf']})",
   bold=True, fc=C_TOT, nf=NF_MAN)

sc(ws_cf, CF["begin_cash"], 1, "期首現金残高")
sc(ws_cf, CF["begin_cash"], 2, value=2000, nf=NF_MAN)
sc(ws_cf, CF["begin_cash"], 3, formula=f"=B{CF['end_cash']}", nf=NF_MAN)
sc(ws_cf, CF["begin_cash"], 4, formula=f"=C{CF['end_cash']}", nf=NF_MAN)
sc(ws_cf, CF["begin_cash"], 5, formula=f"=B{CF['begin_cash']}", nf=NF_MAN)

CF["end_cash"] = 22
sc(ws_cf, CF["end_cash"], 1, "期末現金残高", bold=True, fc=C_PROFT)
for c, col in [(2,"B"),(3,"C"),(4,"D")]:
    sc(ws_cf, CF["end_cash"], c,
       formula=f"={col}{CF['begin_cash']}+{col}{CF['total_cf']}",
       bold=True, fc=C_PROFT, nf=NF_MAN)
sc(ws_cf, CF["end_cash"], 5,
   formula=f"=D{CF['end_cash']}",
   bold=True, fc=C_PROFT, nf=NF_MAN)

# CF32行目を期末現金として参照（BSから）
CF["bs_check"] = 24
sc(ws_cf, CF["bs_check"], 1, "  ✓ BS現金との差異（0が正常）", italic=True, fc=C_NOTE)
for c, col, bs_col in [(2,"B","C"),(3,"C","D"),(4,"D","E")]:
    sc(ws_cf, CF["bs_check"], c,
       formula=f"={col}{CF['end_cash']}-BS（貸借対照表）!{bs_col}{BS['cash']}",
       italic=True, fc=C_NOTE, nf=NF_MAN, align="center")

# CF32 = end_cash row（BSが参照するrow）
# BSのキャッシュ式は CF!B32 を参照 → 行22がend_cash
# 参照がずれているため修正：BSでは CF!B22 等を参照させる
# → BS sheet の cash formula を修正
ws_bs.cell(row=BS["cash"], column=3).value = f"=B{BS['cash']}+CF（キャッシュフロー計算書）!B{CF['total_cf']}"
ws_bs.cell(row=BS["cash"], column=4).value = f"=C{BS['cash']}+CF（キャッシュフロー計算書）!C{CF['total_cf']}"
ws_bs.cell(row=BS["cash"], column=5).value = f"=D{BS['cash']}+CF（キャッシュフロー計算書）!D{CF['total_cf']}"

border_range(ws_cf, 3, CF["bs_check"], 1, 5)
ws_cf.freeze_panes = "B4"


# ══════════════════════════════════════════════════
# SHEET 5: 感度分析
# ══════════════════════════════════════════════════
ws_sa = wb.create_sheet("感度分析")
ws_sa.column_dimensions["A"].width = 28
for col in ["B","C","D","E","F","G","H"]:
    ws_sa.column_dimensions[col].width = 14

sc(ws_sa, 1, 1, "UNLID 感度分析——売上成長率 × 原価率 → 営業利益への影響", bold=True,
   size=13, fc=C_DARK, color="FFFFFF", align="center")
ws_sa.merge_cells("A1:H1"); rh(ws_sa, 1, 28)
sc(ws_sa, 2, 1, "単位：万円  |  Year2を基準にした2軸感度分析（売上成長率 × 原価率）",
   fc="D6E4F0", align="center", size=9, italic=True)
ws_sa.merge_cells("A2:H2")

# ─ 感度分析①: 売上成長率 × 原価率 → Year2営業利益
sc(ws_sa, 4, 1, "【分析①】Year2 営業利益：売上成長率 × 原価率", bold=True,
   fc=C_BLUE, color="FFFFFF", align="center")
ws_sa.merge_cells("A4:H4")

sc(ws_sa, 5, 1, "Year1売上実績（万円）", bold=True)
sc(ws_sa, 5, 2, formula=f"=PL（損益計算書）!B{PL['uriage']}", bold=True, nf=NF_MAN)

sc(ws_sa, 6, 1, "Year2 販管費固定額（万円）", bold=True)
sc(ws_sa, 6, 2, formula=f"=PL（損益計算書）!C{PL['sga']}", bold=True, nf=NF_MAN)

sc(ws_sa, 7, 1, "【見方】縦：売上成長率、横：原価率（売上に対する割合）", italic=True, fc=C_NOTE)
ws_sa.merge_cells("A7:H7")

# ヘッダー行（原価率）
sc(ws_sa, 9, 1, "売上成長率 \\ 原価率", bold=True, fc=C_LBLUE, align="center")
cost_rates = [0.05, 0.08, 0.10, 0.12, 0.15, 0.18, 0.20]
for i, r in enumerate(cost_rates):
    sc(ws_sa, 9, i+2, value=r, bold=True, fc=C_LBLUE,
       nf=NF_PCT, align="center")

# 売上成長率の行
growth_rates = [2.0, 3.0, 4.0, 5.0, 6.0, 7.0, 8.0, 10.0]
for j, g in enumerate(growth_rates):
    row = 10 + j
    sc(ws_sa, row, 1, value=g, nf="0.0x", bold=True,
       fc=C_SUB if j % 2 == 0 else None, align="center")
    for i in range(len(cost_rates)):
        col = i + 2
        cr_cell = get_column_letter(col) + "9"
        # 営業利益 = Year1売上 × 成長率 × (1 - 原価率) - 販管費
        formula = (f"=$B$5*{g}*(1-{ws_sa.cell(row=9,column=col).column_letter}$9)"
                   f"-$B$6")
        sc(ws_sa, row, col, formula=formula, nf=NF_MAN, align="center")

border_range(ws_sa, 9, 9+len(growth_rates), 1, len(cost_rates)+1)

# 注釈
sc(ws_sa, 9+len(growth_rates)+2, 1,
   "※ 原価率 = (ワーカー報酬+外注費) / 売上高。現状モデルは約8〜9%。",
   italic=True, fc=C_NOTE)
ws_sa.merge_cells(f"A{9+len(growth_rates)+2}:H{9+len(growth_rates)+2}")

# ─ 感度分析②: B2B単価 × 契約社数 → Year2売上
sc(ws_sa, 22, 1, "【分析②】Year2 売上：B2B契約社数 × 月額単価", bold=True,
   fc=C_BLUE, color="FFFFFF", align="center")
ws_sa.merge_cells("A22:H22")

sc(ws_sa, 23, 1, "人材紹介フィー固定（Year2）（万円）", bold=True)
sc(ws_sa, 23, 2, formula=f"=PL（損益計算書）!C{PL['jinzai']}", bold=True, nf=NF_MAN)

# ヘッダー（月額単価）
sc(ws_sa, 25, 1, "契約社数 \\ 月額単価（万円）", bold=True, fc=C_LBLUE, align="center")
prices = [13, 15, 17, 18, 20, 22, 25]
for i, p in enumerate(prices):
    sc(ws_sa, 25, i+2, value=p, bold=True, fc=C_LBLUE,
       nf=NF_MAN, align="center")

companies = [20, 30, 40, 50, 60, 70, 80]
for j, co in enumerate(companies):
    row = 26 + j
    sc(ws_sa, row, 1, value=co, nf="#,##0社", bold=True,
       fc=C_SUB if j % 2 == 0 else None, align="center")
    for i, p in enumerate(prices):
        col = i + 2
        formula = f"={co}*{p}*12+$B$23"
        sc(ws_sa, row, col, formula=formula, nf=NF_MAN, align="center")

border_range(ws_sa, 25, 25+len(companies), 1, len(prices)+1)

# ─ 感度分析③: 3シナリオ比較
sc(ws_sa, 36, 1, "【分析③】3シナリオ営業利益比較（Year1〜3）", bold=True,
   fc=C_BLUE, color="FFFFFF", align="center")
ws_sa.merge_cells("A36:H36")

for c,l in [(1,"シナリオ"),(2,"Year1売上"),(3,"Year1営業利益"),
            (4,"Year2売上"),(5,"Year2営業利益"),(6,"Year3売上"),(7,"Year3営業利益")]:
    sc(ws_sa, 37, c, l, bold=True, fc=C_LBLUE, align="center")

scenarios = [
    (38, "楽観",   2220*1.3, -120*0.5, 14400*1.4, 4500*1.8, 47800*1.4, 18500*1.6, C_PROFT),
    (39, "中央値（ベース）",
         f"=PL（損益計算書）!B{PL['uriage']}",
         f"=PL（損益計算書）!B{PL['ebit']}",
         f"=PL（損益計算書）!C{PL['uriage']}",
         f"=PL（損益計算書）!C{PL['ebit']}",
         f"=PL（損益計算書）!D{PL['uriage']}",
         f"=PL（損益計算書）!D{PL['ebit']}",
         C_TOT),
    (40, "保守",   2220*0.7, -120*2,   14400*0.55, 4500*0.3, 47800*0.6, 18500*0.4, C_LOSS),
]
for row, label, v1,v2,v3,v4,v5,v6, fc in scenarios:
    sc(ws_sa, row, 1, label, bold=True, fc=fc)
    for c, v in [(2,v1),(3,v2),(4,v3),(5,v4),(6,v5),(7,v6)]:
        if isinstance(v, str):
            sc(ws_sa, row, c, formula=v, bold=True, fc=fc, nf=NF_MAN, align="center")
        else:
            sc(ws_sa, row, c, value=round(v), bold=True, fc=fc, nf=NF_MAN, align="center")

border_range(ws_sa, 37, 40, 1, 7)

sc(ws_sa, 42, 1,
   "楽観：FDE採用×1.3倍・B2B契約×1.3倍    保守：FDE採用×0.7倍・B2B契約×0.55倍",
   italic=True, fc=C_NOTE)
ws_sa.merge_cells("A42:H42")

ws_sa.freeze_panes = "B9"


# ══════════════════════════════════════════════════
# 保存
# ══════════════════════════════════════════════════
output_path = "/home/user/UNLID/UNLID_財務三表_v1.0.xlsx"
wb.save(output_path)
print(f"✅ 作成完了: {output_path}")
print("シート一覧:")
for s in wb.sheetnames:
    print(f"  ・{s}")
